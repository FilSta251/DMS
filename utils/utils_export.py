# -*- coding: utf-8 -*-
"""
Export Manager pro Motoservis DMS
Exporty do PDF, Excel, CSV s českým formátováním
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict, Union, Optional
import csv
import config

# Import openpyxl (pro Excel)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl není nainstalován. Spusť: pip install openpyxl")

# Import reportlab (pro PDF)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ reportlab není nainstalován. Spusť: pip install reportlab")


class ExportManager:
    """
    Správce exportů do různých formátů
    """

    @staticmethod
    def export_to_csv(data: List[List],
                      filename: str,
                      headers: Optional[List[str]] = None,
                      directory: Optional[Path] = None) -> Union[Path, None]:
        """
        Export dat do CSV

        Args:
            data: Data jako list listů [[r1c1, r1c2], [r2c1, r2c2]]
            filename: Název souboru (bez přípony)
            headers: Seznam hlaviček (volitelné)
            directory: Cílová složka (výchozí: config.EXPORTS_DIR)

        Returns:
            Path: Cesta k souboru nebo None při chybě

        Example:
            >>> data = [['Jan', 'Novák', 123], ['Petr', 'Svoboda', 456]]
            >>> headers = ['Jméno', 'Příjmení', 'ID']
            >>> ExportManager.export_to_csv(data, 'zakaznici', headers)
            Path('data/exports/zakaznici_20251115_143045.csv')
        """
        try:
            # Příprava složky
            if directory is None:
                directory = config.EXPORTS_DIR
            directory = Path(directory)
            directory.mkdir(parents=True, exist_ok=True)

            # Příprava názvu souboru s časovou známkou
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = directory / f"{filename}_{timestamp}.csv"

            # Zápis do CSV
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')

                # Hlavičky
                if headers:
                    writer.writerow(headers)

                # Data
                for row in data:
                    # Převod na stringy
                    row_str = [str(cell) if cell is not None else '' for cell in row]
                    writer.writerow(row_str)

            return file_path

        except Exception as e:
            print(f"❌ Chyba při exportu do CSV: {e}")
            return None

    @staticmethod
    def export_to_excel(data: List[List],
                       filename: str,
                       headers: Optional[List[str]] = None,
                       sheet_name: str = "Data",
                       directory: Optional[Path] = None,
                       auto_width: bool = True) -> Union[Path, None]:
        """
        Export dat do Excel (.xlsx)

        Args:
            data: Data jako list listů
            filename: Název souboru (bez přípony)
            headers: Seznam hlaviček
            sheet_name: Název listu
            directory: Cílová složka
            auto_width: True = automatická šířka sloupců

        Returns:
            Path: Cesta k souboru nebo None

        Example:
            >>> data = [['Jan', 'Novák', 1234.56], ['Petr', 'Svoboda', 789.00]]
            >>> headers = ['Jméno', 'Příjmení', 'Částka']
            >>> ExportManager.export_to_excel(data, 'zakaznici', headers)
        """
        if not EXCEL_AVAILABLE:
            print("❌ openpyxl není k dispozici")
            return None

        try:
            # Příprava složky
            if directory is None:
                directory = config.EXPORTS_DIR
            directory = Path(directory)
            directory.mkdir(parents=True, exist_ok=True)

            # Příprava názvu souboru
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = directory / f"{filename}_{timestamp}.xlsx"

            # Vytvoření workbooku
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Styly
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Hlavičky
            if headers:
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_thin

                start_row = 2
            else:
                start_row = 1

            # Data
            for row_num, row_data in enumerate(data, start_row):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border_thin

                    # Zarovnání čísel doprava
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")

            # Automatická šířka sloupců
            if auto_width:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter

                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

            # Uložení
            wb.save(file_path)

            return file_path

        except Exception as e:
            print(f"❌ Chyba při exportu do Excel: {e}")
            return None

    @staticmethod
    def export_to_excel_multi_sheet(sheets: Dict[str, Dict],
                                    filename: str,
                                    directory: Optional[Path] = None) -> Union[Path, None]:
        """
        Export více listů do jednoho Excel souboru

        Args:
            sheets: Dict s názvy listů a daty
                {
                    'List1': {'data': [[...]], 'headers': [...]},
                    'List2': {'data': [[...]], 'headers': [...]}
                }
            filename: Název souboru
            directory: Cílová složka

        Returns:
            Path: Cesta k souboru

        Example:
            >>> sheets = {
            ...     'Zákazníci': {
            ...         'data': [['Jan', 'Novák'], ['Petr', 'Svoboda']],
            ...         'headers': ['Jméno', 'Příjmení']
            ...     },
            ...     'Vozidla': {
            ...         'data': [['1A2 3456', 'Honda'], ['1B2 7890', 'Yamaha']],
            ...         'headers': ['SPZ', 'Značka']
            ...     }
            ... }
            >>> ExportManager.export_to_excel_multi_sheet(sheets, 'export')
        """
        if not EXCEL_AVAILABLE:
            print("❌ openpyxl není k dispozici")
            return None

        try:
            # Příprava složky
            if directory is None:
                directory = config.EXPORTS_DIR
            directory = Path(directory)
            directory.mkdir(parents=True, exist_ok=True)

            # Příprava názvu
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = directory / f"{filename}_{timestamp}.xlsx"

            # Vytvoření workbooku
            wb = Workbook()
            wb.remove(wb.active)  # Odstranění defaultního listu

            # Vytvoření každého listu
            for sheet_name, sheet_data in sheets.items():
                ws = wb.create_sheet(title=sheet_name)

                data = sheet_data.get('data', [])
                headers = sheet_data.get('headers', None)

                # Styly
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                border_thin = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Hlavičky
                if headers:
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border_thin

                    start_row = 2
                else:
                    start_row = 1

                # Data
                for row_num, row_data in enumerate(data, start_row):
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = value
                        cell.border = border_thin

                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")

                # Auto šířka
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter

                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

            # Uložení
            wb.save(file_path)

            return file_path

        except Exception as e:
            print(f"❌ Chyba při exportu do Excel (multi-sheet): {e}")
            return None

    @staticmethod
    def export_to_pdf_simple(data: List[List],
                            filename: str,
                            headers: Optional[List[str]] = None,
                            title: Optional[str] = None,
                            directory: Optional[Path] = None) -> Union[Path, None]:
        """
        Export jednoduchého PDF s tabulkou

        Args:
            data: Data jako list listů
            filename: Název souboru
            headers: Hlavičky
            title: Titulek dokumentu
            directory: Cílová složka

        Returns:
            Path: Cesta k souboru

        Example:
            >>> data = [['Jan Novák', 1234.56], ['Petr Svoboda', 789.00]]
            >>> headers = ['Zákazník', 'Částka']
            >>> ExportManager.export_to_pdf_simple(data, 'report', headers, 'Finanční přehled')
        """
        if not PDF_AVAILABLE:
            print("❌ reportlab není k dispozici")
            return None

        try:
            # Příprava složky
            if directory is None:
                directory = config.EXPORTS_DIR
            directory = Path(directory)
            directory.mkdir(parents=True, exist_ok=True)

            # Příprava názvu
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = directory / f"{filename}_{timestamp}.pdf"

            # Vytvoření PDF
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            # Kontejner pro elementy
            elements = []

            # Styly
            styles = getSampleStyleSheet()

            # Titulek
            if title:
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor=colors.HexColor("#2c3e50"),
                    spaceAfter=20,
                    alignment=TA_CENTER
                )
                elements.append(Paragraph(title, title_style))
                elements.append(Spacer(1, 0.5*cm))

            # Příprava dat pro tabulku
            table_data = []

            # Hlavičky
            if headers:
                table_data.append(headers)

            # Data
            for row in data:
                table_data.append([str(cell) if cell is not None else '' for cell in row])

            # Vytvoření tabulky
            table = Table(table_data)

            # Styl tabulky
            table_style = TableStyle([
                # Hlavička
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                # Data
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

                # Mřížka
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

                # Střídavé řádky
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")])
            ])

            table.setStyle(table_style)
            elements.append(table)

            # Datum vytvoření
            elements.append(Spacer(1, 1*cm))
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_RIGHT
            )
            date_text = f"Vytvořeno: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            elements.append(Paragraph(date_text, date_style))

            # Build PDF
            doc.build(elements)

            return file_path

        except Exception as e:
            print(f"❌ Chyba při exportu do PDF: {e}")
            return None

    @staticmethod
    def generate_filename(prefix: str, extension: str) -> str:
        """
        Generování názvu souboru s časovou známkou

        Args:
            prefix: Prefix (např. "faktura", "report")
            extension: Přípona (např. "pdf", "xlsx")

        Returns:
            str: "faktura_20251115_143045.pdf"

        Example:
            >>> ExportManager.generate_filename("faktura", "pdf")
            'faktura_20251115_143045.pdf'
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"


# ============================================
# ZKRATKY
# ============================================

def export_csv(data: List[List], filename: str, headers: Optional[List[str]] = None) -> Union[Path, None]:
    """Zkratka: export do CSV"""
    return ExportManager.export_to_csv(data, filename, headers)


def export_excel(data: List[List], filename: str, headers: Optional[List[str]] = None) -> Union[Path, None]:
    """Zkratka: export do Excel"""
    return ExportManager.export_to_excel(data, filename, headers)


def export_pdf(data: List[List], filename: str, headers: Optional[List[str]] = None, title: Optional[str] = None) -> Union[Path, None]:
    """Zkratka: export do PDF"""
    return ExportManager.export_to_pdf_simple(data, filename, headers, title)


# ============================================
# TESTY
# ============================================

if __name__ == "__main__":
    print("=== TEST EXPORT MANAGERU ===\n")

    # Testovací data
    headers = ['Jméno', 'Příjmení', 'Částka (Kč)', 'Datum']
    data = [
        ['Jan', 'Novák', 1234.56, '15.11.2025'],
        ['Petr', 'Svoboda', 789.00, '14.11.2025'],
        ['Marie', 'Nováková', 2345.67, '13.11.2025']
    ]

    # Test 1: CSV
    print("Test 1: Export do CSV")
    csv_file = ExportManager.export_to_csv(data, 'test_export', headers)
    if csv_file:
        print(f"✅ CSV vytvořen: {csv_file}")
    else:
        print("❌ CSV selhal")

    # Test 2: Excel
    print("\nTest 2: Export do Excel")
    if EXCEL_AVAILABLE:
        excel_file = ExportManager.export_to_excel(data, 'test_export', headers)
        if excel_file:
            print(f"✅ Excel vytvořen: {excel_file}")
        else:
            print("❌ Excel selhal")
    else:
        print("⚠️ openpyxl není nainstalován")

    # Test 3: PDF
    print("\nTest 3: Export do PDF")
    if PDF_AVAILABLE:
        pdf_file = ExportManager.export_to_pdf_simple(
            data, 'test_export', headers, 'Testovací report'
        )
        if pdf_file:
            print(f"✅ PDF vytvořen: {pdf_file}")
        else:
            print("❌ PDF selhal")
    else:
        print("⚠️ reportlab není nainstalován")

    print("\n=== TESTY DOKONČENY ===")
