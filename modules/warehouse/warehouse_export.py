# -*- coding: utf-8 -*-
"""
Export a tisk skladu - PROFESIONÁLNÍ
PDF, Excel, štítky, etikety
"""

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import config
from database_manager import db
from datetime import datetime
import os


class WarehouseExporter:
    """Třída pro exporty skladu"""

    def __init__(self):
        """Inicializace exportéra"""
        self.register_fonts()
        self.styles = getSampleStyleSheet()
        self.create_custom_styles()

    def register_fonts(self):
        """Registrace fontů"""
        try:
            # Pokus o registraci DejaVu fontů (pro české znaky)
            font_path = "C:/Windows/Fonts/"
            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path + 'DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_path + 'DejaVuSans-Bold.ttf'))
            self.font_name = 'DejaVuSans'
            self.font_bold = 'DejaVuSans-Bold'
        except:
            # Fallback na standardní fonty
            self.font_name = 'Helvetica'
            self.font_bold = 'Helvetica-Bold'

    def create_custom_styles(self):
        """Vytvoření vlastních stylů"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor(config.COLOR_PRIMARY),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=self.font_bold
        ))

        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(config.COLOR_SECONDARY),
            spaceAfter=10,
            fontName=self.font_bold
        ))

    # ========================================
    # PDF EXPORTY
    # ========================================

    def export_price_list_pdf(self, file_path, category_id=None):
        """Export ceníku do PDF"""
        try:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []

            # Hlavička
            title = Paragraph("CENÍK NÁHRADNÍCH DÍLŮ", self.styles['CustomTitle'])
            elements.append(title)

            subtitle = Paragraph(
                f"Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                self.styles['Normal']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Načtení dat
            query = """
                SELECT w.name, w.code, w.unit, w.price_sale, c.name as category
                FROM warehouse w
                LEFT JOIN warehouse_categories c ON w.category_id = c.id
                WHERE 1=1
            """
            params = []

            if category_id:
                query += " AND w.category_id = ?"
                params.append(category_id)

            query += " ORDER BY c.name, w.name"

            items = db.execute_query(query, params)

            if not items:
                elements.append(Paragraph("Žádné položky k zobrazení", self.styles['Normal']))
            else:
                # Seskupení podle kategorií
                current_category = None
                table_data = []

                for item in items:
                    category = item[4] or "Bez kategorie"

                    # Nová kategorie
                    if category != current_category:
                        if table_data:
                            # Vytvoření tabulky pro předchozí kategorii
                            self._add_price_table(elements, table_data, current_category)
                            table_data = []
                            elements.append(Spacer(1, 15))

                        current_category = category

                    # Přidání položky
                    table_data.append([
                        item[0],  # Název
                        item[1] or "",  # Kód
                        item[2],  # Jednotka
                        f"{item[3]:.2f} Kč"  # Cena
                    ])

                # Poslední kategorie
                if table_data:
                    self._add_price_table(elements, table_data, current_category)

            # Generování PDF
            doc.build(elements)
            return True

        except Exception as e:
            print(f"Chyba při exportu ceníku: {e}")
            return False

    def _add_price_table(self, elements, data, category_name):
        """Přidání tabulky s cenami"""
        # Hlavička kategorie
        cat_heading = Paragraph(f"<b>{category_name}</b>", self.styles['CustomHeading'])
        elements.append(cat_heading)

        # Hlavička tabulky
        table_data = [['Název', 'Kód', 'Jedn.', 'Cena']] + data

        # Vytvoření tabulky
        table = Table(table_data, colWidths=[10*cm, 3*cm, 2*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(config.COLOR_PRIMARY)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), self.font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)

    def export_inventory_list_pdf(self, file_path):
        """Export inventurního seznamu do PDF"""
        try:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []

            # Hlavička
            title = Paragraph("INVENTURNÍ SEZNAM", self.styles['CustomTitle'])
            elements.append(title)

            subtitle = Paragraph(
                f"Datum: {datetime.now().strftime('%d.%m.%Y')}",
                self.styles['Normal']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Načtení dat
            items = db.execute_query("""
                SELECT w.name, w.code, w.quantity, w.unit, w.location, c.name as category
                FROM warehouse w
                LEFT JOIN warehouse_categories c ON w.category_id = c.id
                ORDER BY c.name, w.name
            """)

            if not items:
                elements.append(Paragraph("Sklad je prázdný", self.styles['Normal']))
            else:
                # Tabulka
                table_data = [['Název', 'Kód', 'Množství', 'Jednotka', 'Umístění', 'Skutečnost']]

                for item in items:
                    table_data.append([
                        item[0],  # Název
                        item[1] or "",  # Kód
                        f"{item[2]:.2f}",  # Množství
                        item[3],  # Jednotka
                        item[4] or "",  # Umístění
                        ""  # Prázdné pro vyplnění při inventuře
                    ])

                # Vytvoření tabulky
                table = Table(table_data, colWidths=[6*cm, 2.5*cm, 2*cm, 2*cm, 3*cm, 2.5*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(config.COLOR_PRIMARY)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), self.font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), self.font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))

                elements.append(table)

            # Generování PDF
            doc.build(elements)
            return True

        except Exception as e:
            print(f"Chyba při exportu inventury: {e}")
            return False

    def export_below_minimum_pdf(self, file_path):
        """Export položek pod minimem do PDF"""
        try:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            elements = []

            # Hlavička
            title = Paragraph("POLOŽKY POD MINIMÁLNÍM STAVEM", self.styles['CustomTitle'])
            elements.append(title)

            subtitle = Paragraph(
                f"Datum: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                self.styles['Normal']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Načtení dat
            items = db.execute_query("""
                SELECT w.name, w.code, w.quantity, w.min_quantity, w.unit,
                       s.name as supplier, s.phone
                FROM warehouse w
                LEFT JOIN warehouse_suppliers s ON w.supplier_id = s.id
                WHERE w.quantity < w.min_quantity
                ORDER BY (w.min_quantity - w.quantity) DESC
            """)

            if not items:
                elements.append(Paragraph(
                    "✓ Všechny položky jsou nad minimálním stavem",
                    self.styles['Normal']
                ))
            else:
                # Tabulka
                table_data = [['Položka', 'Kód', 'Stav', 'Min.', 'Chybí', 'Dodavatel']]

                for item in items:
                    shortage = item[3] - item[2]

                    table_data.append([
                        item[0],  # Název
                        item[1] or "",  # Kód
                        f"{item[2]:.2f} {item[4]}",  # Aktuální stav
                        f"{item[3]:.2f}",  # Minimum
                        f"{shortage:.2f}",  # Chybí
                        item[5] or "---"  # Dodavatel
                    ])

                # Vytvoření tabulky
                table = Table(table_data, colWidths=[6*cm, 2*cm, 2.5*cm, 2*cm, 2*cm, 3.5*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(config.COLOR_DANGER)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (2, 1), (4, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), self.font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), self.font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))

                elements.append(table)

                # Poznámka
                elements.append(Spacer(1, 20))
                note = Paragraph(
                    f"<b>Celkem položek pod minimem: {len(items)}</b>",
                    self.styles['Normal']
                )
                elements.append(note)

            # Generování PDF
            doc.build(elements)
            return True

        except Exception as e:
            print(f"Chyba při exportu: {e}")
            return False

    # ========================================
    # EXCEL EXPORTY
    # ========================================

    def export_full_warehouse_excel(self, file_path):
        """Export kompletního skladu do Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Sklad"

            # Hlavička
            headers = [
                'Název', 'Kód', 'EAN', 'Kategorie', 'Množství', 'Jednotka',
                'Min. stav', 'Umístění', 'Nákupní cena', 'Prodejní cena',
                'Marže %', 'Hodnota', 'Dodavatel'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            items = db.execute_query("""
                SELECT w.name, w.code, w.ean, c.name as category, w.quantity, w.unit,
                       w.min_quantity, w.location, w.price_purchase, w.price_sale,
                       s.name as supplier
                FROM warehouse w
                LEFT JOIN warehouse_categories c ON w.category_id = c.id
                LEFT JOIN warehouse_suppliers s ON w.supplier_id = s.id
                ORDER BY w.name
            """)

            if items:
                for row_num, item in enumerate(items, 2):
                    ws.cell(row=row_num, column=1, value=item[0])
                    ws.cell(row=row_num, column=2, value=item[1] or "")
                    ws.cell(row=row_num, column=3, value=item[2] or "")
                    ws.cell(row=row_num, column=4, value=item[3] or "")
                    ws.cell(row=row_num, column=5, value=item[4])
                    ws.cell(row=row_num, column=6, value=item[5])
                    ws.cell(row=row_num, column=7, value=item[6])
                    ws.cell(row=row_num, column=8, value=item[7] or "")
                    ws.cell(row=row_num, column=9, value=item[8])
                    ws.cell(row=row_num, column=10, value=item[9])

                    # Marže
                    if item[8] and item[8] > 0:
                        margin = ((item[9] - item[8]) / item[8]) * 100
                        ws.cell(row=row_num, column=11, value=round(margin, 2))
                    else:
                        ws.cell(row=row_num, column=11, value=0)

                    # Hodnota
                    value = item[4] * item[8]
                    ws.cell(row=row_num, column=12, value=round(value, 2))

                    ws.cell(row=row_num, column=13, value=item[10] or "")

                    # Barevné zvýraznění podle stavu
                    if item[4] < item[6]:
                        # Pod minimem - červená
                        fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
                        ws.cell(row=row_num, column=5).fill = fill

            # Auto-šířka sloupců
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Chyba při exportu: {e}")
            return False

    def export_movements_excel(self, file_path, date_from=None, date_to=None):
        """Export pohybů do Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Pohyby"

            # Hlavička
            headers = ['Datum', 'Typ', 'Položka', 'Množství', 'Jedn.', 'Cena/jedn.', 'Celkem', 'Dodavatel', 'Poznámka']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            query = """
                SELECT wm.date, wm.movement_type, w.name, wm.quantity, w.unit,
                       wm.unit_price, s.name as supplier, wm.note
                FROM warehouse_movements wm
                LEFT JOIN warehouse w ON wm.item_id = w.id
                LEFT JOIN warehouse_suppliers s ON wm.supplier_id = s.id
                WHERE 1=1
            """
            params = []

            if date_from:
                query += " AND wm.date >= ?"
                params.append(date_from)

            if date_to:
                query += " AND wm.date <= ?"
                params.append(date_to)

            query += " ORDER BY wm.date DESC"

            movements = db.execute_query(query, params)

            if movements:
                for row_num, mov in enumerate(movements, 2):
                    ws.cell(row=row_num, column=1, value=mov[0])
                    ws.cell(row=row_num, column=2, value=mov[1])
                    ws.cell(row=row_num, column=3, value=mov[2])
                    ws.cell(row=row_num, column=4, value=mov[3])
                    ws.cell(row=row_num, column=5, value=mov[4])
                    ws.cell(row=row_num, column=6, value=mov[5])
                    ws.cell(row=row_num, column=7, value=mov[3] * mov[5])
                    ws.cell(row=row_num, column=8, value=mov[6] or "")
                    ws.cell(row=row_num, column=9, value=mov[7] or "")

            # Auto-šířka
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Chyba: {e}")
            return False

    def export_abc_analysis_excel(self, file_path):
        """Export ABC analýzy do Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "ABC Analýza"

            # Hlavička
            headers = ['Položka', 'Celkem vydáno', 'Hodnota', '% z celku', 'Kumulativní %', 'Kategorie']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Načtení dat - položky podle celkové vydané hodnoty
            movements = db.execute_query("""
                SELECT w.name,
                       SUM(wm.quantity) as total_quantity,
                       SUM(wm.quantity * wm.unit_price) as total_value
                FROM warehouse_movements wm
                LEFT JOIN warehouse w ON wm.item_id = w.id
                WHERE wm.movement_type = 'Výdej'
                GROUP BY w.id, w.name
                HAVING total_value > 0
                ORDER BY total_value DESC
            """)

            if movements:
                # Výpočet celkové hodnoty
                total_value = sum(m[2] for m in movements)

                cumulative = 0
                row_num = 2

                for mov in movements:
                    name = mov[0]
                    quantity = mov[1]
                    value = mov[2]
                    percent = (value / total_value * 100) if total_value > 0 else 0
                    cumulative += percent

                    # Určení ABC kategorie
                    if cumulative <= 80:
                        category = "A"
                        fill_color = "c8e6c9"  # Světle zelená
                    elif cumulative <= 95:
                        category = "B"
                        fill_color = "fff9c4"  # Světle žlutá
                    else:
                        category = "C"
                        fill_color = "ffccbc"  # Světle oranžová

                    ws.cell(row=row_num, column=1, value=name)
                    ws.cell(row=row_num, column=2, value=round(quantity, 2))
                    ws.cell(row=row_num, column=3, value=round(value, 2))
                    ws.cell(row=row_num, column=4, value=round(percent, 2))
                    ws.cell(row=row_num, column=5, value=round(cumulative, 2))
                    ws.cell(row=row_num, column=6, value=category)

                    # Barevné zvýraznění kategorie
                    ws.cell(row=row_num, column=6).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
                    ws.cell(row=row_num, column=6).font = Font(bold=True)

                    row_num += 1

            # Auto-šířka
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"Chyba: {e}")
            return False

    # ========================================
    # ŠTÍTKY A ETIKETY
    # ========================================

    def generate_barcode_label(self, file_path, item_data):
        """Generování štítku s čárovým kódem"""
        try:
            import barcode
            from barcode.writer import ImageWriter

            # Vytvoření čárového kódu (EAN13 nebo Code128)
            code = item_data.get('ean') or item_data.get('code') or str(item_data['id']).zfill(12)

            # Pro EAN musí být přesně 12 číslic (13. je kontrolní)
            if len(code) == 13 and code.isdigit():
                barcode_class = barcode.get_barcode_class('ean13')
            else:
                barcode_class = barcode.get_barcode_class('code128')

            # Generování
            barcode_instance = barcode_class(code, writer=ImageWriter())

            options = {
                'module_width': 0.3,
                'module_height': 10.0,
                'quiet_zone': 6.5,
                'font_size': 8,
                'text_distance': 3.0,
                'background': 'white',
                'foreground': 'black',
            }

            # Uložení
            barcode_instance.save(file_path.replace('.png', ''), options)

            return True

        except Exception as e:
            print(f"Chyba při generování čárového kódu: {e}")
            return False

    def generate_price_label_pdf(self, file_path, item_data):
        """Generování cenové etikety do PDF"""
        try:
            # Velikost etikety (cca 5x3 cm)
            c = canvas.Canvas(file_path, pagesize=(5*cm, 3*cm))

            # Název
            c.setFont(self.font_bold, 10)
            c.drawString(0.3*cm, 2.3*cm, item_data['name'][:25])

            # Kód
            c.setFont(self.font_name, 8)
            code_text = f"Kód: {item_data.get('code') or '---'}"
            c.drawString(0.3*cm, 1.9*cm, code_text)

            # Cena - VELKÁ
            c.setFont(self.font_bold, 16)
            price = item_data.get('price_sale', 0)
            c.drawString(0.3*cm, 1.2*cm, f"{price:.2f} Kč")

            # DPH info
            c.setFont(self.font_name, 7)
            c.drawString(0.3*cm, 0.3*cm, "Cena včetně DPH")

            c.save()
            return True

        except Exception as e:
            print(f"Chyba: {e}")
            return False


# Globální instance exportéra
exporter = WarehouseExporter()
