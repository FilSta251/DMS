# -*- coding: utf-8 -*-
"""
Správa skladových pohybů - PROFESIONÁLNÍ
Seznam všech transakcí, filtry, storno, export
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLabel, QComboBox,
    QMessageBox, QMenu, QDateEdit, QLineEdit, QFileDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QColor
import config
from database_manager import db
from datetime import datetime, timedelta


class WarehouseMovementsWindow(QMainWindow):
    """Okno pro správu skladových pohybů"""

    movement_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("📊 Historie skladových pohybů")
        self.setMinimumSize(1400, 800)

        self.init_ui()
        self.load_filters()
        self.load_movements()

    def init_ui(self):
        """Inicializace UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # === HORNÍ LIŠTA ===
        self.create_action_bar(main_layout)

        # === FILTRY ===
        self.create_filters(main_layout)

        # === TABULKA POHYBŮ ===
        self.create_table(main_layout)

        # === DOLNÍ LIŠTA SE STATISTIKAMI ===
        self.create_stats_bar(main_layout)

    def create_action_bar(self, parent_layout):
        """Horní lišta s akcemi"""
        action_bar = QWidget()
        action_bar.setFixedHeight(60)
        action_bar.setStyleSheet(f"""
            QWidget {{
                background-color: {config.COLOR_PRIMARY};
                border-bottom: 2px solid #2c3e50;
            }}
        """)
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(15, 10, 15, 10)

        # Nadpis
        title = QLabel("📊 SKLADOVÉ POHYBY")
        title.setStyleSheet("color: white; font-size: 18px; font-weight: bold;")
        action_layout.addWidget(title)

        action_layout.addStretch()

        # === TLAČÍTKA ===

        # Nový příjem
        btn_receive = QPushButton("➕ Nový příjem")
        btn_receive.setStyleSheet(self.get_button_style(config.COLOR_SUCCESS))
        btn_receive.clicked.connect(self.new_receive)
        action_layout.addWidget(btn_receive)

        # Nový výdej
        btn_issue = QPushButton("➖ Nový výdej")
        btn_issue.setStyleSheet(self.get_button_style("#e67e22"))
        btn_issue.clicked.connect(self.new_issue)
        action_layout.addWidget(btn_issue)

        # Inventura
        btn_inventory = QPushButton("📊 Inventura")
        btn_inventory.setStyleSheet(self.get_button_style(config.COLOR_SECONDARY))
        btn_inventory.clicked.connect(self.new_inventory)
        action_layout.addWidget(btn_inventory)

        # Export
        btn_export = QPushButton("📤 Export do Excel")
        btn_export.setStyleSheet(self.get_button_style("#8e44ad"))
        btn_export.clicked.connect(self.export_to_excel)
        action_layout.addWidget(btn_export)

        # Zavřít
        btn_close = QPushButton("✕ Zavřít")
        btn_close.setStyleSheet(self.get_button_style("#7f8c8d"))
        btn_close.clicked.connect(self.close)
        action_layout.addWidget(btn_close)

        parent_layout.addWidget(action_bar)

    def create_filters(self, parent_layout):
        """Filtry"""
        filter_bar = QWidget()
        filter_bar.setStyleSheet("background-color: #ecf0f1; padding: 10px;")
        filter_layout = QVBoxLayout(filter_bar)

        # První řádek filtrů
        row1 = QHBoxLayout()

        # Datum od
        row1.addWidget(QLabel("Datum od:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_from.setFixedWidth(120)
        self.date_from.dateChanged.connect(self.on_filter_changed)
        row1.addWidget(self.date_from)

        # Datum do
        row1.addWidget(QLabel("Datum do:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setFixedWidth(120)
        self.date_to.dateChanged.connect(self.on_filter_changed)
        row1.addWidget(self.date_to)

        # Quick buttons
        btn_today = QPushButton("Dnes")
        btn_today.clicked.connect(lambda: self.set_date_range("today"))
        row1.addWidget(btn_today)

        btn_week = QPushButton("Týden")
        btn_week.clicked.connect(lambda: self.set_date_range("week"))
        row1.addWidget(btn_week)

        btn_month = QPushButton("Měsíc")
        btn_month.clicked.connect(lambda: self.set_date_range("month"))
        row1.addWidget(btn_month)

        btn_all = QPushButton("Vše")
        btn_all.clicked.connect(lambda: self.set_date_range("all"))
        row1.addWidget(btn_all)

        row1.addStretch()

        filter_layout.addLayout(row1)

        # Druhý řádek filtrů
        row2 = QHBoxLayout()

        # Typ pohybu
        row2.addWidget(QLabel("Typ pohybu:"))
        self.combo_movement_type = QComboBox()
        self.combo_movement_type.addItems([
            "Všechny typy",
            "➕ Příjem",
            "➖ Výdej",
            "📊 Inventura",
            "↩️ Storno"
        ])
        self.combo_movement_type.setFixedWidth(150)
        self.combo_movement_type.currentIndexChanged.connect(self.on_filter_changed)
        row2.addWidget(self.combo_movement_type)

        # Položka
        row2.addWidget(QLabel("Položka:"))
        self.combo_item = QComboBox()
        self.combo_item.setFixedWidth(250)
        self.combo_item.currentIndexChanged.connect(self.on_filter_changed)
        row2.addWidget(self.combo_item)

        # Dodavatel (pro příjmy)
        row2.addWidget(QLabel("Dodavatel:"))
        self.combo_supplier = QComboBox()
        self.combo_supplier.setFixedWidth(200)
        self.combo_supplier.currentIndexChanged.connect(self.on_filter_changed)
        row2.addWidget(self.combo_supplier)

        # Vyhledávání v poznámce/dokladu
        row2.addWidget(QLabel("Hledat:"))
        self.input_search = QLineEdit()
        self.input_search.setPlaceholderText("Poznámka, doklad...")
        self.input_search.setFixedWidth(200)
        self.input_search.textChanged.connect(self.on_filter_changed)
        row2.addWidget(self.input_search)

        row2.addStretch()

        # Reset
        btn_reset = QPushButton("↺ Reset filtrů")
        btn_reset.clicked.connect(self.reset_filters)
        row2.addWidget(btn_reset)

        filter_layout.addLayout(row2)

        parent_layout.addWidget(filter_bar)

    def create_table(self, parent_layout):
        """Tabulka pohybů"""
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "Datum", "Čas", "Typ", "Položka", "Množství", "Jedn.",
            "Cena/jedn.", "Celkem", "Dodavatel", "Doklad", "Poznámka", "ID"
        ])

        # Skrytí ID
        self.table.setColumnHidden(11, True)

        # Nastavení
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.horizontalHeader().setStretchLastSection(True)

        # Šířky sloupců
        self.table.setColumnWidth(0, 90)   # Datum
        self.table.setColumnWidth(1, 60)   # Čas
        self.table.setColumnWidth(2, 100)  # Typ
        self.table.setColumnWidth(3, 250)  # Položka
        self.table.setColumnWidth(4, 80)   # Množství
        self.table.setColumnWidth(5, 50)   # Jednotka

        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # Double click pro detail
        self.table.doubleClicked.connect(self.view_movement_detail)

        # Kontextové menu
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        parent_layout.addWidget(self.table)

    def create_stats_bar(self, parent_layout):
        """Dolní lišta se statistikami"""
        stats_bar = QWidget()
        stats_bar.setFixedHeight(45)
        stats_bar.setStyleSheet("background-color: #ecf0f1; border-top: 1px solid #bdc3c7;")
        stats_layout = QHBoxLayout(stats_bar)
        stats_layout.setContentsMargins(15, 5, 15, 5)

        self.lbl_total_movements = QLabel("Celkem pohybů: 0")
        stats_layout.addWidget(self.lbl_total_movements)

        self.lbl_total_received = QLabel("➕ Přijato: 0")
        self.lbl_total_received.setStyleSheet("color: #27ae60; font-weight: bold;")
        stats_layout.addWidget(self.lbl_total_received)

        self.lbl_total_issued = QLabel("➖ Vydáno: 0")
        self.lbl_total_issued.setStyleSheet("color: #e67e22; font-weight: bold;")
        stats_layout.addWidget(self.lbl_total_issued)

        self.lbl_total_value = QLabel("Celková hodnota: 0.00 Kč")
        stats_layout.addWidget(self.lbl_total_value)

        stats_layout.addStretch()

        parent_layout.addWidget(stats_bar)

    def get_button_style(self, color):
        """Styl tlačítek"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """

    def load_filters(self):
        """Načtení filtrů"""
        try:
            # Položky
            self.combo_item.clear()
            self.combo_item.addItem("Všechny položky", None)

            items = db.execute_query(
                "SELECT id, name, code FROM warehouse ORDER BY name"
            )
            if items:
                for item in items:
                    display = f"{item[1]} ({item[2]})" if item[2] else item[1]
                    self.combo_item.addItem(display, item[0])

            # Dodavatelé
            self.combo_supplier.clear()
            self.combo_supplier.addItem("Všichni dodavatelé", None)

            suppliers = db.execute_query(
                "SELECT id, name FROM warehouse_suppliers ORDER BY name"
            )
            if suppliers:
                for sup in suppliers:
                    self.combo_supplier.addItem(sup[1], sup[0])

        except Exception as e:
            print(f"Chyba při načítání filtrů: {e}")

    def load_movements(self):
        """Načtení pohybů"""
        try:
            # Sestavení SQL dotazu
            query = """
                SELECT
                    wm.id, wm.date, wm.movement_type, wm.quantity, wm.unit_price,
                    wm.document_number, wm.note,
                    w.name as item_name, w.unit,
                    ws.name as supplier_name
                FROM warehouse_movements wm
                LEFT JOIN warehouse w ON wm.item_id = w.id
                LEFT JOIN warehouse_suppliers ws ON wm.supplier_id = ws.id
                WHERE 1=1
            """

            params = []

            # Filtr data od
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            query += " AND wm.date >= ?"
            params.append(date_from)

            # Filtr data do
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            query += " AND wm.date <= ?"
            params.append(date_to)

            # Filtr typu pohybu
            type_index = self.combo_movement_type.currentIndex()
            if type_index == 1:  # Příjem
                query += " AND wm.movement_type = 'Příjem'"
            elif type_index == 2:  # Výdej
                query += " AND wm.movement_type = 'Výdej'"
            elif type_index == 3:  # Inventura
                query += " AND wm.movement_type = 'Inventura'"
            elif type_index == 4:  # Storno
                query += " AND wm.movement_type = 'Storno'"

            # Filtr položky
            item_id = self.combo_item.currentData()
            if item_id:
                query += " AND wm.item_id = ?"
                params.append(item_id)

            # Filtr dodavatele
            supplier_id = self.combo_supplier.currentData()
            if supplier_id:
                query += " AND wm.supplier_id = ?"
                params.append(supplier_id)

            # Vyhledávání
            search_text = self.input_search.text().strip()
            if search_text:
                query += " AND (wm.note LIKE ? OR wm.document_number LIKE ?)"
                search_param = f"%{search_text}%"
                params.extend([search_param, search_param])

            query += " ORDER BY wm.date DESC, wm.id DESC LIMIT 1000"

            movements = db.execute_query(query, params)

            # Vyčištění tabulky
            self.table.setRowCount(0)

            if not movements:
                self.update_stats(0, 0, 0, 0)
                return

            # Statistiky
            total_count = len(movements)
            total_received = 0
            total_issued = 0
            total_value = 0

            # Vyplnění tabulky
            for mov in movements:
                row = self.table.rowCount()
                self.table.insertRow(row)

                mov_id = mov[0]
                date = mov[1]
                mov_type = mov[2]
                quantity = mov[3]
                unit_price = mov[4] or 0
                document = mov[5] or ""
                note = mov[6] or ""
                item_name = mov[7] or "---"
                unit = mov[8] or ""
                supplier = mov[9] or "---"

                # Výpočet celkové částky
                total = quantity * unit_price

                # Statistiky
                if mov_type == "Příjem":
                    total_received += quantity
                    total_value += total
                elif mov_type == "Výdej":
                    total_issued += quantity

                # Rozdělení data a času
                date_obj = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                date_str = date_obj.strftime("%d.%m.%Y")
                time_str = date_obj.strftime("%H:%M")

                # Ikona podle typu
                if mov_type == "Příjem":
                    type_text = "➕ Příjem"
                    type_color = QColor(config.COLOR_SUCCESS)
                elif mov_type == "Výdej":
                    type_text = "➖ Výdej"
                    type_color = QColor("#e67e22")
                elif mov_type == "Inventura":
                    type_text = "📊 Inventura"
                    type_color = QColor(config.COLOR_SECONDARY)
                else:
                    type_text = f"↩️ {mov_type}"
                    type_color = QColor("#95a5a6")

                # Vyplnění buněk
                self.table.setItem(row, 0, QTableWidgetItem(date_str))
                self.table.setItem(row, 1, QTableWidgetItem(time_str))

                type_item = QTableWidgetItem(type_text)
                type_item.setForeground(type_color)
                self.table.setItem(row, 2, type_item)

                self.table.setItem(row, 3, QTableWidgetItem(item_name))

                qty_item = QTableWidgetItem(f"{quantity:.2f}")
                qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, 4, qty_item)

                self.table.setItem(row, 5, QTableWidgetItem(unit))

                price_item = QTableWidgetItem(f"{unit_price:.2f} Kč")
                price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, 6, price_item)

                total_item = QTableWidgetItem(f"{total:.2f} Kč")
                total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, 7, total_item)

                self.table.setItem(row, 8, QTableWidgetItem(supplier))
                self.table.setItem(row, 9, QTableWidgetItem(document))
                self.table.setItem(row, 10, QTableWidgetItem(note))
                self.table.setItem(row, 11, QTableWidgetItem(str(mov_id)))

            # Aktualizace statistik
            self.update_stats(total_count, total_received, total_issued, total_value)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při načítání pohybů:\n{str(e)}")

    def update_stats(self, total_count, total_received, total_issued, total_value):
        """Aktualizace statistik"""
        self.lbl_total_movements.setText(f"Celkem pohybů: {total_count}")
        self.lbl_total_received.setText(f"➕ Přijato: {total_received:.2f}")
        self.lbl_total_issued.setText(f"➖ Vydáno: {total_issued:.2f}")
        self.lbl_total_value.setText(f"Celková hodnota: {total_value:,.2f} Kč")

    def on_filter_changed(self):
        """Změna filtru"""
        self.load_movements()

    def set_date_range(self, range_type):
        """Nastavení rozsahu data"""
        today = QDate.currentDate()

        if range_type == "today":
            self.date_from.setDate(today)
            self.date_to.setDate(today)
        elif range_type == "week":
            self.date_from.setDate(today.addDays(-7))
            self.date_to.setDate(today)
        elif range_type == "month":
            self.date_from.setDate(today.addMonths(-1))
            self.date_to.setDate(today)
        elif range_type == "all":
            self.date_from.setDate(QDate(2000, 1, 1))
            self.date_to.setDate(today.addYears(1))

    def reset_filters(self):
        """Reset filtrů"""
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to.setDate(QDate.currentDate())
        self.combo_movement_type.setCurrentIndex(0)
        self.combo_item.setCurrentIndex(0)
        self.combo_supplier.setCurrentIndex(0)
        self.input_search.clear()
        self.load_movements()

    def new_receive(self):
        """Nový příjem"""
        from .warehouse_widgets import ReceiveStockDialog
        dialog = ReceiveStockDialog(parent=self)
        dialog.stock_received.connect(self.load_movements)
        dialog.exec()

    def new_issue(self):
        """Nový výdej"""
        from .warehouse_widgets import IssueStockDialog
        dialog = IssueStockDialog(parent=self)
        dialog.stock_issued.connect(self.load_movements)
        dialog.exec()

    def new_inventory(self):
        """Nová inventura"""
        from .warehouse_widgets import InventoryDialog
        dialog = InventoryDialog(parent=self)
        dialog.inventory_done.connect(self.load_movements)
        dialog.exec()
    def view_movement_detail(self):
        """Zobrazení detailu pohybu"""
        if self.table.currentRow() < 0:
            return

        mov_id = int(self.table.item(self.table.currentRow(), 11).text())

        try:
            # Načtení detailu pohybu
            mov = db.execute_query(
                """SELECT
                    wm.date, wm.movement_type, wm.quantity, wm.unit_price,
                    wm.document_number, wm.note, wm.created_by,
                    w.name, w.code, w.unit,
                    ws.name as supplier_name
                FROM warehouse_movements wm
                LEFT JOIN warehouse w ON wm.item_id = w.id
                LEFT JOIN warehouse_suppliers ws ON wm.supplier_id = ws.id
                WHERE wm.id = ?""",
                [mov_id]
            )

            if not mov:
                return

            m = mov[0]

            detail = f"""
📊 DETAIL POHYBU

Datum: {m[0]}
Typ: {m[1]}

Položka: {m[7]} ({m[8]})
Množství: {m[2]:.2f} {m[9]}
Cena/jedn.: {m[3]:.2f} Kč
Celkem: {m[2] * m[3]:.2f} Kč

Dodavatel: {m[10] or '---'}
Doklad: {m[4] or '---'}

Poznámka:
{m[5] or '---'}

Vytvořil: {m[6] or '---'}
            """

            QMessageBox.information(self, "Detail pohybu", detail)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba:\n{str(e)}")

    def show_context_menu(self, position):
        """Kontextové menu"""
        if self.table.currentRow() < 0:
            return

        menu = QMenu()

        action_detail = menu.addAction("📋 Detail")
        action_detail.triggered.connect(self.view_movement_detail)

        menu.addSeparator()

        action_storno = menu.addAction("↩️ Stornovat pohyb")
        action_storno.triggered.connect(self.storno_movement)

        menu.exec(self.table.viewport().mapToGlobal(position))

    def storno_movement(self):
        """Storno pohybu"""
        if self.table.currentRow() < 0:
            return

        mov_id = int(self.table.item(self.table.currentRow(), 11).text())
        mov_type = self.table.item(self.table.currentRow(), 2).text()
        item_name = self.table.item(self.table.currentRow(), 3).text()

        reply = QMessageBox.question(
            self,
            "Stornovat pohyb?",
            f"Opravdu stornovat pohyb:\n\n{mov_type}\n{item_name}\n\n"
            "Tato akce vrátí množství zpět na sklad.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Načtení pohybu
                mov = db.execute_query(
                    "SELECT item_id, movement_type, quantity FROM warehouse_movements WHERE id = ?",
                    [mov_id]
                )

                if not mov:
                    return

                item_id = mov[0][0]
                original_type = mov[0][1]
                quantity = mov[0][2]

                # Zpětný pohyb na skladě
                if original_type == "Příjem":
                    # Storno příjmu = odečíst ze skladu
                    db.execute_query(
                        "UPDATE warehouse SET quantity = quantity - ? WHERE id = ?",
                        [quantity, item_id]
                    )
                elif original_type == "Výdej":
                    # Storno výdeje = přidat na sklad
                    db.execute_query(
                        "UPDATE warehouse SET quantity = quantity + ? WHERE id = ?",
                        [quantity, item_id]
                    )

                # Označení pohybu jako stornovaný
                db.execute_query(
                    "UPDATE warehouse_movements SET movement_type = 'Storno', note = note || ' [STORNOVÁNO]' WHERE id = ?",
                    [mov_id]
                )

                QMessageBox.information(self, "Úspěch", "Pohyb byl stornován")
                self.movement_changed.emit()
                self.load_movements()

            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Chyba při stornování:\n{str(e)}")

    def export_to_excel(self):
        """Export do Excelu"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            # Dialog pro uložení
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit jako Excel",
                f"pohyby_skladu_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel soubory (*.xlsx)"
            )

            if not file_path:
                return

            # Vytvoření workbooku
            wb = Workbook()
            ws = wb.active
            ws.title = "Skladové pohyby"

            # Hlavička
            headers = [
                "Datum", "Čas", "Typ pohybu", "Položka", "Množství", "Jednotka",
                "Cena/jedn.", "Celkem", "Dodavatel", "Doklad", "Poznámka"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            row_num = 2
            for row in range(self.table.rowCount()):
                for col in range(11):  # Bez ID sloupce
                    value = self.table.item(row, col).text() if self.table.item(row, col) else ""
                    ws.cell(row=row_num, column=col + 1, value=value)
                row_num += 1

            # Auto-šířka sloupců
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Uložení
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Úspěch",
                f"Data byla vyexportována do:\n{file_path}"
            )

            # Otevření souboru
            try:
                import os
                os.startfile(file_path)
            except:
                pass

        except ImportError:
            QMessageBox.warning(
                self,
                "Chybí knihovna",
                "Pro export do Excelu je potřeba nainstalovat:\n\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při exportu:\n{str(e)}")
