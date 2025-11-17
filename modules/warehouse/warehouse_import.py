# -*- coding: utf-8 -*-
"""
Import dat do skladu - PROFESIONÁLNÍ
CSV, Excel, mapování, validace, náhled
"""

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QComboBox, QCheckBox,
    QGroupBox, QFormLayout, QFileDialog, QMessageBox,
    QProgressDialog, QTextEdit
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor
import config
from database_manager import db
import csv
import os


class WarehouseImportDialog(QDialog):
    """Dialog pro import dat do skladu"""

    items_imported = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        self.file_path = None
        self.file_data = []
        self.headers = []
        self.column_mapping = {}

        self.setWindowTitle("📥 Import do skladu")
        self.setModal(True)
        self.setMinimumSize(1000, 700)

        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        layout = QVBoxLayout(self)

        # Hlavička
        header = QLabel("📥 IMPORT POLOŽEK DO SKLADU")
        header.setStyleSheet(f"""
            font-size: 16px;
            font-weight: bold;
            color: {config.COLOR_PRIMARY};
            padding: 15px;
            background-color: #ecf0f1;
            border-radius: 5px;
        """)
        layout.addWidget(header)

        # === KROK 1: Výběr souboru ===
        file_group = QGroupBox("📁 Krok 1: Výběr souboru")
        file_layout = QHBoxLayout(file_group)

        self.lbl_file = QLabel("Žádný soubor nebyl vybrán")
        file_layout.addWidget(self.lbl_file)

        btn_browse = QPushButton("📂 Procházet...")
        btn_browse.clicked.connect(self.browse_file)
        file_layout.addWidget(btn_browse)

        layout.addWidget(file_group)

        # === KROK 2: Mapování sloupců ===
        mapping_group = QGroupBox("🔗 Krok 2: Mapování sloupců")
        self.mapping_layout = QFormLayout(mapping_group)

        self.mapping_combos = {}

        # Definice polí pro mapování
        self.fields = {
            'name': {'label': 'Název *', 'required': True},
            'code': {'label': 'Kód', 'required': False},
            'ean': {'label': 'EAN', 'required': False},
            'category': {'label': 'Kategorie', 'required': False},
            'supplier': {'label': 'Dodavatel', 'required': False},
            'quantity': {'label': 'Množství', 'required': False},
            'unit': {'label': 'Jednotka', 'required': False},
            'min_quantity': {'label': 'Min. množství', 'required': False},
            'location': {'label': 'Umístění', 'required': False},
            'price_purchase': {'label': 'Nákupní cena', 'required': False},
            'price_sale': {'label': 'Prodejní cena', 'required': False},
            'description': {'label': 'Popis', 'required': False}
        }

        for field_id, field_info in self.fields.items():
            combo = QComboBox()
            combo.setMinimumWidth(200)
            self.mapping_combos[field_id] = combo
            self.mapping_layout.addRow(field_info['label'] + ":", combo)

        layout.addWidget(mapping_group)

        # === KROK 3: Náhled dat ===
        preview_group = QGroupBox("👁️ Krok 3: Náhled prvních 5 řádků")
        preview_layout = QVBoxLayout(preview_group)

        self.preview_table = QTableWidget()
        self.preview_table.setMaximumHeight(200)
        self.preview_table.setAlternatingRowColors(True)
        preview_layout.addWidget(self.preview_table)

        layout.addWidget(preview_group)

        # === KROK 4: Nastavení importu ===
        settings_group = QGroupBox("⚙️ Krok 4: Nastavení importu")
        settings_layout = QVBoxLayout(settings_group)

        self.check_update_existing = QCheckBox("Aktualizovat existující položky (podle kódu)")
        self.check_update_existing.setChecked(False)
        settings_layout.addWidget(self.check_update_existing)

        self.check_skip_first_row = QCheckBox("Přeskočit první řádek (hlavička)")
        self.check_skip_first_row.setChecked(True)
        settings_layout.addWidget(self.check_skip_first_row)

        self.check_validate_only = QCheckBox("Pouze validace (bez importu)")
        settings_layout.addWidget(self.check_validate_only)

        layout.addWidget(settings_group)

        # === LOG ===
        log_group = QGroupBox("📋 Log importu")
        log_layout = QVBoxLayout(log_group)

        self.text_log = QTextEdit()
        self.text_log.setMaximumHeight(150)
        self.text_log.setReadOnly(True)
        log_layout.addWidget(self.text_log)

        layout.addWidget(log_group)

        # === TLAČÍTKA ===
        buttons = QHBoxLayout()

        btn_cancel = QPushButton("Zrušit")
        btn_cancel.clicked.connect(self.reject)

        self.btn_import = QPushButton("▶️ Spustit import")
        self.btn_import.setEnabled(False)
        self.btn_import.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 12px 30px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:disabled {{
                background-color: #95a5a6;
            }}
        """)
        self.btn_import.clicked.connect(self.start_import)

        buttons.addStretch()
        buttons.addWidget(btn_cancel)
        buttons.addWidget(self.btn_import)

        layout.addLayout(buttons)

    def browse_file(self):
        """Výběr souboru"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Vyberte soubor k importu",
            "",
            "CSV soubory (*.csv);;Excel soubory (*.xlsx *.xls);;Všechny soubory (*.*)"
        )

        if not file_path:
            return

        self.file_path = file_path
        self.lbl_file.setText(os.path.basename(file_path))

        # Načtení dat
        if self.load_file_data():
            self.setup_column_mapping()
            self.update_preview()
            self.btn_import.setEnabled(True)
            self.log("✓ Soubor byl načten")

    def load_file_data(self):
        """Načtení dat ze souboru"""
        try:
            ext = os.path.splitext(self.file_path)[1].lower()

            if ext == '.csv':
                self.load_csv()
            elif ext in ['.xlsx', '.xls']:
                self.load_excel()
            else:
                QMessageBox.warning(self, "Chyba", "Nepodporovaný formát souboru")
                return False

            return True

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při načítání souboru:\n{str(e)}")
            return False

    def load_csv(self):
        """Načtení CSV souboru"""
        with open(self.file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            self.file_data = list(reader)

        # První řádek jako hlavičky
        if self.file_data:
            self.headers = self.file_data[0]

    def load_excel(self):
        """Načtení Excel souboru"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active

            self.file_data = []
            for row in ws.iter_rows(values_only=True):
                self.file_data.append([str(cell) if cell is not None else "" for cell in row])

            # První řádek jako hlavičky
            if self.file_data:
                self.headers = self.file_data[0]

            wb.close()

        except ImportError:
            QMessageBox.warning(
                self,
                "Chybí knihovna",
                "Pro import z Excelu je potřeba:\n\npip install openpyxl"
            )
            raise

    def setup_column_mapping(self):
        """Nastavení mapování sloupců"""
        # Naplnění combo boxů
        for combo in self.mapping_combos.values():
            combo.clear()
            combo.addItem("-- Nepoužito --", None)

            for i, header in enumerate(self.headers):
                combo.addItem(f"{header} (sloupec {i+1})", i)

        # Automatické mapování podle názvů
        self.auto_map_columns()

    def auto_map_columns(self):
        """Automatické mapování podle názvů sloupců"""
        # Mapování českých názvů
        mapping_hints = {
            'name': ['název', 'nazev', 'name', 'item', 'položka', 'polozka'],
            'code': ['kód', 'kod', 'code', 'katalogové číslo', 'katalogove cislo'],
            'ean': ['ean', 'ean13', 'čárový kód', 'carovy kod', 'barcode'],
            'category': ['kategorie', 'category'],
            'supplier': ['dodavatel', 'supplier', 'vendor'],
            'quantity': ['množství', 'mnozstvi', 'quantity', 'qty', 'stav'],
            'unit': ['jednotka', 'unit', 'jedn', 'mj'],
            'min_quantity': ['minimum', 'min', 'min_quantity'],
            'location': ['umístění', 'umisteni', 'location', 'místo', 'misto'],
            'price_purchase': ['nákupní', 'nakupni', 'purchase', 'nákup', 'nakup'],
            'price_sale': ['prodejní', 'prodejni', 'sale', 'prodej', 'cena'],
            'description': ['popis', 'description', 'poznámka', 'poznamka']
        }

        for field_id, hints in mapping_hints.items():
            combo = self.mapping_combos[field_id]

            for i, header in enumerate(self.headers):
                header_lower = header.lower()

                for hint in hints:
                    if hint in header_lower:
                        combo.setCurrentIndex(i + 1)  # +1 protože první je "Nepoužito"
                        break

    def update_preview(self):
        """Aktualizace náhledu dat"""
        if not self.file_data:
            return

        # Určení řádků k zobrazení
        start_row = 1 if self.check_skip_first_row.isChecked() else 0
        preview_data = self.file_data[start_row:start_row + 5]

        # Nastavení tabulky
        self.preview_table.setRowCount(len(preview_data))
        self.preview_table.setColumnCount(len(self.headers))
        self.preview_table.setHorizontalHeaderLabels(self.headers)

        # Vyplnění dat
        for row_idx, row_data in enumerate(preview_data):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                self.preview_table.setItem(row_idx, col_idx, item)

        # Auto-šířka
        self.preview_table.resizeColumnsToContents()

    def start_import(self):
        """Spuštění importu"""
        # Validace mapování
        if not self.validate_mapping():
            return

        # Potvrzení
        if not self.check_validate_only.isChecked():
            reply = QMessageBox.question(
                self,
                "Potvrdit import",
                f"Připraveno {len(self.file_data) - (1 if self.check_skip_first_row.isChecked() else 0)} řádků k importu.\n\n"
                "Spustit import?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.No:
                return

        # Spuštění
        self.text_log.clear()
        self.perform_import()

    def validate_mapping(self):
        """Validace mapování"""
        # Kontrola povinných polí
        name_col = self.mapping_combos['name'].currentData()

        if name_col is None:
            QMessageBox.warning(
                self,
                "Chybí mapování",
                "Pole 'Název' je povinné!\nMusíte namapovat sloupec s názvem položky."
            )
            return False

        # Kontrola duplicitního mapování
        mapped_columns = []
        for field_id, combo in self.mapping_combos.items():
            col_idx = combo.currentData()
            if col_idx is not None:
                if col_idx in mapped_columns:
                    QMessageBox.warning(
                        self,
                        "Duplicitní mapování",
                        f"Sloupec {self.headers[col_idx]} je namapován vícekrát!"
                    )
                    return False
                mapped_columns.append(col_idx)

        return True

    def perform_import(self):
        """Provedení importu"""
        try:
            # Vytvoření mapování
            self.column_mapping = {}
            for field_id, combo in self.mapping_combos.items():
                col_idx = combo.currentData()
                if col_idx is not None:
                    self.column_mapping[field_id] = col_idx

            # Určení řádků k importu
            start_row = 1 if self.check_skip_first_row.isChecked() else 0
            data_rows = self.file_data[start_row:]

            # Statistiky
            stats = {
                'total': len(data_rows),
                'imported': 0,
                'updated': 0,
                'skipped': 0,
                'errors': 0
            }

            # Progress dialog
            progress = QProgressDialog("Import dat...", "Zrušit", 0, len(data_rows), self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)

            # Načtení kategorií a dodavatelů pro mapování
            categories = self.load_categories()
            suppliers = self.load_suppliers()

            # Import po řádcích
            for row_idx, row_data in enumerate(data_rows):
                progress.setValue(row_idx)

                if progress.wasCanceled():
                    self.log("⚠️ Import byl zrušen")
                    break

                try:
                    result = self.import_row(row_data, categories, suppliers)

                    if result == 'imported':
                        stats['imported'] += 1
                    elif result == 'updated':
                        stats['updated'] += 1
                    elif result == 'skipped':
                        stats['skipped'] += 1

                except Exception as e:
                    stats['errors'] += 1
                    self.log(f"❌ Řádek {row_idx + 1}: {str(e)}")

            progress.setValue(len(data_rows))

            # Výsledky
            self.log("\n" + "="*50)
            self.log("VÝSLEDKY IMPORTU:")
            self.log(f"Celkem řádků: {stats['total']}")
            self.log(f"✓ Importováno: {stats['imported']}")
            self.log(f"↻ Aktualizováno: {stats['updated']}")
            self.log(f"⊘ Přeskočeno: {stats['skipped']}")
            self.log(f"✗ Chyby: {stats['errors']}")

            if not self.check_validate_only.isChecked():
                QMessageBox.information(
                    self,
                    "Import dokončen",
                    f"Import byl dokončen!\n\n"
                    f"Importováno: {stats['imported']}\n"
                    f"Aktualizováno: {stats['updated']}\n"
                    f"Chyby: {stats['errors']}"
                )

                if stats['imported'] > 0 or stats['updated'] > 0:
                    self.items_imported.emit()
                    self.accept()
            else:
                QMessageBox.information(
                    self,
                    "Validace dokončena",
                    f"Validace byla dokončena!\n\n"
                    f"Platných řádků: {stats['imported'] + stats['updated']}\n"
                    f"Chyby: {stats['errors']}"
                )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při importu:\n{str(e)}")

    def import_row(self, row_data, categories, suppliers):
        """Import jednoho řádku"""
        # Extrakce dat
        item_data = {}

        for field_id, col_idx in self.column_mapping.items():
            if col_idx < len(row_data):
                value = row_data[col_idx].strip()

                # Konverze hodnot
                if field_id in ['quantity', 'min_quantity', 'price_purchase', 'price_sale']:
                    try:
                        value = float(value.replace(',', '.')) if value else 0
                    except:
                        value = 0

                item_data[field_id] = value
            else:
                item_data[field_id] = None

        # Validace povinných polí
        if not item_data.get('name'):
            raise ValueError("Chybí název položky")

        # Validace v režimu "pouze validace"
        if self.check_validate_only.isChecked():
            return 'imported'

        # Mapování kategorie
        category_id = None
        if item_data.get('category'):
            category_name = item_data['category'].lower()
            category_id = categories.get(category_name)

        # Mapování dodavatele
        supplier_id = None
        if item_data.get('supplier'):
            supplier_name = item_data['supplier'].lower()
            supplier_id = suppliers.get(supplier_name)

        # Kontrola existence podle kódu
        existing = None
        if item_data.get('code'):
            existing = db.execute_query(
                "SELECT id FROM warehouse WHERE code = ?",
                [item_data['code']]
            )

        if existing and self.check_update_existing.isChecked():
            # Aktualizace existující položky
            db.execute_query(
                """UPDATE warehouse SET
                   name=?, ean=?, category_id=?, supplier_id=?, quantity=?,
                   unit=?, min_quantity=?, location=?, price_purchase=?, price_sale=?,
                   description=?
                   WHERE id=?""",
                [
                    item_data.get('name'),
                    item_data.get('ean'),
                    category_id,
                    supplier_id,
                    item_data.get('quantity', 0),
                    item_data.get('unit', 'ks'),
                    item_data.get('min_quantity', 0),
                    item_data.get('location'),
                    item_data.get('price_purchase', 0),
                    item_data.get('price_sale', 0),
                    item_data.get('description'),
                    existing[0][0]
                ]
            )
            return 'updated'

        elif existing:
            # Existuje, ale neaktualizujeme
            return 'skipped'

        else:
            # Nová položka
            db.execute_query(
                """INSERT INTO warehouse
                   (name, code, ean, category_id, supplier_id, quantity, unit,
                    min_quantity, location, price_purchase, price_sale, description)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                [
                    item_data.get('name'),
                    item_data.get('code'),
                    item_data.get('ean'),
                    category_id,
                    supplier_id,
                    item_data.get('quantity', 0),
                    item_data.get('unit', 'ks'),
                    item_data.get('min_quantity', 0),
                    item_data.get('location'),
                    item_data.get('price_purchase', 0),
                    item_data.get('price_sale', 0),
                    item_data.get('description')
                ]
            )
            return 'imported'

    def load_categories(self):
        """Načtení kategorií pro mapování"""
        categories = {}
        rows = db.execute_query("SELECT id, name FROM warehouse_categories")
        if rows:
            for row in rows:
                categories[row[1].lower()] = row[0]
        return categories

    def load_suppliers(self):
        """Načtení dodavatelů pro mapování"""
        suppliers = {}
        rows = db.execute_query("SELECT id, name FROM warehouse_suppliers")
        if rows:
            for row in rows:
                suppliers[row[1].lower()] = row[0]
        return suppliers

    def log(self, message):
        """Přidání zprávy do logu"""
        self.text_log.append(message)
