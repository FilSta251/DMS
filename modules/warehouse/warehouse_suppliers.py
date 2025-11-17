# -*- coding: utf-8 -*-
"""
Správa dodavatelů - PROFESIONÁLNÍ
CRUD, kontakty, položky, objednávky, export
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLabel, QMessageBox,
    QDialog, QFormLayout, QLineEdit, QTextEdit, QComboBox, QTabWidget,
    QGroupBox, QMenu, QFileDialog
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor
import config
from database_manager import db
from datetime import datetime


class WarehouseSuppliersWindow(QMainWindow):
    """Okno pro správu dodavatelů"""

    suppliers_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("🚚 Dodavatelé")
        self.setMinimumSize(1200, 700)

        self.init_ui()
        self.load_suppliers()

    def init_ui(self):
        """Inicializace UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # === HORNÍ LIŠTA ===
        self.create_action_bar(main_layout)

        # === TABULKA DODAVATELŮ ===
        self.create_table(main_layout)

        # === DOLNÍ LIŠTA ===
        self.create_stats_bar(main_layout)

    def create_action_bar(self, parent_layout):
        """Horní lišta s akcemi"""
        action_bar = QWidget()
        action_bar.setFixedHeight(60)
        action_bar.setStyleSheet(f"""
            QWidget {{
                background-color: {config.COLOR_PRIMARY};
                border-bottom: 2px solid #2c3e50;
            }}
        """)
        action_layout = QHBoxLayout(action_bar)
        action_layout.setContentsMargins(15, 10, 15, 10)

        # Nadpis
        title = QLabel("🚚 DODAVATELÉ")
        title.setStyleSheet("color: white; font-size: 18px; font-weight: bold;")
        action_layout.addWidget(title)

        action_layout.addStretch()

        # === TLAČÍTKA ===

        # Nový dodavatel
        btn_new = QPushButton("➕ Nový dodavatel")
        btn_new.setStyleSheet(self.get_button_style(config.COLOR_SUCCESS))
        btn_new.clicked.connect(self.add_supplier)
        action_layout.addWidget(btn_new)

        # Export
        btn_export = QPushButton("📤 Export")
        btn_export.setStyleSheet(self.get_button_style(config.COLOR_SECONDARY))
        btn_export.clicked.connect(self.export_suppliers)
        action_layout.addWidget(btn_export)

        # Zavřít
        btn_close = QPushButton("✕ Zavřít")
        btn_close.setStyleSheet(self.get_button_style("#7f8c8d"))
        btn_close.clicked.connect(self.close)
        action_layout.addWidget(btn_close)

        parent_layout.addWidget(action_bar)

    def create_table(self, parent_layout):
        """Tabulka dodavatelů"""
        container = QWidget()
        layout = QVBoxLayout(container)

        # Info
        info = QLabel("📋 Seznam dodavatelů")
        info.setStyleSheet("padding: 10px; background-color: #ecf0f1; font-weight: bold;")
        layout.addWidget(info)

        # Tabulka
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Název", "IČO", "Kontakt", "Telefon", "Email",
            "Město", "Platební podmínky", "Položky", "ID"
        ])

        # Skrytí ID
        self.table.setColumnHidden(8, True)

        # Nastavení
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.horizontalHeader().setStretchLastSection(True)

        # Šířky sloupců
        self.table.setColumnWidth(0, 200)  # Název
        self.table.setColumnWidth(1, 100)  # IČO
        self.table.setColumnWidth(2, 150)  # Kontakt
        self.table.setColumnWidth(3, 120)  # Telefon
        self.table.setColumnWidth(4, 180)  # Email

        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # Double click pro detail
        self.table.doubleClicked.connect(self.open_detail)

        # Kontextové menu
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        layout.addWidget(self.table)

        parent_layout.addWidget(container)

    def create_stats_bar(self, parent_layout):
        """Dolní lišta se statistikami"""
        stats_bar = QWidget()
        stats_bar.setFixedHeight(40)
        stats_bar.setStyleSheet("background-color: #ecf0f1; border-top: 1px solid #bdc3c7;")
        stats_layout = QHBoxLayout(stats_bar)
        stats_layout.setContentsMargins(15, 5, 15, 5)

        self.lbl_total_suppliers = QLabel("Celkem dodavatelů: 0")
        stats_layout.addWidget(self.lbl_total_suppliers)

        self.lbl_total_items = QLabel("Celkem položek: 0")
        stats_layout.addWidget(self.lbl_total_items)

        stats_layout.addStretch()

        parent_layout.addWidget(stats_bar)

    def get_button_style(self, color):
        """Styl tlačítek"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """

    def load_suppliers(self):
        """Načtení dodavatelů"""
        try:
            suppliers = db.execute_query(
                """SELECT
                    s.id, s.name, s.ico, s.contact_person, s.phone, s.email,
                    s.city, s.payment_terms,
                    COUNT(w.id) as item_count
                FROM warehouse_suppliers s
                LEFT JOIN warehouse w ON w.supplier_id = s.id
                GROUP BY s.id, s.name, s.ico, s.contact_person, s.phone, s.email, s.city, s.payment_terms
                ORDER BY s.name"""
            )

            self.table.setRowCount(0)

            if not suppliers:
                self.update_stats(0, 0)
                return

            total_items = 0

            for sup in suppliers:
                row = self.table.rowCount()
                self.table.insertRow(row)

                sup_id = sup[0]
                name = sup[1]
                ico = sup[2] or ""
                contact = sup[3] or ""
                phone = sup[4] or ""
                email = sup[5] or ""
                city = sup[6] or ""
                payment = sup[7] or ""
                item_count = sup[8]

                total_items += item_count

                # Vyplnění buněk
                self.table.setItem(row, 0, QTableWidgetItem(name))
                self.table.setItem(row, 1, QTableWidgetItem(ico))
                self.table.setItem(row, 2, QTableWidgetItem(contact))
                self.table.setItem(row, 3, QTableWidgetItem(phone))
                self.table.setItem(row, 4, QTableWidgetItem(email))
                self.table.setItem(row, 5, QTableWidgetItem(city))
                self.table.setItem(row, 6, QTableWidgetItem(payment))

                items_item = QTableWidgetItem(str(item_count))
                items_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 7, items_item)

                self.table.setItem(row, 8, QTableWidgetItem(str(sup_id)))

            # Aktualizace statistik
            self.update_stats(len(suppliers), total_items)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při načítání dodavatelů:\n{str(e)}")

    def update_stats(self, total_suppliers, total_items):
        """Aktualizace statistik"""
        self.lbl_total_suppliers.setText(f"Celkem dodavatelů: {total_suppliers}")
        self.lbl_total_items.setText(f"Celkem položek: {total_items}")

    def add_supplier(self):
        """Přidání dodavatele"""
        dialog = SupplierDetailDialog(parent=self)
        dialog.supplier_saved.connect(self.load_suppliers)
        dialog.supplier_saved.connect(self.suppliers_changed.emit)
        dialog.exec()

    def open_detail(self):
        """Otevření detailu dodavatele"""
        if self.table.currentRow() < 0:
            return

        supplier_id = int(self.table.item(self.table.currentRow(), 8).text())

        dialog = SupplierDetailDialog(supplier_id, self)
        dialog.supplier_saved.connect(self.load_suppliers)
        dialog.supplier_saved.connect(self.suppliers_changed.emit)
        dialog.exec()

    def show_context_menu(self, position):
        """Kontextové menu"""
        if self.table.currentRow() < 0:
            return

        menu = QMenu()

        action_detail = menu.addAction("📋 Detail")
        action_detail.triggered.connect(self.open_detail)

        menu.addSeparator()

        action_delete = menu.addAction("🗑️ Smazat")
        action_delete.triggered.connect(self.delete_supplier)

        menu.exec(self.table.viewport().mapToGlobal(position))

    def delete_supplier(self):
        """Smazání dodavatele"""
        if self.table.currentRow() < 0:
            return

        supplier_id = int(self.table.item(self.table.currentRow(), 8).text())
        supplier_name = self.table.item(self.table.currentRow(), 0).text()
        item_count = int(self.table.item(self.table.currentRow(), 7).text())

        if item_count > 0:
            reply = QMessageBox.question(
                self,
                "Položky dodavatele",
                f"Dodavatel '{supplier_name}' má přiřazeno {item_count} položek.\n\n"
                "Chcete dodavatele přesto smazat?\n(Položky ztratí přiřazení k dodavateli)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.No:
                return

        reply = QMessageBox.question(
            self,
            "Smazat dodavatele?",
            f"Opravdu smazat dodavatele '{supplier_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Odstranění přiřazení u položek
                db.execute_query(
                    "UPDATE warehouse SET supplier_id = NULL WHERE supplier_id = ?",
                    [supplier_id]
                )

                # Odstranění u pohybů
                db.execute_query(
                    "UPDATE warehouse_movements SET supplier_id = NULL WHERE supplier_id = ?",
                    [supplier_id]
                )

                # Smazání dodavatele
                db.execute_query(
                    "DELETE FROM warehouse_suppliers WHERE id = ?",
                    [supplier_id]
                )

                QMessageBox.information(self, "Úspěch", "Dodavatel byl smazán")
                self.load_suppliers()
                self.suppliers_changed.emit()

            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Chyba při mazání:\n{str(e)}")

    def export_suppliers(self):
        """Export dodavatelů"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit jako Excel",
                f"dodavatele_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel soubory (*.xlsx)"
            )

            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Dodavatelé"

            # Hlavička
            headers = [
                "Název", "IČO", "Kontaktní osoba", "Telefon", "Email",
                "Ulice", "Město", "PSČ", "Země", "Platební podmínky", "Poznámka"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            suppliers = db.execute_query(
                """SELECT name, ico, contact_person, phone, email, street, city,
                   postal_code, country, payment_terms, note
                   FROM warehouse_suppliers ORDER BY name"""
            )

            row_num = 2
            if suppliers:
                for sup in suppliers:
                    for col, value in enumerate(sup, 1):
                        ws.cell(row=row_num, column=col, value=value or "")
                    row_num += 1

            # Auto-šířka
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Úspěch",
                f"Dodavatelé byly vyexportováni do:\n{file_path}"
            )

            try:
                import os
                os.startfile(file_path)
            except:
                pass

        except ImportError:
            QMessageBox.warning(
                self,
                "Chybí knihovna",
                "Pro export je potřeba:\n\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při exportu:\n{str(e)}")


class SupplierDetailDialog(QDialog):
    """Dialog s detailem dodavatele"""

    supplier_saved = pyqtSignal()

    def __init__(self, supplier_id=None, parent=None):
        super().__init__(parent)
        self.supplier_id = supplier_id
        self.is_new = supplier_id is None

        self.setWindowTitle("Nový dodavatel" if self.is_new else "Detail dodavatele")
        self.setModal(True)
        self.setMinimumSize(800, 700)

        self.init_ui()

        if not self.is_new:
            self.load_supplier_data()

    def init_ui(self):
        """Inicializace UI"""
        layout = QVBoxLayout(self)

        # Hlavička
        header = QLabel("🚚 " + ("NOVÝ DODAVATEL" if self.is_new else "DETAIL DODAVATELE"))
        header.setStyleSheet(f"""
            font-size: 16px;
            font-weight: bold;
            color: {config.COLOR_PRIMARY};
            padding: 15px;
            background-color: #ecf0f1;
            border-radius: 5px;
        """)
        layout.addWidget(header)

        # === ZÁLOŽKY ===
        self.tabs = QTabWidget()

        # ZÁLOŽKA 1: Základní údaje
        self.tab_basic = self.create_tab_basic()
        self.tabs.addTab(self.tab_basic, "📋 Základní údaje")

        # ZÁLOŽKA 2: Položky dodavatele
        if not self.is_new:
            self.tab_items = self.create_tab_items()
            self.tabs.addTab(self.tab_items, "📦 Položky")

            # ZÁLOŽKA 3: Historie objednávek
            self.tab_history = self.create_tab_history()
            self.tabs.addTab(self.tab_history, "📊 Historie")

        layout.addWidget(self.tabs)

        # === TLAČÍTKA ===
        buttons = QHBoxLayout()

        btn_cancel = QPushButton("Zrušit")
        btn_cancel.clicked.connect(self.reject)

        btn_save = QPushButton("💾 Uložit")
        btn_save.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 12px 30px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
            }}
        """)
        btn_save.clicked.connect(self.save)

        if not self.is_new:
            btn_delete = QPushButton("🗑️ Smazat")
            btn_delete.setStyleSheet(f"""
                QPushButton {{
                    background-color: {config.COLOR_DANGER};
                    color: white;
                    padding: 12px 30px;
                    border-radius: 5px;
                    font-weight: bold;
                    font-size: 14px;
                }}
            """)
            btn_delete.clicked.connect(self.delete)
            buttons.addWidget(btn_delete)

        buttons.addStretch()
        buttons.addWidget(btn_cancel)
        buttons.addWidget(btn_save)

        layout.addLayout(buttons)

    def create_tab_basic(self):
        """ZÁLOŽKA: Základní údaje"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # === ZÁKLADNÍ INFO ===
        basic_group = QGroupBox("📋 Základní informace")
        basic_form = QFormLayout(basic_group)

        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("Název dodavatele...")
        basic_form.addRow("Název *:", self.input_name)

        self.input_ico = QLineEdit()
        self.input_ico.setPlaceholderText("IČO...")
        basic_form.addRow("IČO:", self.input_ico)

        self.input_dic = QLineEdit()
        self.input_dic.setPlaceholderText("DIČ...")
        basic_form.addRow("DIČ:", self.input_dic)

        layout.addWidget(basic_group)

        # === KONTAKT ===
        contact_group = QGroupBox("👤 Kontaktní údaje")
        contact_form = QFormLayout(contact_group)

        self.input_contact_person = QLineEdit()
        self.input_contact_person.setPlaceholderText("Jméno kontaktní osoby...")
        contact_form.addRow("Kontaktní osoba:", self.input_contact_person)

        self.input_phone = QLineEdit()
        self.input_phone.setPlaceholderText("+420 xxx xxx xxx")
        contact_form.addRow("Telefon:", self.input_phone)

        self.input_email = QLineEdit()
        self.input_email.setPlaceholderText("email@dodavatel.cz")
        contact_form.addRow("Email:", self.input_email)

        self.input_web = QLineEdit()
        self.input_web.setPlaceholderText("www.dodavatel.cz")
        contact_form.addRow("Web:", self.input_web)

        layout.addWidget(contact_group)

        # === ADRESA ===
        address_group = QGroupBox("📍 Adresa")
        address_form = QFormLayout(address_group)

        self.input_street = QLineEdit()
        self.input_street.setPlaceholderText("Ulice a číslo popisné...")
        address_form.addRow("Ulice:", self.input_street)

        self.input_city = QLineEdit()
        self.input_city.setPlaceholderText("Město...")
        address_form.addRow("Město:", self.input_city)

        self.input_postal = QLineEdit()
        self.input_postal.setPlaceholderText("PSČ...")
        address_form.addRow("PSČ:", self.input_postal)

        self.input_country = QLineEdit()
        self.input_country.setText("Česká republika")
        address_form.addRow("Země:", self.input_country)

        layout.addWidget(address_group)

        # === PLATEBNÍ PODMÍNKY ===
        payment_group = QGroupBox("💰 Platební podmínky")
        payment_form = QFormLayout(payment_group)

        self.combo_payment_terms = QComboBox()
        self.combo_payment_terms.addItems([
            "7 dní",
            "14 dní",
            "21 dní",
            "30 dní",
            "60 dní",
            "90 dní",
            "Hotově",
            "Záloha",
            "Jiné"
        ])
        self.combo_payment_terms.setEditable(True)
        payment_form.addRow("Splatnost:", self.combo_payment_terms)

        self.input_bank_account = QLineEdit()
        self.input_bank_account.setPlaceholderText("Číslo účtu / IBAN...")
        payment_form.addRow("Bankovní účet:", self.input_bank_account)

        layout.addWidget(payment_group)

        # === POZNÁMKA ===
        note_group = QGroupBox("📝 Poznámka")
        note_layout = QVBoxLayout(note_group)

        self.text_note = QTextEdit()
        self.text_note.setMaximumHeight(100)
        self.text_note.setPlaceholderText("Interní poznámky k dodavateli...")
        note_layout.addWidget(self.text_note)

        layout.addWidget(note_group)

        layout.addStretch()

        return widget

    def create_tab_items(self):
        """ZÁLOŽKA: Položky dodavatele"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Info
        info = QLabel("📦 Položky od tohoto dodavatele")
        info.setStyleSheet("padding: 10px; background-color: #ecf0f1; font-weight: bold;")
        layout.addWidget(info)

        # Tabulka
        self.table_items = QTableWidget()
        self.table_items.setColumnCount(6)
        self.table_items.setHorizontalHeaderLabels([
            "Název", "Kód", "Množství", "Jednotka", "Nákupní cena", "Hodnota"
        ])
        self.table_items.horizontalHeader().setStretchLastSection(True)
        self.table_items.setAlternatingRowColors(True)

        layout.addWidget(self.table_items)

        return widget

    def create_tab_history(self):
        """ZÁLOŽKA: Historie objednávek"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Info
        info = QLabel("📊 Historie příjmů od dodavatele")
        info.setStyleSheet("padding: 10px; background-color: #ecf0f1; font-weight: bold;")
        layout.addWidget(info)

        # Tabulka
        self.table_history = QTableWidget()
        self.table_history.setColumnCount(6)
        self.table_history.setHorizontalHeaderLabels([
            "Datum", "Položka", "Množství", "Cena/jedn.", "Celkem", "Doklad"
        ])
        self.table_history.horizontalHeader().setStretchLastSection(True)
        self.table_history.setAlternatingRowColors(True)

        layout.addWidget(self.table_history)

        return widget

    def load_supplier_data(self):
        """Načtení dat dodavatele"""
        try:
            supplier = db.execute_query(
                """SELECT name, ico, dic, contact_person, phone, email, web,
                   street, city, postal_code, country, payment_terms, bank_account, note
                   FROM warehouse_suppliers WHERE id = ?""",
                [self.supplier_id]
            )

            if not supplier:
                return

            s = supplier[0]

            self.input_name.setText(s[0] or "")
            self.input_ico.setText(s[1] or "")
            self.input_dic.setText(s[2] or "")
            self.input_contact_person.setText(s[3] or "")
            self.input_phone.setText(s[4] or "")
            self.input_email.setText(s[5] or "")
            self.input_web.setText(s[6] or "")
            self.input_street.setText(s[7] or "")
            self.input_city.setText(s[8] or "")
            self.input_postal.setText(s[9] or "")
            self.input_country.setText(s[10] or "Česká republika")
            self.combo_payment_terms.setCurrentText(s[11] or "30 dní")
            self.input_bank_account.setText(s[12] or "")
            self.text_note.setPlainText(s[13] or "")

            # Načtení položek
            self.load_supplier_items()

            # Načtení historie
            self.load_supplier_history()

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba:\n{str(e)}")

    def load_supplier_items(self):
        """Načtení položek dodavatele"""
        try:
            items = db.execute_query(
                """SELECT name, code, quantity, unit, price_purchase
                   FROM warehouse
                   WHERE supplier_id = ?
                   ORDER BY name""",
                [self.supplier_id]
            )

            self.table_items.setRowCount(0)

            if not items:
                return

            for item in items:
                row = self.table_items.rowCount()
                self.table_items.insertRow(row)

                name = item[0]
                code = item[1] or ""
                quantity = item[2]
                unit = item[3]
                price = item[4]
                value = quantity * price

                self.table_items.setItem(row, 0, QTableWidgetItem(name))
                self.table_items.setItem(row, 1, QTableWidgetItem(code))
                self.table_items.setItem(row, 2, QTableWidgetItem(f"{quantity:.2f}"))
                self.table_items.setItem(row, 3, QTableWidgetItem(unit))
                self.table_items.setItem(row, 4, QTableWidgetItem(f"{price:.2f} Kč"))
                self.table_items.setItem(row, 5, QTableWidgetItem(f"{value:.2f} Kč"))

        except Exception as e:
            print(f"Chyba: {e}")

    def load_supplier_history(self):
        """Načtení historie"""
        try:
            history = db.execute_query(
                """SELECT wm.date, w.name, wm.quantity, wm.unit_price, wm.document_number
                   FROM warehouse_movements wm
                   LEFT JOIN warehouse w ON wm.item_id = w.id
                   WHERE wm.supplier_id = ? AND wm.movement_type = 'Příjem'
                   ORDER BY wm.date DESC
                   LIMIT 100""",
                [self.supplier_id]
            )

            self.table_history.setRowCount(0)

            if not history:
                return

            for h in history:
                row = self.table_history.rowCount()
                self.table_history.insertRow(row)

                date = h[0]
                item_name = h[1] or "---"
                quantity = h[2]
                price = h[3]
                total = quantity * price
                document = h[4] or ""

                self.table_history.setItem(row, 0, QTableWidgetItem(date))
                self.table_history.setItem(row, 1, QTableWidgetItem(item_name))
                self.table_history.setItem(row, 2, QTableWidgetItem(f"{quantity:.2f}"))
                self.table_history.setItem(row, 3, QTableWidgetItem(f"{price:.2f} Kč"))
                self.table_history.setItem(row, 4, QTableWidgetItem(f"{total:.2f} Kč"))
                self.table_history.setItem(row, 5, QTableWidgetItem(document))

        except Exception as e:
            print(f"Chyba: {e}")

    def save(self):
        """Uložení dodavatele"""
        if not self.input_name.text():
            QMessageBox.warning(self, "Chyba", "Vyplňte název dodavatele!")
            self.input_name.setFocus()
            return

        try:
            data = {
                'name': self.input_name.text(),
                'ico': self.input_ico.text(),
                'dic': self.input_dic.text(),
                'contact_person': self.input_contact_person.text(),
                'phone': self.input_phone.text(),
                'email': self.input_email.text(),
                'web': self.input_web.text(),
                'street': self.input_street.text(),
                'city': self.input_city.text(),
                'postal_code': self.input_postal.text(),
                'country': self.input_country.text(),
                'payment_terms': self.combo_payment_terms.currentText(),
                'bank_account': self.input_bank_account.text(),
                'note': self.text_note.toPlainText()
            }

            if self.is_new:
                db.execute_query(
                    """INSERT INTO warehouse_suppliers
                       (name, ico, dic, contact_person, phone, email, web, street, city,
                        postal_code, country, payment_terms, bank_account, note)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    list(data.values())
                )
                QMessageBox.information(self, "Úspěch", "Dodavatel byl přidán")
            else:
                db.execute_query(
                    """UPDATE warehouse_suppliers SET
                       name=?, ico=?, dic=?, contact_person=?, phone=?, email=?, web=?,
                       street=?, city=?, postal_code=?, country=?, payment_terms=?,
                       bank_account=?, note=?
                       WHERE id=?""",
                    list(data.values()) + [self.supplier_id]
                )
                QMessageBox.information(self, "Úspěch", "Dodavatel byl aktualizován")

            self.supplier_saved.emit()
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při ukládání:\n{str(e)}")

    def delete(self):
        """Smazání dodavatele"""
        reply = QMessageBox.question(
            self,
            "Smazat dodavatele?",
            f"Opravdu smazat dodavatele '{self.input_name.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                db.execute_query(
                    "UPDATE warehouse SET supplier_id = NULL WHERE supplier_id = ?",
                    [self.supplier_id]
                )

                db.execute_query(
                    "DELETE FROM warehouse_suppliers WHERE id = ?",
                    [self.supplier_id]
                )

                QMessageBox.information(self, "Smazáno", "Dodavatel byl smazán")
                self.supplier_saved.emit()
                self.accept()

            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Chyba:\n{str(e)}")
