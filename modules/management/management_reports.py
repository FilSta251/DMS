# -*- coding: utf-8 -*-
"""
Management Reports - Generování manažerských reportů
"""
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
                             QPushButton, QLabel, QFrame, QScrollArea, QComboBox,
                             QDateEdit, QGroupBox, QRadioButton, QButtonGroup,
                             QTextEdit, QMessageBox, QFileDialog, QProgressDialog)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont
from database_manager import db
from datetime import datetime, timedelta
import json
import os


class ManagementReports(QWidget):
    """Generování manažerských reportů"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_module = parent
        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        content_widget = QWidget()
        scroll.setWidget(content_widget)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(20)

        # Úvodní info
        self.create_info_section(content_layout)

        # Výběr typu reportu
        self.create_report_type_section(content_layout)

        # Výběr období
        self.create_period_section(content_layout)

        # Výběr formátu
        self.create_format_section(content_layout)

        # Náhled obsahu
        self.create_preview_section(content_layout)

        # Tlačítka akcí
        self.create_action_buttons(content_layout)

        content_layout.addStretch()

    def create_info_section(self, parent_layout):
        """Úvodní informace"""
        info_frame = QFrame()
        info_frame.setObjectName("infoFrame")
        info_layout = QVBoxLayout(info_frame)

        title = QLabel("📄 Generování manažerských reportů")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        info_layout.addWidget(title)

        desc = QLabel(
            "Vytvořte profesionální reporty s klíčovými metrikami, grafy a analýzami.\n"
            "Dostupné formáty: PDF, Excel, PowerPoint, JSON, Email."
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        info_layout.addWidget(desc)

        info_frame.setStyleSheet("""
            QFrame#infoFrame {
                background-color: #ecf0f1;
                border-radius: 8px;
                padding: 15px;
            }
        """)

        parent_layout.addWidget(info_frame)

    def create_report_type_section(self, parent_layout):
        """Sekce s výběrem typu reportu"""
        type_group = QGroupBox("📊 Typ reportu")
        type_layout = QVBoxLayout(type_group)

        self.report_type_group = QButtonGroup()

        types = [
            ("monthly", "📅 Měsíční report", "Přehled za jeden měsíc"),
            ("quarterly", "📆 Čtvrtletní report", "Přehled za celé čtvrtletí"),
            ("yearly", "📖 Roční report", "Komplexní roční přehled"),
            ("custom", "🔧 Vlastní období", "Zvolte si vlastní časové období")
        ]

        for value, label, tooltip in types:
            radio = QRadioButton(label)
            radio.setToolTip(tooltip)
            radio.setProperty("report_type", value)
            self.report_type_group.addButton(radio)
            type_layout.addWidget(radio)

        # Výchozí výběr
        self.report_type_group.buttons()[0].setChecked(True)
        self.report_type_group.buttonClicked.connect(self.on_report_type_changed)

        parent_layout.addWidget(type_group)

    def create_period_section(self, parent_layout):
        """Sekce s výběrem období"""
        period_group = QGroupBox("📅 Období")
        period_layout = QVBoxLayout(period_group)

        # Rychlé volby
        quick_layout = QHBoxLayout()

        quick_label = QLabel("Rychlé volby:")
        quick_label.setStyleSheet("font-weight: bold;")
        quick_layout.addWidget(quick_label)

        self.period_combo = QComboBox()
        self.period_combo.addItems([
            "Tento měsíc",
            "Minulý měsíc",
            "Toto čtvrtletí",
            "Minulé čtvrtletí",
            "Tento rok",
            "Minulý rok",
            "Poslední 3 měsíce",
            "Poslední 6 měsíců",
            "Poslední 12 měsíců"
        ])
        self.period_combo.currentIndexChanged.connect(self.on_quick_period_changed)
        quick_layout.addWidget(self.period_combo)

        quick_layout.addStretch()
        period_layout.addLayout(quick_layout)

        # Vlastní období
        custom_layout = QHBoxLayout()

        custom_layout.addWidget(QLabel("Od:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        custom_layout.addWidget(self.date_from)

        custom_layout.addWidget(QLabel("Do:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        custom_layout.addWidget(self.date_to)

        custom_layout.addStretch()
        period_layout.addLayout(custom_layout)

        parent_layout.addWidget(period_group)

    def create_format_section(self, parent_layout):
        """Sekce s výběrem formátu"""
        format_group = QGroupBox("💾 Formát exportu")
        format_layout = QGridLayout(format_group)

        self.format_buttons = {}

        formats = [
            ("pdf", "📄 PDF", "Profesionální PDF dokument", 0, 0),
            ("excel", "📊 Excel", "Tabulkový soubor s daty", 0, 1),
            ("pptx", "📽️ PowerPoint", "Prezentace pro management", 0, 2),
            ("json", "📋 JSON", "Strukturovaná data", 1, 0),
            ("html", "🌐 HTML", "Webová stránka", 1, 1),
            ("email", "📧 Email", "Odeslat na email", 1, 2)
        ]

        for fmt_id, label, tooltip, row, col in formats:
            btn = QPushButton(label)
            btn.setToolTip(tooltip)
            btn.setCheckable(True)
            btn.setMinimumHeight(60)
            btn.setProperty("format_id", fmt_id)
            btn.clicked.connect(self.on_format_selected)
            self.format_buttons[fmt_id] = btn
            format_layout.addWidget(btn, row, col)

        # Výchozí výběr
        self.format_buttons["pdf"].setChecked(True)

        parent_layout.addWidget(format_group)

    def create_preview_section(self, parent_layout):
        """Náhled obsahu reportu"""
        preview_group = QGroupBox("👁️ Náhled obsahu reportu")
        preview_layout = QVBoxLayout(preview_group)

        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMinimumHeight(300)
        self.preview_text.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 11px;
            }
        """)

        preview_layout.addWidget(self.preview_text)

        # Tlačítko náhledu
        preview_btn = QPushButton("🔄 Aktualizovat náhled")
        preview_btn.clicked.connect(self.update_preview)
        preview_layout.addWidget(preview_btn)

        parent_layout.addWidget(preview_group)

    def create_action_buttons(self, parent_layout):
        """Akční tlačítka"""
        buttons_layout = QHBoxLayout()

        buttons_layout.addStretch()

        # Tlačítko generování
        generate_btn = QPushButton("📄 Generovat report")
        generate_btn.setMinimumHeight(50)
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 30px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        generate_btn.clicked.connect(self.generate_report)
        buttons_layout.addWidget(generate_btn)

        # Tlačítko náhledu
        preview_btn = QPushButton("👁️ Náhled")
        preview_btn.setMinimumHeight(50)
        preview_btn.clicked.connect(self.preview_report)
        buttons_layout.addWidget(preview_btn)

        parent_layout.addLayout(buttons_layout)

    def on_report_type_changed(self, button):
        """Změna typu reportu"""
        report_type = button.property("report_type")

        # Automatické nastavení období podle typu
        if report_type == "monthly":
            self.period_combo.setCurrentIndex(0)  # Tento měsíc
        elif report_type == "quarterly":
            self.period_combo.setCurrentIndex(2)  # Toto čtvrtletí
        elif report_type == "yearly":
            self.period_combo.setCurrentIndex(4)  # Tento rok

    def on_quick_period_changed(self, index):
        """Změna rychlého výběru období"""
        today = QDate.currentDate()

        periods = {
            0: (QDate(today.year(), today.month(), 1), today),  # Tento měsíc
            1: (QDate(today.year(), today.month(), 1).addMonths(-1),
                QDate(today.year(), today.month(), 1).addDays(-1)),  # Minulý měsíc
            2: self.get_quarter_dates(today, 0),  # Toto čtvrtletí
            3: self.get_quarter_dates(today, -1),  # Minulé čtvrtletí
            4: (QDate(today.year(), 1, 1), today),  # Tento rok
            5: (QDate(today.year()-1, 1, 1), QDate(today.year()-1, 12, 31)),  # Minulý rok
            6: (today.addMonths(-3), today),  # Poslední 3 měsíce
            7: (today.addMonths(-6), today),  # Poslední 6 měsíců
            8: (today.addMonths(-12), today),  # Poslední 12 měsíců
        }

        if index in periods:
            date_from, date_to = periods[index]
            self.date_from.setDate(date_from)
            self.date_to.setDate(date_to)

    def get_quarter_dates(self, date, offset=0):
        """Získání dat čtvrtletí"""
        quarter = (date.month() - 1) // 3 + 1 + offset
        year = date.year()

        if quarter < 1:
            quarter += 4
            year -= 1
        elif quarter > 4:
            quarter -= 4
            year += 1

        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2

        date_from = QDate(year, start_month, 1)
        date_to = QDate(year, end_month, 1).addMonths(1).addDays(-1)

        return date_from, date_to

    def on_format_selected(self):
        """Výběr formátu"""
        # Zrušit ostatní výběry
        sender = self.sender()
        for btn in self.format_buttons.values():
            if btn != sender:
                btn.setChecked(False)

    def get_selected_format(self):
        """Získání vybraného formátu"""
        for fmt_id, btn in self.format_buttons.items():
            if btn.isChecked():
                return fmt_id
        return "pdf"

    def get_selected_report_type(self):
        """Získání vybraného typu reportu"""
        for button in self.report_type_group.buttons():
            if button.isChecked():
                return button.property("report_type")
        return "monthly"

    def update_preview(self):
        """Aktualizace náhledu"""
        try:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            report_type = self.get_selected_report_type()

            # Načtení dat
            data = self.collect_report_data(date_from, date_to)

            # Vytvoření náhledu
            preview = self.generate_preview_text(data, report_type)
            self.preview_text.setPlainText(preview)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování náhledu:\n{e}")

    def collect_report_data(self, date_from, date_to):
        """Sběr dat pro report"""
        data = {
            'period': {
                'from': date_from,
                'to': date_to
            },
            'metrics': {},
            'orders': {},
            'financials': {},
            'mechanics': {},
            'warehouse': {}
        }

        try:
            # Základní metriky
            query = """
                SELECT
                    COUNT(*) as order_count,
                    COALESCE(SUM(total_price), 0) as total_revenue,
                    COALESCE(SUM(material_cost), 0) as total_costs,
                    COALESCE(AVG(total_price), 0) as avg_order_value
                FROM orders
                WHERE created_date BETWEEN ? AND ?
                AND status != 'Zrušeno'
            """
            result = db.fetch_one(query, (date_from, date_to))
            if result:
                data['metrics']['order_count'] = result[0]
                data['metrics']['total_revenue'] = result[1]
                data['metrics']['total_costs'] = result[2]
                data['metrics']['avg_order_value'] = result[3]
                data['metrics']['profit'] = result[1] - result[2]
                data['metrics']['margin'] = ((result[1] - result[2]) / result[1] * 100) if result[1] > 0 else 0

            # Zakázky podle typu
            query = """
                SELECT order_type, COUNT(*) as count
                FROM orders
                WHERE created_date BETWEEN ? AND ?
                AND status != 'Zrušeno'
                GROUP BY order_type
            """
            results = db.fetch_all(query, (date_from, date_to))
            data['orders']['by_type'] = {r[0]: r[1] for r in results} if results else {}

            # Mechanici
            query = """
                SELECT
                    u.full_name,
                    COALESCE(SUM(wl.hours_worked), 0) as hours,
                    COUNT(DISTINCT wl.order_id) as orders
                FROM users u
                LEFT JOIN order_work_log wl ON u.id = wl.user_id
                    AND wl.date BETWEEN ? AND ?
                WHERE u.role = 'mechanik' AND u.active = 1
                GROUP BY u.id, u.full_name
                ORDER BY hours DESC
                LIMIT 5
            """
            results = db.fetch_all(query, (date_from, date_to))
            data['mechanics']['top'] = [
                {'name': r[0], 'hours': r[1], 'orders': r[2]}
                for r in results
            ] if results else []

            # Sklad
            query = """
                SELECT
                    COALESCE(SUM(quantity * price_purchase), 0) as value,
                    COUNT(*) as item_count,
                    SUM(CASE WHEN quantity <= min_quantity THEN 1 ELSE 0 END) as low_stock
                FROM warehouse
                WHERE quantity > 0
            """
            result = db.fetch_one(query)
            if result:
                data['warehouse']['total_value'] = result[0]
                data['warehouse']['item_count'] = result[1]
                data['warehouse']['low_stock_count'] = result[2]

        except Exception as e:
            print(f"Chyba při sběru dat: {e}")

        return data

    def generate_preview_text(self, data, report_type):
        """Generování textového náhledu"""
        type_names = {
            'monthly': 'MĚSÍČNÍ',
            'quarterly': 'ČTVRTLETNÍ',
            'yearly': 'ROČNÍ',
            'custom': 'VLASTNÍ'
        }

        preview = f"""
═══════════════════════════════════════════════════════════════
                    {type_names.get(report_type, 'VLASTNÍ')} REPORT
                        MOTOSERVIS DMS
═══════════════════════════════════════════════════════════════

OBDOBÍ: {data['period']['from']} až {data['period']['to']}
DATUM GENEROVÁNÍ: {datetime.now().strftime('%d.%m.%Y %H:%M')}

───────────────────────────────────────────────────────────────
📊 KLÍČOVÉ METRIKY
───────────────────────────────────────────────────────────────

Celkový obrat:           {data['metrics'].get('total_revenue', 0):,.0f} Kč
Celkové náklady:         {data['metrics'].get('total_costs', 0):,.0f} Kč
Zisk:                    {data['metrics'].get('profit', 0):,.0f} Kč
Marže:                   {data['metrics'].get('margin', 0):.1f}%

Počet zakázek:           {data['metrics'].get('order_count', 0)}
Průměrná zakázka:        {data['metrics'].get('avg_order_value', 0):,.0f} Kč

───────────────────────────────────────────────────────────────
📋 ZAKÁZKY
───────────────────────────────────────────────────────────────
"""

        # Zakázky podle typu
        if data['orders'].get('by_type'):
            for order_type, count in data['orders']['by_type'].items():
                preview += f"\n{order_type:20s} {count:5d} zakázek"
        else:
            preview += "\nŽádné zakázky v tomto období"

        preview += """

───────────────────────────────────────────────────────────────
👨‍🔧 TOP MECHANICI
───────────────────────────────────────────────────────────────
"""

        # Top mechanici
        if data['mechanics'].get('top'):
            for i, mech in enumerate(data['mechanics']['top'], 1):
                preview += f"\n{i}. {mech['name']:30s} {mech['hours']:6.1f}h  {mech['orders']:3d} zakázek"
        else:
            preview += "\nŽádná data o mechanicích"

        preview += """

───────────────────────────────────────────────────────────────
📦 SKLAD
───────────────────────────────────────────────────────────────
"""

        preview += f"""
Celková hodnota skladu:  {data['warehouse'].get('total_value', 0):,.0f} Kč
Počet položek:           {data['warehouse'].get('item_count', 0)}
Položky pod minimem:     {data['warehouse'].get('low_stock_count', 0)}

═══════════════════════════════════════════════════════════════
                        KONEC NÁHLEDU
═══════════════════════════════════════════════════════════════
"""

        return preview

    def preview_report(self):
        """Náhled reportu"""
        self.update_preview()
        QMessageBox.information(
            self,
            "Náhled",
            "Náhled byl aktualizován níže.\n\n"
            "Pro generování kompletního reportu použijte tlačítko 'Generovat report'."
        )

    def generate_report(self):
        """Generování reportu"""
        try:
            fmt = self.get_selected_format()
            report_type = self.get_selected_report_type()
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")

            # Sběr dat
            data = self.collect_report_data(date_from, date_to)

            # Generování podle formátu
            if fmt == "pdf":
                self.generate_pdf_report(data, report_type)
            elif fmt == "excel":
                self.generate_excel_report(data, report_type)
            elif fmt == "pptx":
                self.generate_pptx_report(data, report_type)
            elif fmt == "json":
                self.generate_json_report(data, report_type)
            elif fmt == "html":
                self.generate_html_report(data, report_type)
            elif fmt == "email":
                self.send_email_report(data, report_type)
            else:
                QMessageBox.warning(self, "Upozornění", f"Formát {fmt} zatím není implementován")

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování reportu:\n{e}")

    def generate_pdf_report(self, data, report_type):
        """Generování PDF reportu"""
        try:
            # Výběr kam uložit
            default_name = f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit PDF report",
                default_name,
                "PDF soubory (*.pdf)"
            )

            if not file_path:
                return

            # Pokus o import ReportLab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib import colors
                from reportlab.lib.units import cm

                # Vytvoření dokumentu
                doc = SimpleDocTemplate(file_path, pagesize=A4)
                story = []
                styles = getSampleStyleSheet()

                # Titulek
                title_style = styles['Heading1']
                title = Paragraph(f"MANAŽERSKÝ REPORT - {report_type.upper()}", title_style)
                story.append(title)
                story.append(Spacer(1, 1*cm))

                # Období
                story.append(Paragraph(f"<b>Období:</b> {data['period']['from']} až {data['period']['to']}", styles['Normal']))
                story.append(Spacer(1, 0.5*cm))

                # Metriky
                story.append(Paragraph("<b>KLÍČOVÉ METRIKY</b>", styles['Heading2']))

                metrics_data = [
                    ['Metrika', 'Hodnota'],
                    ['Celkový obrat', f"{data['metrics'].get('total_revenue', 0):,.0f} Kč"],
                    ['Celkové náklady', f"{data['metrics'].get('total_costs', 0):,.0f} Kč"],
                    ['Zisk', f"{data['metrics'].get('profit', 0):,.0f} Kč"],
                    ['Marže', f"{data['metrics'].get('margin', 0):.1f}%"],
                    ['Počet zakázek', str(data['metrics'].get('order_count', 0))],
                ]

                table = Table(metrics_data, colWidths=[10*cm, 6*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(table)
                story.append(Spacer(1, 1*cm))

                # Build PDF
                doc.build(story)

                QMessageBox.information(
                    self,
                    "Úspěch",
                    f"PDF report byl úspěšně vytvořen:\n{file_path}"
                )

            except ImportError:
                QMessageBox.warning(
                    self,
                    "Chybí knihovna",
                    "Pro generování PDF je potřeba nainstalovat ReportLab:\n\n"
                    "pip install reportlab\n\n"
                    "Zkuste formát JSON nebo HTML jako alternativu."
                )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování PDF:\n{e}")

    def generate_excel_report(self, data, report_type):
        """Generování Excel reportu"""
        try:
            default_name = f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit Excel report",
                default_name,
                "Excel soubory (*.xlsx)"
            )

            if not file_path:
                return

            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill

                wb = Workbook()
                ws = wb.active
                ws.title = "Report"

                # Hlavička
                ws['A1'] = f"MANAŽERSKÝ REPORT - {report_type.upper()}"
                ws['A1'].font = Font(bold=True, size=16)
                ws.merge_cells('A1:D1')

                ws['A2'] = f"Období: {data['period']['from']} až {data['period']['to']}"
                ws['A3'] = f"Datum: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

                # Metriky
                row = 5
                ws[f'A{row}'] = "KLÍČOVÉ METRIKY"
                ws[f'A{row}'].font = Font(bold=True, size=14)

                row += 1
                ws[f'A{row}'] = "Metrika"
                ws[f'B{row}'] = "Hodnota"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="CCCCCC", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="CCCCCC", fill_type="solid")

                metrics = [
                    ('Celkový obrat', f"{data['metrics'].get('total_revenue', 0):,.0f} Kč"),
                    ('Celkové náklady', f"{data['metrics'].get('total_costs', 0):,.0f} Kč"),
                    ('Zisk', f"{data['metrics'].get('profit', 0):,.0f} Kč"),
                    ('Marže', f"{data['metrics'].get('margin', 0):.1f}%"),
                    ('Počet zakázek', str(data['metrics'].get('order_count', 0))),
                ]

                for metric, value in metrics:
                    row += 1
                    ws[f'A{row}'] = metric
                    ws[f'B{row}'] = value

                # Nastavení šířky sloupců
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20

                wb.save(file_path)

                QMessageBox.information(
                    self,
                    "Úspěch",
                    f"Excel report byl úspěšně vytvořen:\n{file_path}"
                )

            except ImportError:
                QMessageBox.warning(
                    self,
                    "Chybí knihovna",
                    "Pro generování Excel souborů je potřeba nainstalovat openpyxl:\n\n"
                    "pip install openpyxl\n\n"
                    "Zkuste formát JSON nebo HTML jako alternativu."
                )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování Excel:\n{e}")

    def generate_pptx_report(self, data, report_type):
        """Generování PowerPoint reportu"""
        QMessageBox.information(
            self,
            "PowerPoint",
            "Generování PowerPoint reportů bude implementováno v příští verzi.\n\n"
            "Prozatím použijte PDF nebo Excel formát."
        )

    def generate_json_report(self, data, report_type):
        """Generování JSON reportu"""
        try:
            default_name = f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.json"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit JSON report",
                default_name,
                "JSON soubory (*.json)"
            )

            if not file_path:
                return

            # Přidání metadat
            report_data = {
                'report_type': report_type,
                'generated_at': datetime.now().isoformat(),
                'period': data['period'],
                'data': data
            }

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2)

            QMessageBox.information(
                self,
                "Úspěch",
                f"JSON report byl úspěšně vytvořen:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování JSON:\n{e}")

    def generate_html_report(self, data, report_type):
        """Generování HTML reportu"""
        try:
            default_name = f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.html"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Uložit HTML report",
                default_name,
                "HTML soubory (*.html)"
            )

            if not file_path:
                return

            html = f"""
<!DOCTYPE html>
<html lang="cs">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Manažerský Report - {report_type}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background-color: #2c3e50;
            color: white;
            padding: 30px;
            text-align: center;
            border-radius: 8px;
            margin-bottom: 30px;
        }}
        .metrics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .metric-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .metric-value {{
            font-size: 24px;
            font-weight: bold;
            color: #27ae60;
            margin-top: 10px;
        }}
        .metric-label {{
            color: #7f8c8d;
            font-size: 14px;
        }}
        table {{
            width: 100%;
            background: white;
            border-collapse: collapse;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background-color: #34495e;
            color: white;
            padding: 15px;
            text-align: left;
        }}
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #ecf0f1;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>MANAŽERSKÝ REPORT</h1>
        <p>{report_type.upper()}</p>
        <p>Období: {data['period']['from']} až {data['period']['to']}</p>
        <p>Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
    </div>

    <h2>📊 Klíčové metriky</h2>
    <div class="metrics">
        <div class="metric-card">
            <div class="metric-label">Celkový obrat</div>
            <div class="metric-value">{data['metrics'].get('total_revenue', 0):,.0f} Kč</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Celkové náklady</div>
            <div class="metric-value">{data['metrics'].get('total_costs', 0):,.0f} Kč</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Zisk</div>
            <div class="metric-value">{data['metrics'].get('profit', 0):,.0f} Kč</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Marže</div>
            <div class="metric-value">{data['metrics'].get('margin', 0):.1f}%</div>
        </div>
    </div>

    <h2>👨‍🔧 Top mechanici</h2>
    <table>
        <thead>
            <tr>
                <th>Pořadí</th>
                <th>Mechanik</th>
                <th>Hodiny</th>
                <th>Zakázky</th>
            </tr>
        </thead>
        <tbody>
"""

            if data['mechanics'].get('top'):
                for i, mech in enumerate(data['mechanics']['top'], 1):
                    html += f"""
            <tr>
                <td>{i}</td>
                <td>{mech['name']}</td>
                <td>{mech['hours']:.1f}h</td>
                <td>{mech['orders']}</td>
            </tr>
"""
            else:
                html += "<tr><td colspan='4'>Žádná data</td></tr>"

            html += """
        </tbody>
    </table>
</body>
</html>
"""

            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html)

            QMessageBox.information(
                self,
                "Úspěch",
                f"HTML report byl úspěšně vytvořen:\n{file_path}\n\n"
                f"Můžete ho otevřít v prohlížeči."
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při generování HTML:\n{e}")

    def send_email_report(self, data, report_type):
        """Odeslání reportu emailem"""
        QMessageBox.information(
            self,
            "Email",
            "Automatické odesílání emailů bude implementováno v příští verzi.\n\n"
            "Prozatím vygenerujte PDF nebo HTML a odešlete manuálně."
        )
