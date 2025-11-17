# -*- coding: utf-8 -*-
"""
Historie servisů vozidla - kompletní přehled všech servisních úkonů
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QComboBox, QFrame, QHeaderView, QMessageBox,
    QDialog, QGridLayout, QTextEdit, QGroupBox, QFormLayout, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor, QBrush
from datetime import datetime, date, timedelta
import config
from database_manager import db


class VehicleServiceHistoryWidget(QWidget):
    """Widget pro zobrazení servisní historie vozidla"""

    def __init__(self, vehicle_id, parent=None):
        super().__init__(parent)
        self.vehicle_id = vehicle_id
        self.all_orders = []
        self.init_ui()
        self.load_history()

    def init_ui(self):
        """Inicializace uživatelského rozhraní"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(0, 0, 0, 0)

        # Hlavička s filtry
        header_panel = self.create_header_panel()
        layout.addWidget(header_panel)

        # Statistiky
        self.stats_panel = self.create_stats_panel()
        layout.addWidget(self.stats_panel)

        # Tabulka historie
        self.table = self.create_table()
        layout.addWidget(self.table)

    def create_header_panel(self):
        """Vytvoření hlavičky s filtry"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background-color: #ecf0f1;
                border-radius: 8px;
                padding: 10px;
            }
        """)

        layout = QHBoxLayout(panel)
        layout.setSpacing(15)

        # Titulek
        title = QLabel("📋 Servisní historie")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        layout.addStretch()

        # Filtr období
        period_label = QLabel("Období:")
        self.period_filter = QComboBox()
        self.period_filter.addItems([
            "Vše",
            "Poslední rok",
            "Poslední 2 roky",
            "Poslední 3 roky",
            "Aktuální rok"
        ])
        self.period_filter.setMinimumWidth(150)
        self.period_filter.currentTextChanged.connect(self.apply_filters)

        layout.addWidget(period_label)
        layout.addWidget(self.period_filter)

        # Filtr typu
        type_label = QLabel("Typ:")
        self.type_filter = QComboBox()
        self.type_filter.addItems([
            "Vše",
            "Zakázka",
            "Volný prodej",
            "Interní zakázka",
            "Reklamace",
            "Nabídka"
        ])
        self.type_filter.setMinimumWidth(150)
        self.type_filter.currentTextChanged.connect(self.apply_filters)

        layout.addWidget(type_label)
        layout.addWidget(self.type_filter)

        # Tlačítko statistiky
        btn_stats = QPushButton("📊 Detailní statistiky")
        btn_stats.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SECONDARY};
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }}
            QPushButton:hover {{
                background-color: #2980b9;
            }}
        """)
        btn_stats.clicked.connect(self.show_detailed_statistics)
        layout.addWidget(btn_stats)

        # Tlačítko export
        btn_export = QPushButton("📤 Export")
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_export.clicked.connect(self.export_history)
        layout.addWidget(btn_export)

        return panel

    def create_stats_panel(self):
        """Vytvoření panelu se statistikami"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                padding: 10px;
                border: 1px solid #ddd;
            }
        """)

        layout = QHBoxLayout(panel)
        layout.setSpacing(20)

        # Statistiky - budou aktualizovány
        self.stat_total_orders = QLabel("<b>Celkem zakázek:</b> 0")
        self.stat_total_cost = QLabel("<b>Celková útrata:</b> 0 Kč")
        self.stat_avg_cost = QLabel("<b>Průměrná zakázka:</b> 0 Kč")
        self.stat_last_service = QLabel("<b>Poslední servis:</b> -")

        layout.addWidget(self.stat_total_orders)
        layout.addWidget(self.stat_total_cost)
        layout.addWidget(self.stat_avg_cost)
        layout.addWidget(self.stat_last_service)
        layout.addStretch()

        return panel

    def create_table(self):
        """Vytvoření tabulky historie"""
        table = QTableWidget()
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 10px;
                font-weight: bold;
                border: none;
            }
        """)

        # Sloupce
        columns = [
            "ID", "Datum", "Číslo zakázky", "Typ", "Stav km",
            "Popis prací", "Počet položek", "Cena bez DPH", "Cena s DPH", "Stav"
        ]
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)

        # Nastavení
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(True)
        table.setColumnHidden(0, True)  # Skrýt ID
        table.doubleClicked.connect(self.show_order_detail)

        # Roztažení sloupců
        header = table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)

        return table

    def load_history(self):
        """Načtení historie z databáze"""
        try:
            self.all_orders = db.fetch_all("""
                SELECT
                    o.id,
                    o.order_number,
                    o.order_type,
                    o.status,
                    o.created_date,
                    o.completed_date,
                    o.total_price,
                    o.note,
                    (SELECT COUNT(*) FROM order_items oi WHERE oi.order_id = o.id) as item_count,
                    (SELECT SUM(oi.quantity * oi.unit_price) FROM order_items oi WHERE oi.order_id = o.id) as subtotal
                FROM orders o
                WHERE o.vehicle_id = ?
                ORDER BY o.created_date DESC
            """, (self.vehicle_id,))

            self.populate_table(self.all_orders)
            self.update_statistics(self.all_orders)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se načíst historii:\n{e}")

    def populate_table(self, orders):
        """Naplnění tabulky daty"""
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)

        for order in orders:
            row = self.table.rowCount()
            self.table.insertRow(row)

            # ID (skrytý)
            id_item = QTableWidgetItem()
            id_item.setData(Qt.ItemDataRole.DisplayRole, order['id'])
            self.table.setItem(row, 0, id_item)

            # Datum
            if order['created_date']:
                try:
                    created = datetime.strptime(str(order['created_date']), "%Y-%m-%d")
                    date_text = created.strftime("%d.%m.%Y")
                except:
                    date_text = str(order['created_date'])
            else:
                date_text = ""
            self.table.setItem(row, 1, QTableWidgetItem(date_text))

            # Číslo zakázky
            order_num_item = QTableWidgetItem(order['order_number'] or '')
            order_num_item.setForeground(QBrush(QColor("#3498db")))
            order_font = order_num_item.font()
            order_font.setUnderline(True)
            order_num_item.setFont(order_font)
            self.table.setItem(row, 2, order_num_item)

            # Typ
            self.table.setItem(row, 3, QTableWidgetItem(order['order_type'] or ''))

            # Stav km - potřebujeme načíst z order_work_log nebo jinak
            km_text = "-"
            # TODO: Implementovat načítání km ze zakázky
            self.table.setItem(row, 4, QTableWidgetItem(km_text))

            # Popis prací (z poznámky, zkráceně)
            note = order['note'] or ''
            if len(note) > 100:
                note = note[:97] + "..."
            self.table.setItem(row, 5, QTableWidgetItem(note))

            # Počet položek
            item_count = order['item_count'] or 0
            self.table.setItem(row, 6, QTableWidgetItem(str(item_count)))

            # Cena bez DPH
            subtotal = order['subtotal'] or 0
            self.table.setItem(row, 7, QTableWidgetItem(f"{subtotal:,.0f} Kč".replace(",", " ")))

            # Cena s DPH
            total = order['total_price'] or 0
            total_item = QTableWidgetItem(f"{total:,.0f} Kč".replace(",", " "))
            total_font = total_item.font()
            total_font.setBold(True)
            total_item.setFont(total_font)
            self.table.setItem(row, 8, total_item)

            # Stav s barvou
            status = order['status'] or ''
            status_item = QTableWidgetItem(status)
            status_color = config.ORDER_STATUS_COLORS.get(status, '#95a5a6')
            status_item.setBackground(QBrush(QColor(status_color)))
            status_item.setForeground(QBrush(QColor("white")))
            self.table.setItem(row, 9, status_item)

        self.table.setSortingEnabled(True)

    def update_statistics(self, orders):
        """Aktualizace statistik"""
        total_orders = len(orders)
        total_cost = sum(o['total_price'] or 0 for o in orders)
        avg_cost = total_cost / total_orders if total_orders > 0 else 0

        # Poslední servis
        last_service = "-"
        for o in orders:
            if o['completed_date']:
                try:
                    last_date = datetime.strptime(str(o['completed_date']), "%Y-%m-%d")
                    last_service = last_date.strftime("%d.%m.%Y")
                    break
                except:
                    pass

        self.stat_total_orders.setText(f"<b>Celkem zakázek:</b> {total_orders}")
        self.stat_total_cost.setText(f"<b>Celková útrata:</b> {total_cost:,.0f} Kč".replace(",", " "))
        self.stat_avg_cost.setText(f"<b>Průměrná zakázka:</b> {avg_cost:,.0f} Kč".replace(",", " "))
        self.stat_last_service.setText(f"<b>Poslední servis:</b> {last_service}")

    def apply_filters(self):
        """Aplikace filtrů"""
        period = self.period_filter.currentText()
        order_type = self.type_filter.currentText()

        today = date.today()
        filtered_orders = []

        for order in self.all_orders:
            # Filtr období
            if period != "Vše" and order['created_date']:
                try:
                    order_date = datetime.strptime(str(order['created_date']), "%Y-%m-%d").date()

                    if period == "Poslední rok":
                        cutoff = today - timedelta(days=365)
                        if order_date < cutoff:
                            continue
                    elif period == "Poslední 2 roky":
                        cutoff = today - timedelta(days=730)
                        if order_date < cutoff:
                            continue
                    elif period == "Poslední 3 roky":
                        cutoff = today - timedelta(days=1095)
                        if order_date < cutoff:
                            continue
                    elif period == "Aktuální rok":
                        if order_date.year != today.year:
                            continue
                except:
                    pass

            # Filtr typu
            if order_type != "Vše":
                if order['order_type'] != order_type:
                    continue

            filtered_orders.append(order)

        self.populate_table(filtered_orders)
        self.update_statistics(filtered_orders)

    def show_order_detail(self):
        """Zobrazení detailu zakázky"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return

        order_id = self.table.item(selected_row, 0).data(Qt.ItemDataRole.DisplayRole)

        dialog = OrderDetailDialog(order_id, self)
        dialog.exec()

    def show_detailed_statistics(self):
        """Zobrazení detailních statistik"""
        dialog = ServiceStatisticsDialog(self.all_orders, self)
        dialog.exec()

    def export_history(self):
        """Export historie do Excelu"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Exportovat historii",
                f"servisni_historie_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel soubory (*.xlsx)"
            )

            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Servisní historie"

            # Hlavička
            headers = ["Datum", "Číslo zakázky", "Typ", "Popis", "Počet položek", "Cena bez DPH", "Cena s DPH", "Stav"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")

            # Data
            for row_idx in range(self.table.rowCount()):
                if not self.table.isRowHidden(row_idx):
                    ws.cell(row=row_idx + 2, column=1, value=self.table.item(row_idx, 1).text())
                    ws.cell(row=row_idx + 2, column=2, value=self.table.item(row_idx, 2).text())
                    ws.cell(row=row_idx + 2, column=3, value=self.table.item(row_idx, 3).text())
                    ws.cell(row=row_idx + 2, column=4, value=self.table.item(row_idx, 5).text())
                    ws.cell(row=row_idx + 2, column=5, value=self.table.item(row_idx, 6).text())
                    ws.cell(row=row_idx + 2, column=6, value=self.table.item(row_idx, 7).text())
                    ws.cell(row=row_idx + 2, column=7, value=self.table.item(row_idx, 8).text())
                    ws.cell(row=row_idx + 2, column=8, value=self.table.item(row_idx, 9).text())

            wb.save(file_path)
            QMessageBox.information(self, "Export dokončen", f"Historie byla exportována do:\n{file_path}")

        except ImportError:
            QMessageBox.warning(self, "Chybí knihovna", "Pro export je potřeba nainstalovat openpyxl:\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se exportovat historii:\n{e}")


class OrderDetailDialog(QDialog):
    """Dialog s detailem zakázky"""

    def __init__(self, order_id, parent=None):
        super().__init__(parent)
        self.order_id = order_id
        self.init_ui()
        self.load_data()

    def init_ui(self):
        """Inicializace UI"""
        self.setWindowTitle("📋 Detail zakázky")
        self.setMinimumWidth(700)
        self.setMinimumHeight(500)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Info o zakázce
        self.info_group = QGroupBox("Informace o zakázce")
        self.info_layout = QFormLayout(self.info_group)
        layout.addWidget(self.info_group)

        # Položky zakázky
        items_group = QGroupBox("Položky zakázky")
        items_layout = QVBoxLayout(items_group)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(5)
        self.items_table.setHorizontalHeaderLabels(["Název", "Množství", "Jednotka", "Cena/ks", "Celkem"])
        self.items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.items_table.setAlternatingRowColors(True)

        header = self.items_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        items_layout.addWidget(self.items_table)
        layout.addWidget(items_group)

        # Poznámky
        notes_group = QGroupBox("Poznámky")
        notes_layout = QVBoxLayout(notes_group)

        self.notes_text = QTextEdit()
        self.notes_text.setReadOnly(True)
        self.notes_text.setMaximumHeight(100)
        notes_layout.addWidget(self.notes_text)

        layout.addWidget(notes_group)

        # Tlačítko zavřít
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_close = QPushButton("✅ Zavřít")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def load_data(self):
        """Načtení dat zakázky"""
        try:
            order = db.fetch_one("""
                SELECT * FROM orders WHERE id = ?
            """, (self.order_id,))

            if order:
                self.info_layout.addRow("Číslo zakázky:", QLabel(order['order_number'] or ''))
                self.info_layout.addRow("Typ:", QLabel(order['order_type'] or ''))
                self.info_layout.addRow("Stav:", QLabel(order['status'] or ''))

                if order['created_date']:
                    try:
                        created = datetime.strptime(str(order['created_date']), "%Y-%m-%d").strftime("%d.%m.%Y")
                    except:
                        created = str(order['created_date'])
                else:
                    created = '-'
                self.info_layout.addRow("Vytvořeno:", QLabel(created))

                if order['completed_date']:
                    try:
                        completed = datetime.strptime(str(order['completed_date']), "%Y-%m-%d").strftime("%d.%m.%Y")
                    except:
                        completed = str(order['completed_date'])
                else:
                    completed = '-'
                self.info_layout.addRow("Dokončeno:", QLabel(completed))

                total = order['total_price'] or 0
                self.info_layout.addRow("Celková cena:", QLabel(f"<b>{total:,.0f} Kč</b>".replace(",", " ")))

                self.notes_text.setPlainText(order['note'] or '')

            # Načtení položek
            items = db.fetch_all("""
                SELECT * FROM order_items WHERE order_id = ?
            """, (self.order_id,))

            self.items_table.setRowCount(0)
            for item in items:
                row = self.items_table.rowCount()
                self.items_table.insertRow(row)

                self.items_table.setItem(row, 0, QTableWidgetItem(item['name'] or ''))
                self.items_table.setItem(row, 1, QTableWidgetItem(str(item['quantity'] or 0)))
                self.items_table.setItem(row, 2, QTableWidgetItem(item['unit'] or 'ks'))

                unit_price = item['unit_price'] or 0
                self.items_table.setItem(row, 3, QTableWidgetItem(f"{unit_price:,.0f} Kč".replace(",", " ")))

                total = (item['quantity'] or 0) * unit_price
                self.items_table.setItem(row, 4, QTableWidgetItem(f"{total:,.0f} Kč".replace(",", " ")))

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se načíst detail zakázky:\n{e}")


class ServiceStatisticsDialog(QDialog):
    """Dialog s detailními statistikami"""

    def __init__(self, orders, parent=None):
        super().__init__(parent)
        self.orders = orders
        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        self.setWindowTitle("📊 Detailní statistiky")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Grid se statistikami
        stats_grid = QGridLayout()
        stats_grid.setSpacing(15)

        # Výpočet statistik
        total_orders = len(self.orders)
        total_cost = sum(o['total_price'] or 0 for o in self.orders)
        avg_cost = total_cost / total_orders if total_orders > 0 else 0

        # Nejdražší zakázka
        max_cost = max((o['total_price'] or 0 for o in self.orders), default=0)

        # Nejlevnější zakázka
        min_cost = min((o['total_price'] or 0 for o in self.orders if o['total_price']), default=0)

        # Typ zakázek
        order_types = {}
        for o in self.orders:
            otype = o['order_type'] or 'Jiný'
            order_types[otype] = order_types.get(otype, 0) + 1

        # Roční statistiky
        yearly_stats = {}
        for o in self.orders:
            if o['created_date']:
                try:
                    year = datetime.strptime(str(o['created_date']), "%Y-%m-%d").year
                    if year not in yearly_stats:
                        yearly_stats[year] = {'count': 0, 'total': 0}
                    yearly_stats[year]['count'] += 1
                    yearly_stats[year]['total'] += o['total_price'] or 0
                except:
                    pass

        # Zobrazení
        stats_grid.addWidget(self.create_stat_card("📋 Celkem zakázek", str(total_orders), "#3498db"), 0, 0)
        stats_grid.addWidget(self.create_stat_card("💰 Celková útrata", f"{total_cost:,.0f} Kč".replace(",", " "), "#27ae60"), 0, 1)
        stats_grid.addWidget(self.create_stat_card("📊 Průměrná zakázka", f"{avg_cost:,.0f} Kč".replace(",", " "), "#f39c12"), 0, 2)
        stats_grid.addWidget(self.create_stat_card("💎 Nejdražší zakázka", f"{max_cost:,.0f} Kč".replace(",", " "), "#e74c3c"), 1, 0)
        stats_grid.addWidget(self.create_stat_card("🪙 Nejlevnější zakázka", f"{min_cost:,.0f} Kč".replace(",", " "), "#9b59b6"), 1, 1)

        # Nejčastější typ
        if order_types:
            most_common = max(order_types.items(), key=lambda x: x[1])
            stats_grid.addWidget(self.create_stat_card("📈 Nejčastější typ", f"{most_common[0]} ({most_common[1]}x)", "#1abc9c"), 1, 2)

        layout.addLayout(stats_grid)

        # Roční přehled
        if yearly_stats:
            yearly_group = QGroupBox("📅 Roční přehled")
            yearly_layout = QVBoxLayout(yearly_group)

            for year in sorted(yearly_stats.keys(), reverse=True):
                stats = yearly_stats[year]
                year_label = QLabel(f"<b>{year}:</b> {stats['count']} zakázek, {stats['total']:,.0f} Kč".replace(",", " "))
                yearly_layout.addWidget(year_label)

            layout.addWidget(yearly_group)

        # Tlačítko zavřít
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_close = QPushButton("✅ Zavřít")
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)

        layout.addLayout(btn_layout)

    def create_stat_card(self, label, value, color):
        """Vytvoření karty se statistikou"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 10px;
                padding: 15px;
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setSpacing(5)

        value_label = QLabel(value)
        value_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 16px;
                font-weight: bold;
            }
        """)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        text_label = QLabel(label)
        text_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 10px;
            }
        """)
        text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout.addWidget(value_label)
        layout.addWidget(text_label)

        return card
