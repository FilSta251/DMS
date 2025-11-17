# -*- coding: utf-8 -*-
"""
Modul Vozidla - Hlavní seznam všech motorek s filtry a statistikami
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QLabel, QComboBox, QMessageBox, QHeaderView,
    QFrame, QGridLayout, QMenu, QApplication, QFileDialog
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QColor, QBrush, QAction, QCursor
from datetime import datetime, date, timedelta
import config
from database_manager import db


class VehiclesModule(QWidget):
    """Hlavní modul pro správu vozidel/motorek"""

    vehicle_selected = pyqtSignal(int)  # Signal pro komunikaci s jinými moduly

    def __init__(self):
        super().__init__()
        self.current_filters = {}
        self.init_ui()

    def init_ui(self):
        """Inicializace uživatelského rozhraní"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Horní panel s tlačítky
        header_panel = self.create_header_panel()
        layout.addWidget(header_panel)

        # Statistiky - rychlý přehled
        stats_panel = self.create_stats_panel()
        layout.addWidget(stats_panel)

        # Filtry
        filters_panel = self.create_filters_panel()
        layout.addWidget(filters_panel)

        # Tabulka vozidel
        self.table = self.create_table()
        layout.addWidget(self.table)

        # Načtení dat
        self.refresh()

    def create_header_panel(self):
        """Vytvoření horního panelu s tlačítky"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                padding: 10px;
            }
        """)

        layout = QHBoxLayout(panel)
        layout.setContentsMargins(15, 10, 15, 10)

        # Titulek
        title = QLabel("🏍️ Vozidla")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        layout.addStretch()

        # Tlačítka
        btn_new = QPushButton("➕ Nové vozidlo")
        btn_new.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: #219a52;
            }}
        """)
        btn_new.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_new.clicked.connect(self.add_vehicle)
        layout.addWidget(btn_new)

        btn_import = QPushButton("📥 Import")
        btn_import.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SECONDARY};
                color: white;
                padding: 10px 15px;
                border-radius: 5px;
            }}
            QPushButton:hover {{
                background-color: #2980b9;
            }}
        """)
        btn_import.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_import.clicked.connect(self.import_vehicles)
        layout.addWidget(btn_import)

        btn_export = QPushButton("📤 Export")
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 10px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export.clicked.connect(self.export_vehicles)
        layout.addWidget(btn_export)

        return panel

    def create_stats_panel(self):
        """Vytvoření panelu se statistikami"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                padding: 15px;
            }
        """)

        layout = QGridLayout(panel)
        layout.setSpacing(20)

        # Statistiky - budou aktualizovány v refresh()
        self.stat_total = self.create_stat_card("Celkem vozidel", "0", "#3498db")
        self.stat_active_customers = self.create_stat_card("Aktivních zákazníků", "0", "#27ae60")
        self.stat_stk_valid = self.create_stat_card("Platná STK", "0", "#2ecc71")
        self.stat_stk_warning = self.create_stat_card("STK expiruje < 30 dní", "0", "#f39c12")
        self.stat_no_customer = self.create_stat_card("Bez zákazníka", "0", "#e74c3c")

        layout.addWidget(self.stat_total, 0, 0)
        layout.addWidget(self.stat_active_customers, 0, 1)
        layout.addWidget(self.stat_stk_valid, 0, 2)
        layout.addWidget(self.stat_stk_warning, 0, 3)
        layout.addWidget(self.stat_no_customer, 0, 4)

        return panel

    def create_stat_card(self, label, value, color):
        """Vytvoření karty se statistikou"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 10px;
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(5)

        value_label = QLabel(value)
        value_label.setObjectName("stat_value")
        value_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
        """)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        text_label = QLabel(label)
        text_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 11px;
            }
        """)
        text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        text_label.setWordWrap(True)

        layout.addWidget(value_label)
        layout.addWidget(text_label)

        return card

    def create_filters_panel(self):
        """Vytvoření panelu s filtry"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                padding: 15px;
            }
        """)

        layout = QHBoxLayout(panel)
        layout.setSpacing(15)

        # Vyhledávání
        search_label = QLabel("🔍")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("SPZ, značka, model, VIN, zákazník...")
        self.search_input.setMinimumWidth(250)
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #ddd;
                border-radius: 5px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        self.search_input.textChanged.connect(self.apply_filters)

        layout.addWidget(search_label)
        layout.addWidget(self.search_input)

        # Filtr značky
        brand_label = QLabel("Značka:")
        self.brand_filter = QComboBox()
        self.brand_filter.setMinimumWidth(150)
        self.brand_filter.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #ddd;
                border-radius: 5px;
            }
        """)
        self.brand_filter.currentTextChanged.connect(self.apply_filters)

        layout.addWidget(brand_label)
        layout.addWidget(self.brand_filter)

        # Filtr STK
        stk_label = QLabel("Stav STK:")
        self.stk_filter = QComboBox()
        self.stk_filter.addItems(["Vše", "Platná", "Expiruje brzy", "Neplatná", "Neuvedeno"])
        self.stk_filter.setMinimumWidth(130)
        self.stk_filter.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #ddd;
                border-radius: 5px;
            }
        """)
        self.stk_filter.currentTextChanged.connect(self.apply_filters)

        layout.addWidget(stk_label)
        layout.addWidget(self.stk_filter)

        # Filtr zákazníka
        customer_label = QLabel("Zákazník:")
        self.customer_filter = QComboBox()
        self.customer_filter.addItems(["Vše", "S přiřazeným", "Bez zákazníka"])
        self.customer_filter.setMinimumWidth(130)
        self.customer_filter.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #ddd;
                border-radius: 5px;
            }
        """)
        self.customer_filter.currentTextChanged.connect(self.apply_filters)

        layout.addWidget(customer_label)
        layout.addWidget(self.customer_filter)

        layout.addStretch()

        # Tlačítko reset filtrů
        btn_reset = QPushButton("🔄 Reset filtrů")
        btn_reset.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_reset.clicked.connect(self.reset_filters)
        layout.addWidget(btn_reset)

        return panel

    def create_table(self):
        """Vytvoření tabulky s vozidly"""
        table = QTableWidget()
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ecf0f1;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QTableWidget::item:hover {
                background-color: #ebf5fb;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 12px;
                font-weight: bold;
                border: none;
                font-size: 12px;
            }
        """)

        # Sloupce
        columns = [
            "ID", "SPZ", "Značka", "Model", "Rok", "Zákazník",
            "Telefon", "Stav km", "STK platná do", "Poslední servis", "Poznámka"
        ]
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)

        # Nastavení chování
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(True)
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(self.show_context_menu)

        # Roztažení sloupců
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # SPZ
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)           # Značka
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)           # Model
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Rok
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)           # Zákazník
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Telefon
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)  # km
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)  # STK
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)  # Servis
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)          # Poznámka

        # Skrytí ID sloupce
        table.setColumnHidden(0, True)

        # Dvojklik pro detail
        table.doubleClicked.connect(self.show_detail)

        return table

    def refresh(self):
        """Obnovení dat v tabulce a statistik"""
        try:
            # Zajištění migrace databáze pro nové sloupce
            self.ensure_vehicle_columns()

            # Načtení vozidel z databáze
            vehicles = db.fetch_all("""
                SELECT
                    v.id,
                    v.license_plate,
                    v.brand,
                    v.model,
                    v.year,
                    v.vin,
                    v.customer_id,
                    v.mileage,
                    v.stk_valid_until,
                    v.notes,
                    c.first_name || ' ' || c.last_name as customer_name,
                    c.phone,
                    (SELECT MAX(o.completed_date)
                     FROM orders o
                     WHERE o.vehicle_id = v.id AND o.status = 'Dokončená') as last_service
                FROM vehicles v
                LEFT JOIN customers c ON v.customer_id = c.id
                ORDER BY v.license_plate
            """)

            # Aktualizace filtru značek
            brands = set()
            for v in vehicles:
                if v['brand']:
                    brands.add(v['brand'])

            current_brand = self.brand_filter.currentText()
            self.brand_filter.blockSignals(True)
            self.brand_filter.clear()
            self.brand_filter.addItem("Vše")
            for brand in sorted(brands):
                self.brand_filter.addItem(brand)
            if current_brand in ["Vše"] + list(brands):
                self.brand_filter.setCurrentText(current_brand)
            self.brand_filter.blockSignals(False)

            # Uložení dat pro filtrování
            self.all_vehicles = vehicles

            # Aktualizace statistik
            self.update_statistics(vehicles)

            # Naplnění tabulky
            self.populate_table(vehicles)

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se načíst vozidla:\n{e}")

    def ensure_vehicle_columns(self):
        """Zajištění existence nových sloupců v tabulce vehicles"""
        try:
            # Přidání sloupce stk_valid_until pokud neexistuje
            db.cursor.execute("PRAGMA table_info(vehicles)")
            columns = [col[1] for col in db.cursor.fetchall()]

            if 'stk_valid_until' not in columns:
                db.cursor.execute("ALTER TABLE vehicles ADD COLUMN stk_valid_until DATE")
                db.connection.commit()
                print("✅ Přidán sloupec stk_valid_until do vehicles")
        except Exception as e:
            print(f"Chyba při kontrole sloupců: {e}")

    def update_statistics(self, vehicles):
        """Aktualizace statistik"""
        total = len(vehicles)
        active_customers = len(set(v['customer_id'] for v in vehicles if v['customer_id']))
        no_customer = sum(1 for v in vehicles if not v['customer_id'])

        today = date.today()
        warning_date = today + timedelta(days=30)

        stk_valid = 0
        stk_warning = 0

        for v in vehicles:
            if v['stk_valid_until']:
                try:
                    stk_date = datetime.strptime(str(v['stk_valid_until']), "%Y-%m-%d").date()
                    if stk_date > warning_date:
                        stk_valid += 1
                    elif stk_date >= today:
                        stk_warning += 1
                except:
                    pass

        # Aktualizace hodnot v kartách
        self.stat_total.findChild(QLabel, "stat_value").setText(str(total))
        self.stat_active_customers.findChild(QLabel, "stat_value").setText(str(active_customers))
        self.stat_stk_valid.findChild(QLabel, "stat_value").setText(str(stk_valid))
        self.stat_stk_warning.findChild(QLabel, "stat_value").setText(str(stk_warning))
        self.stat_no_customer.findChild(QLabel, "stat_value").setText(str(no_customer))

    def populate_table(self, vehicles):
        """Naplnění tabulky daty"""
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)

        today = date.today()
        warning_date = today + timedelta(days=30)

        for vehicle in vehicles:
            row = self.table.rowCount()
            self.table.insertRow(row)

            # ID
            id_item = QTableWidgetItem()
            id_item.setData(Qt.ItemDataRole.DisplayRole, vehicle['id'])
            self.table.setItem(row, 0, id_item)

            # SPZ
            spz_item = QTableWidgetItem(vehicle['license_plate'] or '')
            spz_font = spz_item.font()
            spz_font.setBold(True)
            spz_item.setFont(spz_font)
            self.table.setItem(row, 1, spz_item)

            # Značka
            self.table.setItem(row, 2, QTableWidgetItem(vehicle['brand'] or ''))

            # Model
            self.table.setItem(row, 3, QTableWidgetItem(vehicle['model'] or ''))

            # Rok
            year_item = QTableWidgetItem()
            if vehicle['year']:
                year_item.setData(Qt.ItemDataRole.DisplayRole, vehicle['year'])
            self.table.setItem(row, 4, year_item)

            # Zákazník
            customer_name = vehicle['customer_name'] or 'Bez zákazníka'
            customer_item = QTableWidgetItem(customer_name)
            if not vehicle['customer_id']:
                customer_item.setForeground(QBrush(QColor("#e74c3c")))
            self.table.setItem(row, 5, customer_item)

            # Telefon
            self.table.setItem(row, 6, QTableWidgetItem(vehicle['phone'] or ''))

            # Stav km
            km_item = QTableWidgetItem()
            if vehicle['mileage']:
                km_item.setText(f"{vehicle['mileage']:,} km".replace(",", " "))
            self.table.setItem(row, 7, km_item)

            # STK platná do - s barevným pozadím
            stk_item = QTableWidgetItem()
            if vehicle['stk_valid_until']:
                try:
                    stk_date = datetime.strptime(str(vehicle['stk_valid_until']), "%Y-%m-%d").date()
                    stk_item.setText(stk_date.strftime("%d.%m.%Y"))

                    # Barevné pozadí podle stavu
                    if stk_date < today:
                        stk_item.setBackground(QBrush(QColor("#e74c3c")))  # Červená - neplatná
                        stk_item.setForeground(QBrush(QColor("white")))
                    elif stk_date <= warning_date:
                        stk_item.setBackground(QBrush(QColor("#f39c12")))  # Oranžová - expiruje brzy
                        stk_item.setForeground(QBrush(QColor("white")))
                    else:
                        stk_item.setBackground(QBrush(QColor("#27ae60")))  # Zelená - platná
                        stk_item.setForeground(QBrush(QColor("white")))
                except:
                    stk_item.setText("Neplatné datum")
            else:
                stk_item.setText("Neuvedeno")
                stk_item.setForeground(QBrush(QColor("#95a5a6")))
            self.table.setItem(row, 8, stk_item)

            # Poslední servis
            last_service_item = QTableWidgetItem()
            if vehicle['last_service']:
                try:
                    service_date = datetime.strptime(str(vehicle['last_service']), "%Y-%m-%d").date()
                    last_service_item.setText(service_date.strftime("%d.%m.%Y"))
                except:
                    last_service_item.setText(str(vehicle['last_service']))
            self.table.setItem(row, 9, last_service_item)

            # Poznámka (zkrácená)
            notes = vehicle['notes'] or ''
            if len(notes) > 50:
                notes = notes[:47] + "..."
            self.table.setItem(row, 10, QTableWidgetItem(notes))

        self.table.setSortingEnabled(True)

    def apply_filters(self):
        """Aplikace filtrů na tabulku"""
        search_text = self.search_input.text().lower()
        brand_filter = self.brand_filter.currentText()
        stk_filter = self.stk_filter.currentText()
        customer_filter = self.customer_filter.currentText()

        today = date.today()
        warning_date = today + timedelta(days=30)

        for row in range(self.table.rowCount()):
            show_row = True

            # Textové vyhledávání
            if search_text:
                row_matches = False
                for col in range(1, self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item and search_text in item.text().lower():
                        row_matches = True
                        break
                if not row_matches:
                    show_row = False

            # Filtr značky
            if show_row and brand_filter != "Vše":
                brand_item = self.table.item(row, 2)
                if not brand_item or brand_item.text() != brand_filter:
                    show_row = False

            # Filtr STK
            if show_row and stk_filter != "Vše":
                stk_item = self.table.item(row, 8)
                if stk_item:
                    stk_text = stk_item.text()
                    if stk_filter == "Neuvedeno":
                        if stk_text != "Neuvedeno":
                            show_row = False
                    elif stk_text == "Neuvedeno":
                        show_row = False
                    else:
                        try:
                            stk_date = datetime.strptime(stk_text, "%d.%m.%Y").date()
                            if stk_filter == "Platná":
                                if stk_date <= warning_date:
                                    show_row = False
                            elif stk_filter == "Expiruje brzy":
                                if not (today <= stk_date <= warning_date):
                                    show_row = False
                            elif stk_filter == "Neplatná":
                                if stk_date >= today:
                                    show_row = False
                        except:
                            show_row = False

            # Filtr zákazníka
            if show_row and customer_filter != "Vše":
                customer_item = self.table.item(row, 5)
                if customer_item:
                    has_customer = customer_item.text() != "Bez zákazníka"
                    if customer_filter == "S přiřazeným" and not has_customer:
                        show_row = False
                    elif customer_filter == "Bez zákazníka" and has_customer:
                        show_row = False

            self.table.setRowHidden(row, not show_row)

    def reset_filters(self):
        """Reset všech filtrů"""
        self.search_input.clear()
        self.brand_filter.setCurrentIndex(0)
        self.stk_filter.setCurrentIndex(0)
        self.customer_filter.setCurrentIndex(0)

        for row in range(self.table.rowCount()):
            self.table.setRowHidden(row, False)

    def show_context_menu(self, position):
        """Zobrazení kontextového menu"""
        if self.table.currentRow() < 0:
            return

        menu = QMenu(self)

        action_detail = QAction("📖 Otevřít detail", self)
        action_detail.triggered.connect(self.show_detail)
        menu.addAction(action_detail)

        action_edit = QAction("✏️ Upravit vozidlo", self)
        action_edit.triggered.connect(self.edit_vehicle)
        menu.addAction(action_edit)

        menu.addSeparator()

        action_copy_spz = QAction("📋 Kopírovat SPZ", self)
        action_copy_spz.triggered.connect(self.copy_spz)
        menu.addAction(action_copy_spz)

        action_copy_vin = QAction("📋 Kopírovat VIN", self)
        action_copy_vin.triggered.connect(self.copy_vin)
        menu.addAction(action_copy_vin)

        menu.addSeparator()

        action_customer = QAction("👤 Přejít na zákazníka", self)
        action_customer.triggered.connect(self.go_to_customer)
        menu.addAction(action_customer)

        action_orders = QAction("📋 Zobrazit zakázky", self)
        action_orders.triggered.connect(self.show_orders)
        menu.addAction(action_orders)

        action_schedule = QAction("🗓️ Naplánovat servis", self)
        action_schedule.triggered.connect(self.schedule_service)
        menu.addAction(action_schedule)

        menu.addSeparator()

        action_delete = QAction("🗑️ Smazat vozidlo", self)
        action_delete.triggered.connect(self.delete_vehicle)
        menu.addAction(action_delete)

        menu.exec(QCursor.pos())

    def get_selected_vehicle_id(self):
        """Získání ID vybraného vozidla"""
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return None
        return int(self.table.item(selected_row, 0).data(Qt.ItemDataRole.DisplayRole))

    def add_vehicle(self):
        """Přidání nového vozidla"""
        from .vehicle_form import VehicleFormDialog

        dialog = VehicleFormDialog(self)
        if dialog.exec():
            self.refresh()
            QMessageBox.information(self, "Úspěch", "Vozidlo bylo úspěšně přidáno.")

    def edit_vehicle(self):
        """Editace vybraného vozidla"""
        vehicle_id = self.get_selected_vehicle_id()
        if not vehicle_id:
            QMessageBox.warning(self, "Upozornění", "Vyberte vozidlo k úpravě.")
            return

        from .vehicle_form import VehicleFormDialog

        dialog = VehicleFormDialog(self, vehicle_id=vehicle_id)
        if dialog.exec():
            self.refresh()
            QMessageBox.information(self, "Úspěch", "Vozidlo bylo úspěšně upraveno.")

    def show_detail(self):
        """Zobrazení detailu vozidla"""
        vehicle_id = self.get_selected_vehicle_id()
        if not vehicle_id:
            QMessageBox.warning(self, "Upozornění", "Vyberte vozidlo pro zobrazení detailu.")
            return

        from .vehicle_detail import VehicleDetailWindow

        dialog = VehicleDetailWindow(self, vehicle_id=vehicle_id)
        dialog.exec()
        self.refresh()

    def delete_vehicle(self):
        """Smazání vozidla"""
        vehicle_id = self.get_selected_vehicle_id()
        if not vehicle_id:
            QMessageBox.warning(self, "Upozornění", "Vyberte vozidlo ke smazání.")
            return

        # Kontrola zakázek
        orders = db.fetch_all(
            "SELECT COUNT(*) as count FROM orders WHERE vehicle_id = ?",
            (vehicle_id,)
        )

        if orders and orders[0]['count'] > 0:
            reply = QMessageBox.question(
                self,
                "Varování",
                f"Vozidlo má {orders[0]['count']} přiřazených zakázek.\n"
                "Opravdu chcete vozidlo smazat?\n\n"
                "Zakázky zůstanou zachovány, ale ztratí odkaz na vozidlo.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
        else:
            reply = QMessageBox.question(
                self,
                "Potvrzení",
                "Opravdu chcete smazat vybrané vozidlo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                db.execute_query("DELETE FROM vehicles WHERE id = ?", (vehicle_id,))
                self.refresh()
                QMessageBox.information(self, "Úspěch", "Vozidlo bylo smazáno.")
            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se smazat vozidlo:\n{e}")

    def copy_spz(self):
        """Kopírování SPZ do schránky"""
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            spz = self.table.item(selected_row, 1).text()
            QApplication.clipboard().setText(spz)
            QMessageBox.information(self, "Zkopírováno", f"SPZ '{spz}' byla zkopírována do schránky.")

    def copy_vin(self):
        """Kopírování VIN do schránky"""
        vehicle_id = self.get_selected_vehicle_id()
        if vehicle_id:
            vehicle = db.fetch_one("SELECT vin FROM vehicles WHERE id = ?", (vehicle_id,))
            if vehicle and vehicle['vin']:
                QApplication.clipboard().setText(vehicle['vin'])
                QMessageBox.information(self, "Zkopírováno", f"VIN '{vehicle['vin']}' byl zkopírován do schránky.")
            else:
                QMessageBox.warning(self, "Upozornění", "Vozidlo nemá zadaný VIN.")

    def go_to_customer(self):
        """Přechod na zákazníka"""
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            customer_name = self.table.item(selected_row, 5).text()
            if customer_name == "Bez zákazníka":
                QMessageBox.warning(self, "Upozornění", "Vozidlo nemá přiřazeného zákazníka.")
                return

            main_window = self.window()
            if hasattr(main_window, 'switch_module'):
                main_window.switch_module('customers')

    def show_orders(self):
        """Zobrazení zakázek pro vozidlo"""
        vehicle_id = self.get_selected_vehicle_id()
        if vehicle_id:
            main_window = self.window()
            if hasattr(main_window, 'switch_module'):
                main_window.switch_module('orders')
                # Zde by se mohl předat filtr na orders modul

    def schedule_service(self):
        """Naplánování servisu"""
        vehicle_id = self.get_selected_vehicle_id()
        if vehicle_id:
            main_window = self.window()
            if hasattr(main_window, 'switch_module'):
                main_window.switch_module('calendar')
                # Zde by se mohlo otevřít okno pro naplánování

    def import_vehicles(self):
        """Import vozidel z CSV"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Importovat vozidla",
            "",
            "CSV soubory (*.csv);;Všechny soubory (*.*)"
        )

        if file_path:
            QMessageBox.information(
                self,
                "Import",
                "Funkce importu bude implementována v příští verzi."
            )

    def export_vehicles(self):
        """Export vozidel do Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Exportovat vozidla",
                f"vozidla_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel soubory (*.xlsx)"
            )

            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vozidla"

            # Hlavička
            headers = ["SPZ", "Značka", "Model", "Rok", "VIN", "Zákazník", "Telefon", "Stav km", "STK platná do", "Poznámka"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row_idx in range(self.table.rowCount()):
                if not self.table.isRowHidden(row_idx):
                    ws.cell(row=row_idx + 2, column=1, value=self.table.item(row_idx, 1).text())
                    ws.cell(row=row_idx + 2, column=2, value=self.table.item(row_idx, 2).text())
                    ws.cell(row=row_idx + 2, column=3, value=self.table.item(row_idx, 3).text())
                    ws.cell(row=row_idx + 2, column=4, value=self.table.item(row_idx, 4).text())

                    # VIN z databáze
                    vehicle_id = self.table.item(row_idx, 0).data(Qt.ItemDataRole.DisplayRole)
                    vehicle = db.fetch_one("SELECT vin FROM vehicles WHERE id = ?", (vehicle_id,))
                    ws.cell(row=row_idx + 2, column=5, value=vehicle['vin'] if vehicle else '')

                    ws.cell(row=row_idx + 2, column=6, value=self.table.item(row_idx, 5).text())
                    ws.cell(row=row_idx + 2, column=7, value=self.table.item(row_idx, 6).text())
                    ws.cell(row=row_idx + 2, column=8, value=self.table.item(row_idx, 7).text())
                    ws.cell(row=row_idx + 2, column=9, value=self.table.item(row_idx, 8).text())
                    ws.cell(row=row_idx + 2, column=10, value=self.table.item(row_idx, 10).text())

            # Šířka sloupců
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            wb.save(file_path)
            QMessageBox.information(self, "Export dokončen", f"Vozidla byla exportována do:\n{file_path}")

        except ImportError:
            QMessageBox.warning(
                self,
                "Chybí knihovna",
                "Pro export do Excelu je potřeba nainstalovat knihovnu openpyxl:\n\n"
                "pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Chyba exportu", f"Nepodařilo se exportovat vozidla:\n{e}")
