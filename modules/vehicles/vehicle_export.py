# -*- coding: utf-8 -*-
"""
Export dat vozidel - PDF karty, servisní knihy, Excel seznamy, kompletní ZIP archivy
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QComboBox,
    QFrame, QMessageBox, QDialog, QFormLayout, QCheckBox, QProgressBar,
    QFileDialog, QGroupBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from datetime import datetime
from pathlib import Path
import json
import zipfile
import config
from database_manager import db


class VehicleExporter:
    """Třída pro export dat vozidel"""

    def __init__(self, vehicle_id=None):
        self.vehicle_id = vehicle_id
        self.vehicle_data = None
        if vehicle_id:
            self.load_vehicle_data()

    def load_vehicle_data(self):
        """Načtení dat vozidla"""
        try:
            self.vehicle_data = db.fetch_one("""
                SELECT v.*,
                       c.first_name || ' ' || c.last_name as customer_name,
                       c.phone as customer_phone,
                       c.email as customer_email,
                       c.company as customer_company,
                       c.address as customer_address,
                       c.city as customer_city,
                       c.postal_code as customer_postal_code
                FROM vehicles v
                LEFT JOIN customers c ON v.customer_id = c.id
                WHERE v.id = ?
            """, (self.vehicle_id,))
        except Exception as e:
            print(f"Chyba při načítání dat vozidla: {e}")

    def export_vehicle_card(self, output_path):
        """Export karty vozidla do PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            if not self.vehicle_data:
                return False, "Data vozidla nejsou k dispozici"

            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            styles = getSampleStyleSheet()

            # Vlastní styly
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#2c3e50')
            )

            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading2'],
                fontSize=14,
                spaceBefore=20,
                spaceAfter=10,
                textColor=colors.HexColor('#3498db')
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=6
            )

            elements = []

            # Hlavička
            elements.append(Paragraph("KARTA VOZIDLA", title_style))
            elements.append(Paragraph(
                f"Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                ParagraphStyle('Date', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)
            ))
            elements.append(Spacer(1, 20))

            # Základní údaje
            elements.append(Paragraph("🏍️ ZÁKLADNÍ ÚDAJE", section_style))

            basic_data = [
                ['SPZ:', self.vehicle_data['license_plate'] or '-'],
                ['Značka:', self.vehicle_data['brand'] or '-'],
                ['Model:', self.vehicle_data['model'] or '-'],
                ['Rok výroby:', str(self.vehicle_data['year']) if self.vehicle_data['year'] else '-'],
                ['VIN:', self.vehicle_data['vin'] or '-'],
                ['Barva:', self.vehicle_data['color'] or '-'],
            ]

            basic_table = Table(basic_data, colWidths=[4*cm, 12*cm])
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2c3e50')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(basic_table)
            elements.append(Spacer(1, 15))

            # Technické údaje
            elements.append(Paragraph("⚙️ TECHNICKÉ ÚDAJE", section_style))

            tech_data = [
                ['Typ motoru:', self.vehicle_data['engine_type'] or '-'],
                ['Palivo:', self.vehicle_data['fuel_type'] or '-'],
                ['Stav km:', f"{self.vehicle_data['mileage']:,} km".replace(",", " ") if self.vehicle_data['mileage'] else '-'],
            ]

            # STK s barvou
            stk_text = '-'
            if self.vehicle_data['stk_valid_until']:
                try:
                    stk_date = datetime.strptime(str(self.vehicle_data['stk_valid_until']), "%Y-%m-%d")
                    stk_text = stk_date.strftime("%d.%m.%Y")
                except:
                    pass
            tech_data.append(['STK platná do:', stk_text])

            tech_table = Table(tech_data, colWidths=[4*cm, 12*cm])
            tech_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2c3e50')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(tech_table)
            elements.append(Spacer(1, 15))

            # Majitel
            elements.append(Paragraph("👤 MAJITEL VOZIDLA", section_style))

            if self.vehicle_data['customer_id']:
                owner_data = [
                    ['Jméno:', self.vehicle_data['customer_name'] or '-'],
                    ['Firma:', self.vehicle_data['customer_company'] or '-'],
                    ['Telefon:', self.vehicle_data['customer_phone'] or '-'],
                    ['Email:', self.vehicle_data['customer_email'] or '-'],
                    ['Adresa:', f"{self.vehicle_data['customer_address'] or ''}, {self.vehicle_data['customer_city'] or ''} {self.vehicle_data['customer_postal_code'] or ''}".strip(', ')],
                ]
            else:
                owner_data = [['', 'Vozidlo nemá přiřazeného majitele']]

            owner_table = Table(owner_data, colWidths=[4*cm, 12*cm])
            owner_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2c3e50')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(owner_table)
            elements.append(Spacer(1, 15))

            # Poznámky
            if self.vehicle_data['notes']:
                elements.append(Paragraph("📝 POZNÁMKY", section_style))
                elements.append(Paragraph(self.vehicle_data['notes'], normal_style))

            # Patička
            elements.append(Spacer(1, 30))
            elements.append(Paragraph(
                f"Vytvořeno v systému {config.APP_NAME} v{config.APP_VERSION}",
                ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray)
            ))

            doc.build(elements)
            return True, "Karta vozidla byla exportována"

        except ImportError:
            return False, "Pro export PDF je potřeba nainstalovat reportlab:\npip install reportlab"
        except Exception as e:
            return False, f"Chyba při exportu: {e}"

    def export_service_book(self, output_path):
        """Export servisní knihy do PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.enums import TA_CENTER

            if not self.vehicle_data:
                return False, "Data vozidla nejsou k dispozici"

            # Načtení historie servisů
            orders = db.fetch_all("""
                SELECT
                    o.order_number,
                    o.order_type,
                    o.created_date,
                    o.completed_date,
                    o.total_price,
                    o.note,
                    (SELECT GROUP_CONCAT(oi.name, ', ') FROM order_items oi WHERE oi.order_id = o.id) as items
                FROM orders o
                WHERE o.vehicle_id = ?
                ORDER BY o.created_date DESC
            """, (self.vehicle_id,))

            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#2c3e50')
            )

            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading2'],
                fontSize=12,
                spaceBefore=15,
                spaceAfter=10,
                textColor=colors.HexColor('#3498db')
            )

            elements = []

            # Hlavička
            elements.append(Paragraph("SERVISNÍ KNIHA", title_style))
            elements.append(Paragraph(
                f"{self.vehicle_data['brand']} {self.vehicle_data['model']} - {self.vehicle_data['license_plate']}",
                ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=14, alignment=TA_CENTER, spaceAfter=20)
            ))

            # Shrnutí
            elements.append(Paragraph("📊 SHRNUTÍ", section_style))

            total_orders = len(orders)
            total_spent = sum(o['total_price'] or 0 for o in orders)

            summary_data = [
                ['Celkem servisů:', str(total_orders)],
                ['Celková útrata:', f"{total_spent:,.0f} Kč".replace(",", " ")],
                ['Aktuální stav km:', f"{self.vehicle_data['mileage']:,} km".replace(",", " ") if self.vehicle_data['mileage'] else '-'],
            ]

            summary_table = Table(summary_data, colWidths=[5*cm, 10*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            # Historie servisů
            elements.append(Paragraph("🔧 HISTORIE SERVISŮ", section_style))

            if orders:
                for i, order in enumerate(orders, 1):
                    # Hlavička servisu
                    order_date = '-'
                    if order['created_date']:
                        try:
                            order_date = datetime.strptime(str(order['created_date']), "%Y-%m-%d").strftime("%d.%m.%Y")
                        except:
                            pass

                    elements.append(Paragraph(
                        f"<b>Servis #{i}</b> - {order_date} | {order['order_number']} | {order['order_type']}",
                        ParagraphStyle('OrderHeader', parent=styles['Normal'], fontSize=11, spaceAfter=5, textColor=colors.HexColor('#34495e'))
                    ))

                    # Detaily
                    service_data = []

                    if order['note']:
                        service_data.append(['Popis:', order['note'][:200]])

                    if order['items']:
                        service_data.append(['Položky:', order['items'][:200]])

                    service_data.append(['Cena:', f"{order['total_price'] or 0:,.0f} Kč".replace(",", " ")])

                    if service_data:
                        service_table = Table(service_data, colWidths=[3*cm, 14*cm])
                        service_table.setStyle(TableStyle([
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('PADDING', (0, 0), (-1, -1), 4),
                            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
                        ]))
                        elements.append(service_table)

                    if i < len(orders):
                        elements.append(Spacer(1, 10))
            else:
                elements.append(Paragraph("Zatím nebyly provedeny žádné servisy.", styles['Normal']))

            # Patička
            elements.append(Spacer(1, 30))
            elements.append(Paragraph(
                f"Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')} | {config.APP_NAME}",
                ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray)
            ))

            doc.build(elements)
            return True, "Servisní kniha byla exportována"

        except ImportError:
            return False, "Pro export PDF je potřeba nainstalovat reportlab:\npip install reportlab"
        except Exception as e:
            return False, f"Chyba při exportu: {e}"

    def export_vehicle_list(self, output_path, vehicle_ids=None):
        """Export seznamu vozidel do Excelu"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Načtení dat
            if vehicle_ids:
                placeholders = ','.join(['?' for _ in vehicle_ids])
                vehicles = db.fetch_all(f"""
                    SELECT v.*,
                           c.first_name || ' ' || c.last_name as customer_name,
                           c.phone as customer_phone
                    FROM vehicles v
                    LEFT JOIN customers c ON v.customer_id = c.id
                    WHERE v.id IN ({placeholders})
                    ORDER BY v.license_plate
                """, tuple(vehicle_ids))
            else:
                vehicles = db.fetch_all("""
                    SELECT v.*,
                           c.first_name || ' ' || c.last_name as customer_name,
                           c.phone as customer_phone
                    FROM vehicles v
                    LEFT JOIN customers c ON v.customer_id = c.id
                    ORDER BY v.license_plate
                """)

            wb = openpyxl.Workbook()

            # List 1: Seznam vozidel
            ws1 = wb.active
            ws1.title = "Seznam vozidel"

            headers1 = ["SPZ", "Značka", "Model", "Rok", "VIN", "Barva", "Zákazník", "Telefon"]
            for col, header in enumerate(headers1, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row_idx, v in enumerate(vehicles, 2):
                ws1.cell(row=row_idx, column=1, value=v['license_plate'])
                ws1.cell(row=row_idx, column=2, value=v['brand'])
                ws1.cell(row=row_idx, column=3, value=v['model'])
                ws1.cell(row=row_idx, column=4, value=v['year'])
                ws1.cell(row=row_idx, column=5, value=v['vin'])
                ws1.cell(row=row_idx, column=6, value=v['color'])
                ws1.cell(row=row_idx, column=7, value=v['customer_name'] or 'Bez zákazníka')
                ws1.cell(row=row_idx, column=8, value=v['customer_phone'])

            # Šířky sloupců
            for col in range(1, len(headers1) + 1):
                ws1.column_dimensions[get_column_letter(col)].width = 15

            # List 2: Technické údaje
            ws2 = wb.create_sheet("Technické údaje")

            headers2 = ["SPZ", "Motor", "Palivo", "Stav km", "STK do", "Poznámky"]
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")

            for row_idx, v in enumerate(vehicles, 2):
                ws2.cell(row=row_idx, column=1, value=v['license_plate'])
                ws2.cell(row=row_idx, column=2, value=v['engine_type'])
                ws2.cell(row=row_idx, column=3, value=v['fuel_type'])
                ws2.cell(row=row_idx, column=4, value=v['mileage'])

                stk_text = ''
                if v['stk_valid_until']:
                    try:
                        stk_date = datetime.strptime(str(v['stk_valid_until']), "%Y-%m-%d")
                        stk_text = stk_date.strftime("%d.%m.%Y")
                    except:
                        pass
                ws2.cell(row=row_idx, column=5, value=stk_text)
                ws2.cell(row=row_idx, column=6, value=v['notes'])

            for col in range(1, len(headers2) + 1):
                ws2.column_dimensions[get_column_letter(col)].width = 15

            # List 3: Statistiky
            ws3 = wb.create_sheet("Statistiky")

            ws3.cell(row=1, column=1, value="Celkem vozidel:")
            ws3.cell(row=1, column=2, value=len(vehicles))

            # Počet podle značek
            brands = {}
            for v in vehicles:
                brand = v['brand'] or 'Neznámá'
                brands[brand] = brands.get(brand, 0) + 1

            ws3.cell(row=3, column=1, value="Vozidla podle značek:")
            ws3.cell(row=3, column=1).font = Font(bold=True)

            row = 4
            for brand, count in sorted(brands.items(), key=lambda x: x[1], reverse=True):
                ws3.cell(row=row, column=1, value=brand)
                ws3.cell(row=row, column=2, value=count)
                row += 1

            wb.save(str(output_path))
            return True, "Seznam vozidel byl exportován"

        except ImportError:
            return False, "Pro export do Excelu je potřeba nainstalovat openpyxl:\npip install openpyxl"
        except Exception as e:
            return False, f"Chyba při exportu: {e}"

    def export_complete(self, output_path):
        """Kompletní export vozidla do ZIP archivu"""
        try:
            if not self.vehicle_data:
                return False, "Data vozidla nejsou k dispozici"

            # Vytvoření dočasné složky
            temp_dir = Path(config.DATA_DIR) / "temp_export" / str(self.vehicle_id)
            temp_dir.mkdir(parents=True, exist_ok=True)

            # Export karty vozidla
            card_path = temp_dir / "karta_vozidla.pdf"
            self.export_vehicle_card(card_path)

            # Export servisní knihy
            service_path = temp_dir / "servisni_kniha.pdf"
            self.export_service_book(service_path)

            # Export dat do JSON
            json_path = temp_dir / "data.json"
            vehicle_json = {
                'id': self.vehicle_data['id'],
                'license_plate': self.vehicle_data['license_plate'],
                'brand': self.vehicle_data['brand'],
                'model': self.vehicle_data['model'],
                'year': self.vehicle_data['year'],
                'vin': self.vehicle_data['vin'],
                'color': self.vehicle_data['color'],
                'engine_type': self.vehicle_data['engine_type'],
                'fuel_type': self.vehicle_data['fuel_type'],
                'mileage': self.vehicle_data['mileage'],
                'stk_valid_until': str(self.vehicle_data['stk_valid_until']) if self.vehicle_data['stk_valid_until'] else None,
                'notes': self.vehicle_data['notes'],
                'customer': {
                    'name': self.vehicle_data['customer_name'],
                    'phone': self.vehicle_data['customer_phone'],
                    'email': self.vehicle_data['customer_email'],
                } if self.vehicle_data['customer_id'] else None,
                'exported_at': datetime.now().isoformat()
            }

            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(vehicle_json, f, ensure_ascii=False, indent=2)

            # Kopírování fotek
            photos_dir = Path(config.DATA_DIR) / "vehicle_photos" / str(self.vehicle_id)
            if photos_dir.exists():
                export_photos_dir = temp_dir / "fotografie"
                export_photos_dir.mkdir(exist_ok=True)

                import shutil
                for photo_file in photos_dir.glob("*"):
                    if photo_file.is_file() and not photo_file.name.startswith("thumb_"):
                        shutil.copy2(photo_file, export_photos_dir / photo_file.name)

            # Kopírování dokumentů
            docs_dir = Path(config.DATA_DIR) / "vehicle_documents" / str(self.vehicle_id)
            if docs_dir.exists():
                export_docs_dir = temp_dir / "dokumenty"
                export_docs_dir.mkdir(exist_ok=True)

                import shutil
                for doc_file in docs_dir.glob("*"):
                    if doc_file.is_file():
                        shutil.copy2(doc_file, export_docs_dir / doc_file.name)

            # Vytvoření ZIP archivu
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in temp_dir.rglob("*"):
                    if file_path.is_file():
                        arcname = file_path.relative_to(temp_dir)
                        zipf.write(file_path, arcname)

            # Úklid dočasné složky
            import shutil
            shutil.rmtree(temp_dir)

            return True, "Kompletní export byl vytvořen"

        except Exception as e:
            return False, f"Chyba při exportu: {e}"


class ExportDialog(QDialog):
    """Dialog pro export vozidla"""

    def __init__(self, vehicle_id, parent=None):
        super().__init__(parent)
        self.vehicle_id = vehicle_id
        self.exporter = VehicleExporter(vehicle_id)
        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        self.setWindowTitle("📤 Export vozidla")
        self.setMinimumWidth(500)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Info o vozidle
        if self.exporter.vehicle_data:
            info_label = QLabel(
                f"<b>{self.exporter.vehicle_data['brand']} {self.exporter.vehicle_data['model']}</b><br>"
                f"SPZ: {self.exporter.vehicle_data['license_plate']}"
            )
            info_label.setStyleSheet("""
                QLabel {
                    padding: 10px;
                    background-color: #ecf0f1;
                    border-radius: 5px;
                }
            """)
            layout.addWidget(info_label)

        # Typ exportu
        type_group = QGroupBox("Typ exportu")
        type_layout = QVBoxLayout(type_group)

        self.export_card = QCheckBox("📄 Karta vozidla (PDF)")
        self.export_card.setChecked(True)
        type_layout.addWidget(self.export_card)

        self.export_service = QCheckBox("📋 Servisní kniha (PDF)")
        self.export_service.setChecked(True)
        type_layout.addWidget(self.export_service)

        self.export_complete = QCheckBox("📁 Kompletní export (ZIP s fotkami a dokumenty)")
        type_layout.addWidget(self.export_complete)

        layout.addWidget(type_group)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # Status
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        btn_cancel = QPushButton("❌ Zrušit")
        btn_cancel.clicked.connect(self.reject)
        buttons_layout.addWidget(btn_cancel)

        btn_export = QPushButton("📤 Exportovat")
        btn_export.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #219a52;
            }}
        """)
        btn_export.clicked.connect(self.perform_export)
        buttons_layout.addWidget(btn_export)

        layout.addLayout(buttons_layout)

    def perform_export(self):
        """Provedení exportu"""
        if not any([self.export_card.isChecked(), self.export_service.isChecked(), self.export_complete.isChecked()]):
            QMessageBox.warning(self, "Upozornění", "Vyberte alespoň jeden typ exportu.")
            return

        # Výběr složky
        export_dir = QFileDialog.getExistingDirectory(self, "Vybrat složku pro export")
        if not export_dir:
            return

        export_path = Path(export_dir)
        spz = self.exporter.vehicle_data['license_plate'].replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        results = []
        total_exports = sum([
            self.export_card.isChecked(),
            self.export_service.isChecked(),
            self.export_complete.isChecked()
        ])
        current = 0

        try:
            # Karta vozidla
            if self.export_card.isChecked():
                self.status_label.setText("Exportuji kartu vozidla...")
                QApplication.processEvents()

                card_path = export_path / f"karta_{spz}_{timestamp}.pdf"
                success, msg = self.exporter.export_vehicle_card(card_path)
                results.append(f"{'✅' if success else '❌'} Karta vozidla: {msg}")

                current += 1
                self.progress.setValue(int(current / total_exports * 100))

            # Servisní kniha
            if self.export_service.isChecked():
                self.status_label.setText("Exportuji servisní knihu...")
                QApplication.processEvents()

                service_path = export_path / f"servisni_kniha_{spz}_{timestamp}.pdf"
                success, msg = self.exporter.export_service_book(service_path)
                results.append(f"{'✅' if success else '❌'} Servisní kniha: {msg}")

                current += 1
                self.progress.setValue(int(current / total_exports * 100))

            # Kompletní export
            if self.export_complete.isChecked():
                self.status_label.setText("Vytvářím kompletní export...")
                QApplication.processEvents()

                zip_path = export_path / f"kompletni_export_{spz}_{timestamp}.zip"
                success, msg = self.exporter.export_complete(zip_path)
                results.append(f"{'✅' if success else '❌'} Kompletní export: {msg}")

                current += 1
                self.progress.setValue(int(current / total_exports * 100))

            self.status_label.setText("Export dokončen!")

            # Zobrazení výsledků
            result_text = "\n".join(results)
            QMessageBox.information(
                self,
                "Export dokončen",
                f"Výsledky exportu:\n\n{result_text}\n\nSoubory uloženy do:\n{export_dir}"
            )

            self.accept()

        except Exception as e:
            self.status_label.setText(f"Chyba: {e}")
            QMessageBox.critical(self, "Chyba", f"Chyba při exportu:\n{e}")

        finally:
            self.progress.setVisible(False)


# Pro import v hlavním widgetu
from PyQt6.QtWidgets import QApplication
