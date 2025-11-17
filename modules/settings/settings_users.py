# -*- coding: utf-8 -*-
"""
Správa uživatelů systému
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QDialog, QFormLayout, QLineEdit, QComboBox,
    QCheckBox, QMessageBox, QHeaderView, QFrame, QGroupBox
)
from PyQt6.QtCore import Qt, QDateTime
from PyQt6.QtGui import QFont, QColor
import hashlib
import secrets
from database_manager import db
import config


class UsersSettingsWidget(QWidget):
    """Widget pro správu uživatelů"""

    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_users()

    def init_ui(self):
        """Inicializace UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Horní panel s tlačítky
        top_panel = QHBoxLayout()

        add_btn = QPushButton("➕ Nový uživatel")
        add_btn.clicked.connect(self.add_user)
        add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        add_btn.setObjectName("primaryButton")

        export_btn = QPushButton("📤 Export")
        export_btn.clicked.connect(self.export_users)
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        refresh_btn = QPushButton("🔄 Obnovit")
        refresh_btn.clicked.connect(self.load_users)
        refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        top_panel.addWidget(add_btn)
        top_panel.addWidget(export_btn)
        top_panel.addWidget(refresh_btn)
        top_panel.addStretch()

        # Statistiky
        self.stats_label = QLabel()
        self.stats_label.setStyleSheet("color: #7f8c8d;")
        top_panel.addWidget(self.stats_label)

        layout.addLayout(top_panel)

        # Tabulka uživatelů
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ID", "Jméno", "Příjmení", "Uživatelské jméno",
            "Email", "Role", "Poslední přihlášení", "Stav"
        ])

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)

        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(5, 120)
        self.table.setColumnWidth(6, 150)
        self.table.setColumnWidth(7, 100)

        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.doubleClicked.connect(self.edit_user)

        layout.addWidget(self.table)

        # Spodní panel s akcemi
        bottom_panel = QHBoxLayout()

        edit_btn = QPushButton("✏️ Upravit")
        edit_btn.clicked.connect(self.edit_user)
        edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        reset_pwd_btn = QPushButton("🔑 Reset hesla")
        reset_pwd_btn.clicked.connect(self.reset_password)
        reset_pwd_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        toggle_btn = QPushButton("🔄 Aktivovat/Deaktivovat")
        toggle_btn.clicked.connect(self.toggle_active)
        toggle_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        delete_btn = QPushButton("🗑️ Smazat")
        delete_btn.clicked.connect(self.delete_user)
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.setObjectName("dangerButton")

        activity_btn = QPushButton("📋 Historie aktivit")
        activity_btn.clicked.connect(self.view_activity)
        activity_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        bottom_panel.addWidget(edit_btn)
        bottom_panel.addWidget(reset_pwd_btn)
        bottom_panel.addWidget(toggle_btn)
        bottom_panel.addWidget(delete_btn)
        bottom_panel.addWidget(activity_btn)
        bottom_panel.addStretch()

        layout.addLayout(bottom_panel)

        self.set_styles()

    def load_users(self):
        """Načtení uživatelů z databáze"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, first_name, last_name, username, email, role,
                       last_login, is_active
                FROM users
                ORDER BY last_name, first_name
            """)

            rows = cursor.fetchall()
            self.table.setRowCount(len(rows))

            active_count = 0

            for i, row in enumerate(rows):
                user_id, first_name, last_name, username, email, role, last_login, is_active = row

                self.table.setItem(i, 0, QTableWidgetItem(str(user_id)))
                self.table.setItem(i, 1, QTableWidgetItem(first_name or ""))
                self.table.setItem(i, 2, QTableWidgetItem(last_name or ""))
                self.table.setItem(i, 3, QTableWidgetItem(username))
                self.table.setItem(i, 4, QTableWidgetItem(email or ""))
                self.table.setItem(i, 5, QTableWidgetItem(role or ""))

                # Poslední přihlášení
                if last_login:
                    login_item = QTableWidgetItem(last_login)
                else:
                    login_item = QTableWidgetItem("Nikdy")
                    login_item.setForeground(QColor("#7f8c8d"))
                self.table.setItem(i, 6, login_item)

                # Stav
                if is_active:
                    status_item = QTableWidgetItem("✅ Aktivní")
                    status_item.setForeground(QColor(config.COLOR_SUCCESS))
                    active_count += 1
                else:
                    status_item = QTableWidgetItem("❌ Neaktivní")
                    status_item.setForeground(QColor(config.COLOR_DANGER))
                self.table.setItem(i, 7, status_item)

            # Aktualizace statistik
            self.stats_label.setText(
                f"Celkem: {len(rows)} uživatelů | Aktivních: {active_count}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se načíst uživatele:\n{str(e)}")

    def add_user(self):
        """Přidání nového uživatele"""
        dialog = UserDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_users()

    def edit_user(self):
        """Úprava vybraného uživatele"""
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Upozornění", "Vyberte uživatele pro úpravu.")
            return

        user_id = int(self.table.item(self.table.currentRow(), 0).text())
        dialog = UserDialog(self, user_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_users()

    def delete_user(self):
        """Smazání uživatele"""
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Upozornění", "Vyberte uživatele pro smazání.")
            return

        user_id = int(self.table.item(self.table.currentRow(), 0).text())
        username = self.table.item(self.table.currentRow(), 3).text()

        reply = QMessageBox.question(
            self,
            "Smazat uživatele",
            f"Opravdu chcete smazat uživatele '{username}'?\n\n"
            "Tato akce nelze vrátit zpět!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = db.get_connection()
                cursor = conn.cursor()

                # Kontrola, zda uživatel nemá přiřazená data
                cursor.execute("SELECT COUNT(*) FROM orders WHERE user_id = ?", (user_id,))
                if cursor.fetchone()[0] > 0:
                    QMessageBox.warning(
                        self,
                        "Nelze smazat",
                        "Uživatel má přiřazené zakázky.\n"
                        "Nejdříve je přesuňte na jiného uživatele."
                    )
                    return

                cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
                conn.commit()

                QMessageBox.information(self, "Hotovo", "Uživatel byl smazán.")
                self.load_users()

            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se smazat uživatele:\n{str(e)}")

    def reset_password(self):
        """Reset hesla uživatele"""
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Upozornění", "Vyberte uživatele pro reset hesla.")
            return

        user_id = int(self.table.item(self.table.currentRow(), 0).text())
        username = self.table.item(self.table.currentRow(), 3).text()

        reply = QMessageBox.question(
            self,
            "Reset hesla",
            f"Opravdu chcete resetovat heslo uživatele '{username}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Generování nového hesla
                new_password = secrets.token_urlsafe(12)
                password_hash = hashlib.sha256(new_password.encode()).hexdigest()

                conn = db.get_connection()
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE users
                    SET password_hash = ?, force_password_change = 1
                    WHERE id = ?
                """, (password_hash, user_id))

                conn.commit()

                QMessageBox.information(
                    self,
                    "Heslo resetováno",
                    f"Nové heslo pro uživatele '{username}':\n\n"
                    f"{new_password}\n\n"
                    "Uživatel bude vyzván ke změně hesla při příštím přihlášení.\n"
                    "Zkopírujte heslo a předejte ho uživateli."
                )

            except Exception as e:
                QMessageBox.critical(self, "Chyba", f"Nepodařilo se resetovat heslo:\n{str(e)}")

    def toggle_active(self):
        """Aktivace/deaktivace uživatele"""
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Upozornění", "Vyberte uživatele.")
            return

        user_id = int(self.table.item(self.table.currentRow(), 0).text())
        current_status = "Aktivní" in self.table.item(self.table.currentRow(), 7).text()

        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            new_status = 0 if current_status else 1
            cursor.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
            conn.commit()

            self.load_users()

            status_text = "aktivován" if new_status else "deaktivován"
            QMessageBox.information(self, "Hotovo", f"Uživatel byl {status_text}.")

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Chyba při změně stavu:\n{str(e)}")

    def view_activity(self):
        """Zobrazení historie aktivit"""
        selected = self.table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Upozornění", "Vyberte uživatele.")
            return

        user_id = int(self.table.item(self.table.currentRow(), 0).text())
        username = self.table.item(self.table.currentRow(), 3).text()

        # TODO: Implementovat zobrazení historie aktivit
        QMessageBox.information(
            self,
            "Historie aktivit",
            f"Historie aktivit pro uživatele '{username}'\n\n"
            "Tato funkce bude implementována v další verzi."
        )

    def export_users(self):
        """Export uživatelů"""
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Uživatelé"

            # Hlavička
            headers = ["ID", "Jméno", "Příjmení", "Uživatelské jméno", "Email", "Role", "Stav"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data
            for row in range(self.table.rowCount()):
                ws.cell(row=row + 2, column=1, value=self.table.item(row, 0).text())
                ws.cell(row=row + 2, column=2, value=self.table.item(row, 1).text())
                ws.cell(row=row + 2, column=3, value=self.table.item(row, 2).text())
                ws.cell(row=row + 2, column=4, value=self.table.item(row, 3).text())
                ws.cell(row=row + 2, column=5, value=self.table.item(row, 4).text())
                ws.cell(row=row + 2, column=6, value=self.table.item(row, 5).text())
                ws.cell(row=row + 2, column=7, value=self.table.item(row, 7).text())

            file_path = config.EXPORTS_DIR / f"uzivatele_{QDateTime.currentDateTime().toString('yyyyMMdd_HHmmss')}.xlsx"
            wb.save(str(file_path))

            QMessageBox.information(
                self,
                "Export dokončen",
                f"Uživatelé byli exportováni do:\n{file_path}"
            )

        except ImportError:
            QMessageBox.warning(
                self,
                "Chybí knihovna",
                "Pro export do Excelu nainstalujte knihovnu openpyxl:\n\n"
                "pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Chyba exportu", f"Chyba při exportu:\n{str(e)}")

    def save_settings(self):
        """Uložení (není potřeba - vše se ukládá okamžitě)"""
        pass

    def refresh(self):
        """Obnovení"""
        self.load_users()

    def set_styles(self):
        """Nastavení stylů"""
        self.setStyleSheet(f"""
            QTableWidget {{
                background-color: white;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                gridline-color: #ecf0f1;
            }}

            QTableWidget::item {{
                padding: 8px;
            }}

            QTableWidget::item:selected {{
                background-color: {config.COLOR_SECONDARY};
                color: white;
            }}

            QHeaderView::section {{
                background-color: #34495e;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }}

            QPushButton {{
                padding: 8px 16px;
                border-radius: 4px;
                background-color: #ecf0f1;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
            }}

            QPushButton:hover {{
                background-color: #d5dbdb;
            }}

            #primaryButton {{
                background-color: {config.COLOR_SECONDARY};
                color: white;
                border: none;
            }}

            #primaryButton:hover {{
                background-color: #2980b9;
            }}

            #dangerButton {{
                background-color: {config.COLOR_DANGER};
                color: white;
                border: none;
            }}

            #dangerButton:hover {{
                background-color: #c0392b;
            }}
        """)


class UserDialog(QDialog):
    """Dialog pro vytvoření/úpravu uživatele"""

    def __init__(self, parent=None, user_id=None):
        super().__init__(parent)
        self.user_id = user_id
        self.setWindowTitle("Nový uživatel" if not user_id else "Upravit uživatele")
        self.setMinimumWidth(500)
        self.init_ui()
        if user_id:
            self.load_user_data()

    def init_ui(self):
        """Inicializace UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Základní údaje
        basic_group = QGroupBox("👤 Základní údaje")
        basic_layout = QFormLayout(basic_group)
        basic_layout.setSpacing(10)

        self.first_name = QLineEdit()
        self.first_name.setPlaceholderText("Jméno *")
        basic_layout.addRow("Jméno *:", self.first_name)

        self.last_name = QLineEdit()
        self.last_name.setPlaceholderText("Příjmení *")
        basic_layout.addRow("Příjmení *:", self.last_name)

        self.email = QLineEdit()
        self.email.setPlaceholderText("email@example.com *")
        basic_layout.addRow("Email *:", self.email)

        self.phone = QLineEdit()
        self.phone.setPlaceholderText("+420 xxx xxx xxx")
        basic_layout.addRow("Telefon:", self.phone)

        layout.addWidget(basic_group)

        # Přihlašovací údaje
        login_group = QGroupBox("🔐 Přihlašovací údaje")
        login_layout = QFormLayout(login_group)
        login_layout.setSpacing(10)

        self.username = QLineEdit()
        self.username.setPlaceholderText("Unikátní uživatelské jméno *")
        login_layout.addRow("Uživatelské jméno *:", self.username)

        if not self.user_id:
            self.password = QLineEdit()
            self.password.setPlaceholderText("Min. 8 znaků, velké/malé, číslo *")
            self.password.setEchoMode(QLineEdit.EchoMode.Password)
            login_layout.addRow("Heslo *:", self.password)

            self.password_confirm = QLineEdit()
            self.password_confirm.setPlaceholderText("Zopakujte heslo *")
            self.password_confirm.setEchoMode(QLineEdit.EchoMode.Password)
            login_layout.addRow("Potvrdit heslo *:", self.password_confirm)

            self.force_change = QCheckBox("Vynutit změnu hesla při prvním přihlášení")
            self.force_change.setChecked(True)
            login_layout.addRow("", self.force_change)

        layout.addWidget(login_group)

        # Role a oprávnění
        role_group = QGroupBox("🎭 Role a oprávnění")
        role_layout = QFormLayout(role_group)
        role_layout.setSpacing(10)

        self.role = QComboBox()
        self.role.addItems([
            "Administrátor",
            "Manažer",
            "Mechanik",
            "Recepce",
            "Účetní",
            "Sklad"
        ])
        role_layout.addRow("Role:", self.role)

        self.is_active = QCheckBox("Aktivní účet")
        self.is_active.setChecked(True)
        role_layout.addRow("", self.is_active)

        layout.addWidget(role_group)

        # Notifikace
        notif_group = QGroupBox("🔔 Notifikace")
        notif_layout = QFormLayout(notif_group)
        notif_layout.setSpacing(10)

        self.email_notifications = QCheckBox("Emailové notifikace")
        self.email_notifications.setChecked(True)
        notif_layout.addRow("", self.email_notifications)

        self.daily_summary = QCheckBox("Denní souhrn")
        notif_layout.addRow("", self.daily_summary)

        layout.addWidget(notif_group)

        # Tlačítka
        buttons_layout = QHBoxLayout()

        save_btn = QPushButton("💾 Uložit")
        save_btn.clicked.connect(self.save_user)
        save_btn.setObjectName("saveButton")

        cancel_btn = QPushButton("❌ Zrušit")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        self.set_styles()

    def load_user_data(self):
        """Načtení dat existujícího uživatele"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT first_name, last_name, username, email, phone, role,
                       is_active, email_notifications, daily_summary
                FROM users WHERE id = ?
            """, (self.user_id,))

            row = cursor.fetchone()
            if row:
                self.first_name.setText(row[0] or "")
                self.last_name.setText(row[1] or "")
                self.username.setText(row[2] or "")
                self.email.setText(row[3] or "")
                self.phone.setText(row[4] or "")

                role_index = self.role.findText(row[5] or "")
                if role_index >= 0:
                    self.role.setCurrentIndex(role_index)

                self.is_active.setChecked(bool(row[6]))
                self.email_notifications.setChecked(bool(row[7]))
                self.daily_summary.setChecked(bool(row[8]))

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se načíst data:\n{str(e)}")

    def validate_password(self, password):
        """Validace hesla"""
        if len(password) < 8:
            return False, "Heslo musí mít alespoň 8 znaků"
        if not any(c.isupper() for c in password):
            return False, "Heslo musí obsahovat velké písmeno"
        if not any(c.islower() for c in password):
            return False, "Heslo musí obsahovat malé písmeno"
        if not any(c.isdigit() for c in password):
            return False, "Heslo musí obsahovat číslo"
        return True, ""

    def save_user(self):
        """Uložení uživatele"""
        # Validace
        if not self.first_name.text().strip():
            QMessageBox.warning(self, "Chyba", "Jméno je povinné.")
            return

        if not self.last_name.text().strip():
            QMessageBox.warning(self, "Chyba", "Příjmení je povinné.")
            return

        if not self.email.text().strip():
            QMessageBox.warning(self, "Chyba", "Email je povinný.")
            return

        if not self.username.text().strip():
            QMessageBox.warning(self, "Chyba", "Uživatelské jméno je povinné.")
            return

        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            # Kontrola unikátnosti uživatelského jména
            if self.user_id:
                cursor.execute(
                    "SELECT id FROM users WHERE username = ? AND id != ?",
                    (self.username.text(), self.user_id)
                )
            else:
                cursor.execute(
                    "SELECT id FROM users WHERE username = ?",
                    (self.username.text(),)
                )

            if cursor.fetchone():
                QMessageBox.warning(self, "Chyba", "Toto uživatelské jméno již existuje.")
                return

            if self.user_id:
                # Úprava existujícího uživatele
                cursor.execute("""
                    UPDATE users SET
                        first_name = ?, last_name = ?, username = ?, email = ?,
                        phone = ?, role = ?, is_active = ?, email_notifications = ?,
                        daily_summary = ?
                    WHERE id = ?
                """, (
                    self.first_name.text(),
                    self.last_name.text(),
                    self.username.text(),
                    self.email.text(),
                    self.phone.text(),
                    self.role.currentText(),
                    int(self.is_active.isChecked()),
                    int(self.email_notifications.isChecked()),
                    int(self.daily_summary.isChecked()),
                    self.user_id
                ))
            else:
                # Nový uživatel
                if not self.password.text():
                    QMessageBox.warning(self, "Chyba", "Heslo je povinné.")
                    return

                valid, msg = self.validate_password(self.password.text())
                if not valid:
                    QMessageBox.warning(self, "Chyba", msg)
                    return

                if self.password.text() != self.password_confirm.text():
                    QMessageBox.warning(self, "Chyba", "Hesla se neshodují.")
                    return

                password_hash = hashlib.sha256(self.password.text().encode()).hexdigest()

                cursor.execute("""
                    INSERT INTO users (
                        first_name, last_name, username, email, phone, role,
                        password_hash, is_active, force_password_change,
                        email_notifications, daily_summary
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    self.first_name.text(),
                    self.last_name.text(),
                    self.username.text(),
                    self.email.text(),
                    self.phone.text(),
                    self.role.currentText(),
                    password_hash,
                    int(self.is_active.isChecked()),
                    int(self.force_change.isChecked()),
                    int(self.email_notifications.isChecked()),
                    int(self.daily_summary.isChecked())
                ))

            conn.commit()
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se uložit uživatele:\n{str(e)}")

    def set_styles(self):
        """Nastavení stylů"""
        self.setStyleSheet(f"""
            QGroupBox {{
                font-weight: bold;
                border: 1px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }}

            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }}

            QLineEdit, QComboBox {{
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
            }}

            QLineEdit:focus, QComboBox:focus {{
                border: 2px solid {config.COLOR_SECONDARY};
            }}

            QPushButton {{
                padding: 10px 20px;
                border-radius: 4px;
                background-color: #ecf0f1;
                border: 1px solid #bdc3c7;
            }}

            QPushButton:hover {{
                background-color: #d5dbdb;
            }}

            #saveButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                border: none;
                font-weight: bold;
            }}

            #saveButton:hover {{
                background-color: #229954;
            }}
        """)
