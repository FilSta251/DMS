# -*- coding: utf-8 -*-
"""
Modul Administrativa - Exporty dat a dokumentů (PRODUKČNÍ VERZE)
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QTableWidget, QTableWidgetItem, QFrame,
                             QComboBox, QLineEdit, QDateEdit, QHeaderView,
                             QMessageBox, QDialog, QFormLayout, QTextEdit,
                             QSpinBox, QDoubleSpinBox, QFileDialog, QCheckBox,
                             QGroupBox, QTabWidget, QScrollArea, QProgressBar,
                             QListWidget, QListWidgetItem, QRadioButton, QButtonGroup)
from PyQt6.QtCore import Qt, QDate, pyqtSignal, QThread
from PyQt6.QtGui import QFont, QColor
from datetime import datetime, timedelta, date
from pathlib import Path
import json
import csv
import config
from database_manager import db


class ExportsWidget(QWidget):
    """Widget pro exporty"""

    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Hlavička
        header_label = QLabel("📤 Exporty dat a dokumentů")
        header_font = QFont()
        header_font.setPointSize(16)
        header_font.setBold(True)
        header_label.setFont(header_font)
        layout.addWidget(header_label)

        # Záložky
        tabs = QTabWidget()

        # Záložka: Export faktur
        self.tab_invoices = self.create_invoices_tab()
        tabs.addTab(self.tab_invoices, "💰 Export faktur")

        # Záložka: Export plateb
        self.tab_payments = self.create_payments_tab()
        tabs.addTab(self.tab_payments, "💳 Export plateb")

        # Záložka: Export účetních dat
        self.tab_accounting = self.create_accounting_tab()
        tabs.addTab(self.tab_accounting, "📊 Účetní data")

        # Záložka: Plánované exporty
        self.tab_scheduled = self.create_scheduled_tab()
        tabs.addTab(self.tab_scheduled, "⏰ Plánované exporty")

        # Záložka: Šablony exportů
        self.tab_templates = self.create_templates_tab()
        tabs.addTab(self.tab_templates, "📋 Šablony")

        layout.addWidget(tabs)

    def create_invoices_tab(self):
        """Záložka: Export faktur"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Výběr faktur
        selection_group = QGroupBox("📋 Výběr faktur")
        selection_layout = QVBoxLayout(selection_group)

        # Typ výběru
        self.invoice_selection_type = QButtonGroup()

        all_radio = QRadioButton("Všechny faktury")
        all_radio.setChecked(True)
        self.invoice_selection_type.addButton(all_radio, 0)
        selection_layout.addWidget(all_radio)

        period_radio = QRadioButton("Faktury za období")
        self.invoice_selection_type.addButton(period_radio, 1)
        selection_layout.addWidget(period_radio)

        selected_radio = QRadioButton("Vybrané faktury")
        self.invoice_selection_type.addButton(selected_radio, 2)
        selection_layout.addWidget(selected_radio)

        # Období
        period_widget = QWidget()
        period_layout = QHBoxLayout(period_widget)
        period_layout.setContentsMargins(20, 0, 0, 0)

        period_layout.addWidget(QLabel("Od:"))
        self.invoice_date_from = QDateEdit()
        self.invoice_date_from.setDate(QDate.currentDate().addMonths(-1))
        self.invoice_date_from.setCalendarPopup(True)
        self.invoice_date_from.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.invoice_date_from)

        period_layout.addWidget(QLabel("Do:"))
        self.invoice_date_to = QDateEdit()
        self.invoice_date_to.setDate(QDate.currentDate())
        self.invoice_date_to.setCalendarPopup(True)
        self.invoice_date_to.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.invoice_date_to)

        period_layout.addStretch()
        selection_layout.addWidget(period_widget)

        # Filtry
        filters_layout = QHBoxLayout()

        filters_layout.addWidget(QLabel("Typ:"))
        self.invoice_type_filter = QComboBox()
        self.invoice_type_filter.addItem("Všechny typy", "all")
        self.invoice_type_filter.addItem("Vydané faktury", "issued")
        self.invoice_type_filter.addItem("Přijaté faktury", "received")
        filters_layout.addWidget(self.invoice_type_filter)

        filters_layout.addWidget(QLabel("Status:"))
        self.invoice_status_filter = QComboBox()
        self.invoice_status_filter.addItem("Všechny statusy", "all")
        self.invoice_status_filter.addItem("Zaplaceno", "paid")
        self.invoice_status_filter.addItem("Nezaplaceno", "unpaid")
        self.invoice_status_filter.addItem("Po splatnosti", "overdue")
        filters_layout.addWidget(self.invoice_status_filter)

        filters_layout.addStretch()
        selection_layout.addLayout(filters_layout)

        layout.addWidget(selection_group)

        # Formát exportu
        format_group = QGroupBox("📄 Formát exportu")
        format_layout = QVBoxLayout(format_group)

        format_buttons_layout = QHBoxLayout()

        self.invoice_format = QButtonGroup()

        pdf_radio = QRadioButton("PDF (faktury)")
        pdf_radio.setChecked(True)
        self.invoice_format.addButton(pdf_radio, 0)
        format_buttons_layout.addWidget(pdf_radio)

        excel_radio = QRadioButton("Excel (přehled)")
        self.invoice_format.addButton(excel_radio, 1)
        format_buttons_layout.addWidget(excel_radio)

        csv_radio = QRadioButton("CSV")
        self.invoice_format.addButton(csv_radio, 2)
        format_buttons_layout.addWidget(csv_radio)

        xml_radio = QRadioButton("XML")
        self.invoice_format.addButton(xml_radio, 3)
        format_buttons_layout.addWidget(xml_radio)

        format_buttons_layout.addStretch()
        format_layout.addLayout(format_buttons_layout)

        # Volby PDF
        pdf_options_widget = QWidget()
        pdf_options_layout = QVBoxLayout(pdf_options_widget)
        pdf_options_layout.setContentsMargins(20, 0, 0, 0)

        self.pdf_individual = QRadioButton("Jednotlivé PDF soubory")
        self.pdf_individual.setChecked(True)
        pdf_options_layout.addWidget(self.pdf_individual)

        self.pdf_merged = QRadioButton("Sloučit do jednoho PDF")
        pdf_options_layout.addWidget(self.pdf_merged)

        self.pdf_options_group = pdf_options_widget
        format_layout.addWidget(pdf_options_widget)

        layout.addWidget(format_group)

        # Nastavení
        options_group = QGroupBox("⚙️ Nastavení")
        options_layout = QVBoxLayout(options_group)

        self.invoice_include_attachments = QCheckBox("Zahrnout přílohy")
        options_layout.addWidget(self.invoice_include_attachments)

        self.invoice_include_payments = QCheckBox("Zahrnout platby")
        options_layout.addWidget(self.invoice_include_payments)

        self.invoice_include_items = QCheckBox("Zahrnout položky")
        self.invoice_include_items.setChecked(True)
        options_layout.addWidget(self.invoice_include_items)

        layout.addWidget(options_group)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        preview_btn = QPushButton("👁️ Náhled")
        preview_btn.clicked.connect(lambda: self.preview_export("invoices"))
        buttons_layout.addWidget(preview_btn)

        export_btn = QPushButton("📤 Exportovat")
        export_btn.clicked.connect(self.export_invoices)
        export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 10px 30px;
                font-weight: bold;
            }}
        """)
        buttons_layout.addWidget(export_btn)

        layout.addLayout(buttons_layout)

        layout.addStretch()

        return widget

    def create_payments_tab(self):
        """Záložka: Export plateb"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Období
        period_group = QGroupBox("📅 Období")
        period_layout = QHBoxLayout(period_group)

        period_layout.addWidget(QLabel("Od:"))
        self.payment_date_from = QDateEdit()
        self.payment_date_from.setDate(QDate.currentDate().addMonths(-1))
        self.payment_date_from.setCalendarPopup(True)
        self.payment_date_from.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.payment_date_from)

        period_layout.addWidget(QLabel("Do:"))
        self.payment_date_to = QDateEdit()
        self.payment_date_to.setDate(QDate.currentDate())
        self.payment_date_to.setCalendarPopup(True)
        self.payment_date_to.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.payment_date_to)

        period_layout.addStretch()
        layout.addWidget(period_group)

        # Filtry
        filters_group = QGroupBox("🔍 Filtry")
        filters_layout = QFormLayout(filters_group)

        self.payment_type_filter = QComboBox()
        self.payment_type_filter.addItem("Všechny typy plateb", "all")
        self.payment_type_filter.addItem("Příjmy", "income")
        self.payment_type_filter.addItem("Výdaje", "expense")
        filters_layout.addRow("Typ platby:", self.payment_type_filter)

        self.payment_method_filter = QComboBox()
        self.payment_method_filter.addItem("Všechny způsoby", "all")
        self.payment_method_filter.addItem("Bankovní převod", "bank_transfer")
        self.payment_method_filter.addItem("Hotovost", "cash")
        self.payment_method_filter.addItem("Karta", "card")
        filters_layout.addRow("Způsob platby:", self.payment_method_filter)

        layout.addWidget(filters_group)

        # Formát exportu
        format_group = QGroupBox("📄 Formát exportu")
        format_layout = QVBoxLayout(format_group)

        format_buttons_layout = QHBoxLayout()

        self.payment_format = QButtonGroup()

        excel_radio = QRadioButton("Excel")
        excel_radio.setChecked(True)
        self.payment_format.addButton(excel_radio, 0)
        format_buttons_layout.addWidget(excel_radio)

        csv_radio = QRadioButton("CSV (pro import do banky)")
        self.payment_format.addButton(csv_radio, 1)
        format_buttons_layout.addWidget(csv_radio)

        pdf_radio = QRadioButton("PDF (přehled)")
        self.payment_format.addButton(pdf_radio, 2)
        format_buttons_layout.addWidget(pdf_radio)

        format_buttons_layout.addStretch()
        format_layout.addLayout(format_buttons_layout)

        layout.addWidget(format_group)

        # Nastavení
        options_group = QGroupBox("⚙️ Nastavení")
        options_layout = QVBoxLayout(options_group)

        self.payment_include_unpaid = QCheckBox("Zahrnout nezaplacené faktury")
        options_layout.addWidget(self.payment_include_unpaid)

        self.payment_group_by_customer = QCheckBox("Seskupit podle zákazníka")
        options_layout.addWidget(self.payment_group_by_customer)

        self.payment_include_summary = QCheckBox("Zahrnout součty")
        self.payment_include_summary.setChecked(True)
        options_layout.addWidget(self.payment_include_summary)

        layout.addWidget(options_group)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        preview_btn = QPushButton("👁️ Náhled")
        preview_btn.clicked.connect(lambda: self.preview_export("payments"))
        buttons_layout.addWidget(preview_btn)

        export_btn = QPushButton("📤 Exportovat")
        export_btn.clicked.connect(self.export_payments)
        export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 10px 30px;
                font-weight: bold;
            }}
        """)
        buttons_layout.addWidget(export_btn)

        layout.addLayout(buttons_layout)

        layout.addStretch()

        return widget

    def create_accounting_tab(self):
        """Záložka: Export účetních dat"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Období
        period_group = QGroupBox("📅 Období")
        period_layout = QHBoxLayout(period_group)

        period_layout.addWidget(QLabel("Od:"))
        self.accounting_date_from = QDateEdit()
        self.accounting_date_from.setDate(QDate.currentDate().addMonths(-1))
        self.accounting_date_from.setCalendarPopup(True)
        self.accounting_date_from.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.accounting_date_from)

        period_layout.addWidget(QLabel("Do:"))
        self.accounting_date_to = QDateEdit()
        self.accounting_date_to.setDate(QDate.currentDate())
        self.accounting_date_to.setCalendarPopup(True)
        self.accounting_date_to.setDisplayFormat("dd.MM.yyyy")
        period_layout.addWidget(self.accounting_date_to)

        period_layout.addStretch()
        layout.addWidget(period_group)

        # Typ exportu
        type_group = QGroupBox("📊 Typ exportu")
        type_layout = QVBoxLayout(type_group)

        self.accounting_type = QButtonGroup()

        complete_radio = QRadioButton("Kompletní účetní data")
        complete_radio.setChecked(True)
        self.accounting_type.addButton(complete_radio, 0)
        type_layout.addWidget(complete_radio)

        invoices_only_radio = QRadioButton("Pouze faktury")
        self.accounting_type.addButton(invoices_only_radio, 1)
        type_layout.addWidget(invoices_only_radio)

        vat_only_radio = QRadioButton("Pouze DPH přehled")
        self.accounting_type.addButton(vat_only_radio, 2)
        type_layout.addWidget(vat_only_radio)

        layout.addWidget(type_group)

        # Formát pro účetní software
        software_group = QGroupBox("💼 Účetní software")
        software_layout = QFormLayout(software_group)

        self.accounting_software = QComboBox()
        self.accounting_software.addItem("Univerzální Excel", "excel")
        self.accounting_software.addItem("CSV", "csv")
        self.accounting_software.addItem("XML", "xml")
        self.accounting_software.addItem("Pohoda (XML)", "pohoda")
        self.accounting_software.addItem("Money S3 (XML)", "money")
        self.accounting_software.addItem("ABRA Gen (CSV)", "abra")
        software_layout.addRow("Formát:", self.accounting_software)

        layout.addWidget(software_group)

        # Obsah exportu
        content_group = QGroupBox("📋 Obsah exportu")
        content_layout = QVBoxLayout(content_group)

        self.accounting_include_invoices = QCheckBox("Faktury")
        self.accounting_include_invoices.setChecked(True)
        content_layout.addWidget(self.accounting_include_invoices)

        self.accounting_include_payments = QCheckBox("Platby")
        self.accounting_include_payments.setChecked(True)
        content_layout.addWidget(self.accounting_include_payments)

        self.accounting_include_vat = QCheckBox("DPH přehled")
        self.accounting_include_vat.setChecked(True)
        content_layout.addWidget(self.accounting_include_vat)

        self.accounting_include_summary = QCheckBox("Souhrny a součty")
        self.accounting_include_summary.setChecked(True)
        content_layout.addWidget(self.accounting_include_summary)

        layout.addWidget(content_group)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        preview_btn = QPushButton("👁️ Náhled")
        preview_btn.clicked.connect(lambda: self.preview_export("accounting"))
        buttons_layout.addWidget(preview_btn)

        export_btn = QPushButton("📤 Exportovat")
        export_btn.clicked.connect(self.export_accounting)
        export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {config.COLOR_SUCCESS};
                color: white;
                padding: 10px 30px;
                font-weight: bold;
            }}
        """)
        buttons_layout.addWidget(export_btn)

        layout.addLayout(buttons_layout)

        layout.addStretch()

        return widget

    def create_scheduled_tab(self):
        """Záložka: Plánované exporty"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Info
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #e3f2fd;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        info_layout = QVBoxLayout(info_frame)

        info_title = QLabel("💡 Plánované exporty")
        info_font = QFont()
        info_font.setBold(True)
        info_title.setFont(info_font)
        info_layout.addWidget(info_title)

        info_text = QLabel(
            "Nastavte automatické exporty, které se budou provádět v pravidelných intervalech.\n"
            "Exporty mohou být odesílány emailem nebo ukládány do složky."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)

        layout.addWidget(info_frame)

        # Tlačítka
        buttons_layout = QHBoxLayout()

        new_btn = QPushButton("➕ Nový plánovaný export")
        new_btn.clicked.connect(self.create_scheduled_export)
        new_btn.setStyleSheet(f"background-color: {config.COLOR_SUCCESS}; color: white; padding: 8px 15px;")
        buttons_layout.addWidget(new_btn)

        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)

        # Seznam plánovaných exportů
        self.scheduled_table = QTableWidget()
        self.scheduled_table.setColumnCount(6)
        self.scheduled_table.setHorizontalHeaderLabels([
            "Název",
            "Typ",
            "Frekvence",
            "Poslední export",
            "Aktivní",
            "Akce"
        ])
        self.scheduled_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.scheduled_table)

        self.load_scheduled_exports()

        layout.addStretch()

        return widget

    def create_templates_tab(self):
        """Záložka: Šablony exportů"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Info
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #e3f2fd;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        info_layout = QVBoxLayout(info_frame)

        info_title = QLabel("💡 Šablony exportů")
        info_font = QFont()
        info_font.setBold(True)
        info_title.setFont(info_font)
        info_layout.addWidget(info_title)

        info_text = QLabel(
            "Uložte si často používané konfigurace exportů jako šablony pro rychlý přístup."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)

        layout.addWidget(info_frame)

        # Tlačítka
        buttons_layout = QHBoxLayout()

        new_btn = QPushButton("➕ Nová šablona")
        new_btn.clicked.connect(self.create_export_template)
        new_btn.setStyleSheet(f"background-color: {config.COLOR_SUCCESS}; color: white; padding: 8px 15px;")
        buttons_layout.addWidget(new_btn)

        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)

        # Seznam šablon
        self.templates_table = QTableWidget()
        self.templates_table.setColumnCount(4)
        self.templates_table.setHorizontalHeaderLabels([
            "Název",
            "Typ",
            "Formát",
            "Akce"
        ])
        self.templates_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.templates_table)

        self.load_export_templates()

        layout.addStretch()

        return widget

    # =====================================================
    # EXPORT FUNKCÍ
    # =====================================================

    def export_invoices(self):
        """Export faktur"""
        try:
            # Získat parametry
            selection_type = self.invoice_selection_type.checkedId()
            invoice_type = self.invoice_type_filter.currentData()
            status = self.invoice_status_filter.currentData()
            format_type = self.invoice_format.checkedId()

            # Načíst faktury
            query = "SELECT * FROM invoices WHERE 1=1"
            params = []

            if selection_type == 1:  # Období
                query += " AND issue_date BETWEEN ? AND ?"
                params.extend([
                    self.invoice_date_from.date().toString("yyyy-MM-dd"),
                    self.invoice_date_to.date().toString("yyyy-MM-dd")
                ])

            if invoice_type != "all":
                query += " AND invoice_type = ?"
                params.append(invoice_type)

            if status != "all":
                query += " AND status = ?"
                params.append(status)

            query += " ORDER BY issue_date DESC"

            invoices = db.fetch_all(query, tuple(params) if params else None)

            if not invoices:
                QMessageBox.warning(self, "Upozornění", "Nebyly nalezeny žádné faktury odpovídající kritériím.")
                return

            # Vybrat cestu pro uložení
            if format_type == 0:  # PDF
                if hasattr(self, 'pdf_merged') and self.pdf_merged.isChecked():
                    file_path, _ = QFileDialog.getSaveFileName(
                        self,
                        "Uložit PDF",
                        f"faktury_{datetime.now().strftime('%Y%m%d')}.pdf",
                        "PDF soubory (*.pdf)"
                    )
                else:
                    file_path = QFileDialog.getExistingDirectory(
                        self,
                        "Vyberte složku pro PDF soubory"
                    )
            elif format_type == 1:  # Excel
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit Excel",
                    f"faktury_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "Excel soubory (*.xlsx)"
                )
            elif format_type == 2:  # CSV
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit CSV",
                    f"faktury_{datetime.now().strftime('%Y%m%d')}.csv",
                    "CSV soubory (*.csv)"
                )
            else:  # XML
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit XML",
                    f"faktury_{datetime.now().strftime('%Y%m%d')}.xml",
                    "XML soubory (*.xml)"
                )

            if not file_path:
                return

            # Provést export
            if format_type == 0:  # PDF
                self.export_invoices_pdf(invoices, file_path)
            elif format_type == 1:  # Excel
                self.export_invoices_excel(invoices, file_path)
            elif format_type == 2:  # CSV
                self.export_invoices_csv(invoices, file_path)
            else:  # XML
                self.export_invoices_xml(invoices, file_path)

            QMessageBox.information(
                self,
                "Úspěch",
                f"Export {len(invoices)} faktur byl dokončen.\n\n"
                f"Soubor: {file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se exportovat faktury:\n{e}")

    def export_payments(self):
        """Export plateb"""
        try:
            # Získat parametry
            date_from = self.payment_date_from.date().toString("yyyy-MM-dd")
            date_to = self.payment_date_to.date().toString("yyyy-MM-dd")
            payment_type = self.payment_type_filter.currentData()
            format_type = self.payment_format.checkedId()

            # Načíst platby
            query = """
                SELECT
                    p.*,
                    i.invoice_number,
                    c.first_name || ' ' || c.last_name as customer_name
                FROM payments p
                LEFT JOIN invoices i ON p.invoice_id = i.id
                LEFT JOIN customers c ON i.customer_id = c.id
                WHERE p.payment_date BETWEEN ? AND ?
            """
            params = [date_from, date_to]

            if payment_type != "all":
                if payment_type == "income":
                    query += " AND p.amount > 0"
                else:
                    query += " AND p.amount < 0"

            query += " ORDER BY p.payment_date DESC"

            payments = db.fetch_all(query, tuple(params))

            if not payments:
                QMessageBox.warning(self, "Upozornění", "Nebyly nalezeny žádné platby.")
                return

            # Vybrat cestu pro uložení
            if format_type == 0:  # Excel
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit Excel",
                    f"platby_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "Excel soubory (*.xlsx)"
                )
            elif format_type == 1:  # CSV
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit CSV",
                    f"platby_{datetime.now().strftime('%Y%m%d')}.csv",
                    "CSV soubory (*.csv)"
                )
            else:  # PDF
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit PDF",
                    f"platby_{datetime.now().strftime('%Y%m%d')}.pdf",
                    "PDF soubory (*.pdf)"
                )

            if not file_path:
                return

            # Provést export
            if format_type == 0:  # Excel
                self.export_payments_excel(payments, file_path)
            elif format_type == 1:  # CSV
                self.export_payments_csv(payments, file_path)
            else:  # PDF
                self.export_payments_pdf(payments, file_path)

            QMessageBox.information(
                self,
                "Úspěch",
                f"Export {len(payments)} plateb byl dokončen.\n\n"
                f"Soubor: {file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se exportovat platby:\n{e}")

    def export_accounting(self):
        """Export účetních dat"""
        try:
            # Získat parametry
            date_from = self.accounting_date_from.date().toString("yyyy-MM-dd")
            date_to = self.accounting_date_to.date().toString("yyyy-MM-dd")
            software = self.accounting_software.currentData()

            # Vybrat cestu pro uložení
            if software == "excel":
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit Excel",
                    f"ucetni_data_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "Excel soubory (*.xlsx)"
                )
            elif software == "csv":
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit CSV",
                    f"ucetni_data_{datetime.now().strftime('%Y%m%d')}.csv",
                    "CSV soubory (*.csv)"
                )
            else:  # XML
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Uložit XML",
                    f"ucetni_data_{datetime.now().strftime('%Y%m%d')}.xml",
                    "XML soubory (*.xml)"
                )

            if not file_path:
                return

            # Načíst data
            data = {
                "invoices": [],
                "payments": [],
                "vat_summary": {}
            }

            if self.accounting_include_invoices.isChecked():
                query = """
                    SELECT * FROM invoices
                    WHERE issue_date BETWEEN ? AND ?
                    ORDER BY issue_date
                """
                data["invoices"] = db.fetch_all(query, (date_from, date_to))

            if self.accounting_include_payments.isChecked():
                query = """
                    SELECT * FROM payments
                    WHERE payment_date BETWEEN ? AND ?
                    ORDER BY payment_date
                """
                data["payments"] = db.fetch_all(query, (date_from, date_to))

            if self.accounting_include_vat.isChecked():
                # Výpočet DPH
                data["vat_summary"] = self.calculate_vat_summary(date_from, date_to)

            # Provést export
            if software == "excel":
                self.export_accounting_excel(data, file_path)
            elif software == "csv":
                self.export_accounting_csv(data, file_path)
            elif software == "pohoda":
                self.export_accounting_pohoda(data, file_path)
            elif software == "money":
                self.export_accounting_money(data, file_path)
            elif software == "abra":
                self.export_accounting_abra(data, file_path)
            else:
                self.export_accounting_xml(data, file_path)

            QMessageBox.information(
                self,
                "Úspěch",
                f"Export účetních dat byl dokončen.\n\n"
                f"Soubor: {file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Chyba", f"Nepodařilo se exportovat účetní data:\n{e}")

    # =====================================================
    # EXPORT METODY
    # =====================================================

    def export_invoices_pdf(self, invoices, file_path):
        """Export faktur do PDF"""
        # TODO: Implementovat generování PDF faktur
        pass

    def export_invoices_excel(self, invoices, file_path):
        """Export faktur do Excelu"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Faktury"

            # Hlavička
            headers = [
                "Číslo faktury", "Typ", "Zákazník", "Datum vystavení",
                "Datum splatnosti", "Částka bez DPH", "DPH", "Částka s DPH",
                "Zaplaceno", "Zbývá", "Status"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row, invoice in enumerate(invoices, 2):
                ws.cell(row=row, column=1, value=invoice["invoice_number"])
                ws.cell(row=row, column=2, value="Vydaná" if invoice["invoice_type"] == "issued" else "Přijatá")

                # Načíst zákazníka
                customer = db.fetch_one("SELECT first_name, last_name, company FROM customers WHERE id = ?",
                                       (invoice["customer_id"],))
                customer_name = customer["company"] if customer and customer["company"] else f"{customer['first_name']} {customer['last_name']}" if customer else ""
                ws.cell(row=row, column=3, value=customer_name)

                ws.cell(row=row, column=4, value=datetime.fromisoformat(invoice["issue_date"]).strftime("%d.%m.%Y"))
                ws.cell(row=row, column=5, value=datetime.fromisoformat(invoice["due_date"]).strftime("%d.%m.%Y"))
                ws.cell(row=row, column=6, value=invoice["total_without_vat"])
                ws.cell(row=row, column=7, value=invoice["total_vat"])
                ws.cell(row=row, column=8, value=invoice["total_with_vat"])
                ws.cell(row=row, column=9, value=invoice["paid_amount"])
                ws.cell(row=row, column=10, value=invoice["total_with_vat"] - invoice["paid_amount"])

                status_map = {
                    "paid": "Zaplaceno",
                    "unpaid": "Nezaplaceno",
                    "partial": "Částečně",
                    "overdue": "Po splatnosti"
                }
                ws.cell(row=row, column=11, value=status_map.get(invoice["status"], invoice["status"]))

            # Automatická šířka sloupců
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)

        except ImportError:
            QMessageBox.critical(
                self,
                "Chyba",
                "Pro export do Excelu je potřeba nainstalovat balíček 'openpyxl'.\n\n"
                "Instalace: pip install openpyxl"
            )

    def export_invoices_csv(self, invoices, file_path):
        """Export faktur do CSV"""
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')

                # Hlavička
                writer.writerow([
                    "Číslo faktury", "Typ", "Zákazník", "Datum vystavení",
                    "Datum splatnosti", "Částka bez DPH", "DPH", "Částka s DPH",
                    "Zaplaceno", "Zbývá", "Status"
                ])

                # Data
                for invoice in invoices:
                    # Načíst zákazníka
                    customer = db.fetch_one("SELECT first_name, last_name, company FROM customers WHERE id = ?",
                                           (invoice["customer_id"],))
                    customer_name = customer["company"] if customer and customer["company"] else f"{customer['first_name']} {customer['last_name']}" if customer else ""

                    writer.writerow([
                        invoice["invoice_number"],
                        "Vydaná" if invoice["invoice_type"] == "issued" else "Přijatá",
                        customer_name,
                        datetime.fromisoformat(invoice["issue_date"]).strftime("%d.%m.%Y"),
                        datetime.fromisoformat(invoice["due_date"]).strftime("%d.%m.%Y"),
                        f"{invoice['total_without_vat']:.2f}",
                        f"{invoice['total_vat']:.2f}",
                        f"{invoice['total_with_vat']:.2f}",
                        f"{invoice['paid_amount']:.2f}",
                        f"{invoice['total_with_vat'] - invoice['paid_amount']:.2f}",
                        invoice["status"]
                    ])

        except Exception as e:
            raise Exception(f"Chyba při exportu CSV: {e}")

    def export_invoices_xml(self, invoices, file_path):
        """Export faktur do XML"""
        # TODO: Implementovat XML export
        pass

    def export_payments_excel(self, payments, file_path):
        """Export plateb do Excelu"""
        # TODO: Implementovat podobně jako export_invoices_excel
        pass

    def export_payments_csv(self, payments, file_path):
        """Export plateb do CSV"""
        # TODO: Implementovat podobně jako export_invoices_csv
        pass

    def export_payments_pdf(self, payments, file_path):
        """Export plateb do PDF"""
        # TODO: Implementovat generování PDF přehledu plateb
        pass

    def export_accounting_excel(self, data, file_path):
        """Export účetních dat do Excelu (více listů)"""
        # TODO: Implementovat export s více listy (faktury, platby, DPH)
        pass

    def export_accounting_csv(self, data, file_path):
        """Export účetních dat do CSV"""
        # TODO: Implementovat
        pass

    def export_accounting_xml(self, data, file_path):
        """Export účetních dat do XML"""
        # TODO: Implementovat
        pass

    def export_accounting_pohoda(self, data, file_path):
        """Export pro Pohodu"""
        # TODO: Implementovat XML formát pro Pohodu
        pass

    def export_accounting_money(self, data, file_path):
        """Export pro Money S3"""
        # TODO: Implementovat XML formát pro Money S3
        pass

    def export_accounting_abra(self, data, file_path):
        """Export pro ABRA Gen"""
        # TODO: Implementovat CSV formát pro ABRA Gen
        pass

    # =====================================================
    # POMOCNÉ METODY
    # =====================================================

    def calculate_vat_summary(self, date_from, date_to):
        """Výpočet souhrnu DPH"""
        try:
            # DPH na výstupu (vydané faktury)
            query_output = """
                SELECT
                    SUM(total_without_vat) as base,
                    SUM(total_vat) as vat
                FROM invoices
                WHERE invoice_type = 'issued'
                  AND issue_date BETWEEN ? AND ?
            """
            output = db.fetch_one(query_output, (date_from, date_to))

            # DPH na vstupu (přijaté faktury)
            query_input = """
                SELECT
                    SUM(total_without_vat) as base,
                    SUM(total_vat) as vat
                FROM invoices
                WHERE invoice_type = 'received'
                  AND issue_date BETWEEN ? AND ?
            """
            input_data = db.fetch_one(query_input, (date_from, date_to))

            return {
                "output_base": output["base"] or 0,
                "output_vat": output["vat"] or 0,
                "input_base": input_data["base"] or 0,
                "input_vat": input_data["vat"] or 0,
                "vat_to_pay": (output["vat"] or 0) - (input_data["vat"] or 0)
            }

        except Exception as e:
            print(f"Chyba při výpočtu DPH: {e}")
            return {}

    def preview_export(self, export_type):
        """Náhled exportu"""
        QMessageBox.information(
            self,
            "Náhled",
            f"Funkce náhledu exportu ({export_type}) bude implementována.\n\n"
            "Zobrazí se tabulka s daty, která budou exportována."
        )

    def load_scheduled_exports(self):
        """Načtení plánovaných exportů"""
        # TODO: Načíst z databáze
        self.scheduled_table.setRowCount(0)

    def load_export_templates(self):
        """Načtení šablon exportů"""
        # TODO: Načíst z databáze
        self.templates_table.setRowCount(0)

    def create_scheduled_export(self):
        """Vytvoření plánovaného exportu"""
        dialog = ScheduledExportDialog(self)
        dialog.exec()

    def create_export_template(self):
        """Vytvoření šablony exportu"""
        dialog = ExportTemplateDialog(self)
        dialog.exec()

    def refresh(self):
        """Obnovení"""
        self.load_scheduled_exports()
        self.load_export_templates()


# =====================================================
# DIALOGY
# =====================================================

class ScheduledExportDialog(QDialog):
    """Dialog pro vytvoření plánovaného exportu"""

    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Nový plánovaný export")
        self.setMinimumWidth(500)

        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        layout = QFormLayout(self)

        # Název
        self.name_input = QLineEdit()
        layout.addRow("Název:", self.name_input)

        # Typ exportu
        self.type_combo = QComboBox()
        self.type_combo.addItem("Faktury", "invoices")
        self.type_combo.addItem("Platby", "payments")
        self.type_combo.addItem("Účetní data", "accounting")
        layout.addRow("Typ exportu:", self.type_combo)

        # Frekvence
        self.frequency_combo = QComboBox()
        self.frequency_combo.addItem("Denně", "daily")
        self.frequency_combo.addItem("Týdně", "weekly")
        self.frequency_combo.addItem("Měsíčně", "monthly")
        layout.addRow("Frekvence:", self.frequency_combo)

        # Email
        self.email_input = QLineEdit()
        layout.addRow("Odeslat na email:", self.email_input)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Zrušit")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        save_btn = QPushButton("💾 Uložit")
        save_btn.clicked.connect(self.accept)
        save_btn.setStyleSheet(f"background-color: {config.COLOR_SUCCESS}; color: white; padding: 8px 20px;")
        buttons_layout.addWidget(save_btn)

        layout.addRow(buttons_layout)


class ExportTemplateDialog(QDialog):
    """Dialog pro vytvoření šablony exportu"""

    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Nová šablona exportu")
        self.setMinimumWidth(500)

        self.init_ui()

    def init_ui(self):
        """Inicializace UI"""
        layout = QFormLayout(self)

        # Název
        self.name_input = QLineEdit()
        layout.addRow("Název šablony:", self.name_input)

        # Typ
        self.type_combo = QComboBox()
        self.type_combo.addItem("Faktury", "invoices")
        self.type_combo.addItem("Platby", "payments")
        self.type_combo.addItem("Účetní data", "accounting")
        layout.addRow("Typ:", self.type_combo)

        # Formát
        self.format_combo = QComboBox()
        self.format_combo.addItem("PDF", "pdf")
        self.format_combo.addItem("Excel", "excel")
        self.format_combo.addItem("CSV", "csv")
        self.format_combo.addItem("XML", "xml")
        layout.addRow("Formát:", self.format_combo)

        # Tlačítka
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("Zrušit")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        save_btn = QPushButton("💾 Uložit")
        save_btn.clicked.connect(self.accept)
        save_btn.setStyleSheet(f"background-color: {config.COLOR_SUCCESS}; color: white; padding: 8px 20px;")
        buttons_layout.addWidget(save_btn)

        layout.addRow(buttons_layout)
